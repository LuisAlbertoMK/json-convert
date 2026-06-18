"""
test_integration.py — Tests de integración para json-convert.

Evalúa el pipeline completo sin navegador real:
  - Excel I/O con openpyxl real
  - setup_multisheet + update_control + update_vars_sheet
  - validate_sheet + validate_url + clasificación de errores
  - save_workbook con fallback
  - Config loading desde audit.json

Requiere: openpyxl (no requiere playwright ni navegador)
"""

import asyncio
import json
import os
import sys
import tempfile
import unittest
from unittest.mock import patch

# Importar funciones del script principal
sys.path.insert(0, os.path.dirname(__file__))
from extract_browser import (
    CONTROL_HEADERS,
    SHEET_HEADERS,
    _error_code_from_detail,
    classify_errors,
    compute_score,
    save_workbook,
    setup_multisheet,
    update_control,
    validate_sheet,
    validate_url,
)

# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _make_urls_json(path: str, entries: list | None = None):
    """Crea un archivo urls.json de prueba."""
    if entries is None:
        entries = [
            {"nombre": "Ford Home", "url": "https://www.ford.com"},
            "https://brandpr.ford.com/preview/mach-e",
        ]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(entries, f)
    return path


# ═══════════════════════════════════════════════════════════════════════════
# TESTS: setup_multisheet + Excel I/O pipeline
# ═══════════════════════════════════════════════════════════════════════════

class TestMultisheetPipeline(unittest.TestCase):
    """Pipeline completo: urls.json → setup_multisheet → validación → guardado."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.urls_path = os.path.join(self.tmpdir, "urls.json")
        self.output_path = os.path.join(self.tmpdir, "historial.xlsx")
        _make_urls_json(self.urls_path)

    def tearDown(self):
        for f in os.listdir(self.tmpdir):
            try:
                os.remove(os.path.join(self.tmpdir, f))
            except PermissionError:
                pass
        os.rmdir(self.tmpdir)

    def test_setup_multisheet_creates_new(self):
        """setup_multisheet crea workbook desde 0."""
        wb, ws, audit_date, skipped = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        self.assertIsNotNone(ws)
        self.assertFalse(skipped)
        self.assertIn("_control", wb.sheetnames)
        self.assertIn(audit_date, wb.sheetnames)

        # Verificar headers (nuevo layout: 7 columnas)
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        self.assertEqual(headers, SHEET_HEADERS)

        # setup_multisheet crea estructura + headers.
        # Las URLs se agregan en amain() después del setup.
        # Verificar que hay header row + filas vacías listas.
        self.assertGreaterEqual(ws.max_row, 1)

        # Verificar _control tiene headers
        ctrl = wb["_control"]
        ctrl_headers = [ctrl.cell(1, c).value for c in range(1, 11)]
        self.assertEqual(ctrl_headers, CONTROL_HEADERS)

        wb.close()

    def test_setup_with_url_data(self):
        """Simula el flujo real: setup + carga de URLs desde JSON."""
        wb, ws, audit_date, skipped = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        # Cargar URLs igual que amain()
        with open(self.urls_path, encoding="utf-8") as f:
            entries = json.load(f)
        from openpyxl.styles import Alignment
        wrap = Alignment(wrap_text=True, vertical="top")
        for i, entry in enumerate(entries, start=2):
            if isinstance(entry, str):
                name = url = entry
            elif isinstance(entry, dict):
                name = entry.get("nombre", entry.get("url", ""))
                url = entry.get("url", "")
            ws.cell(i, 1).value = name
            ws.cell(i, 1).alignment = wrap
            ws.cell(i, 2).value = url
            ws.cell(i, 2).alignment = wrap

        url_rows = [(ws.cell(r, 1).value, ws.cell(r, 2).value)
                     for r in range(2, ws.max_row + 1)]
        self.assertEqual(len(url_rows), 2)
        self.assertEqual(url_rows[0], ("Ford Home", "https://www.ford.com"))
        self.assertEqual(url_rows[1], ("https://brandpr.ford.com/preview/mach-e",
                                        "https://brandpr.ford.com/preview/mach-e"))
        wb.close()

    def test_setup_multisheet_resume_skips(self):
        """--resume con sheet existente → skipped=True."""
        # Primera corrida: crear + guardar
        wb1, ws1, date1, skipped1 = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        save_workbook(wb1, self.output_path)
        wb1.close()

        # Segunda corrida con --resume
        wb2, ws2, date2, skipped2 = setup_multisheet(
            self.output_path, self.urls_path, resume=True
        )
        self.assertTrue(skipped2)
        self.assertEqual(date1, date2)
        wb2.close()

    def test_setup_multisheet_replaces_existing(self):
        """Sin --resume, reemplaza sheet existente."""
        wb1, ws1, _, _ = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        wb1.close()

        # Segunda corrida sin resume → reemplaza
        wb2, ws2, _, skipped2 = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        self.assertFalse(skipped2)
        self.assertIsNotNone(ws2)
        wb2.close()

    def test_validate_sheet_ok(self):
        """validate_sheet con sheet válido + URLs → 0 errores."""
        wb, ws, _, _ = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        # Agregar URLs como hace amain()
        ws.cell(2, 1).value = "Ford Home"
        ws.cell(2, 2).value = "https://www.ford.com"
        ws.cell(3, 1).value = "Mach-E"
        ws.cell(3, 2).value = "https://brandpr.ford.com/preview/mach-e"
        errs = validate_sheet(ws)
        self.assertEqual(errs, [])
        wb.close()

    def test_validate_sheet_missing_urls(self):
        """validate_sheet sin URLs en col B → error."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 2).value = "pagina auditada (URL)"
        # Sin datos en col B
        errs = validate_sheet(ws)
        self.assertTrue(any("No hay URLs" in e for e in errs))
        wb.close()

    def test_update_control(self):
        """update_control agrega fila correctamente."""
        wb, ws, audit_date, _ = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        update_control(wb, audit_date, "urls.json",
                        total=5, ok_aa=4, ok_dd=3, errors=1, retries=2,
                        score=85, elapsed_s=120.5, workers=3)

        ctrl = wb["_control"]
        self.assertEqual(ctrl.max_row, 2)  # header + 1 data row
        row = [ctrl.cell(2, c).value for c in range(1, 11)]
        self.assertEqual(row[0], audit_date)
        self.assertEqual(row[1], "urls.json")
        self.assertEqual(row[2], 5)
        self.assertEqual(row[3], 4)
        self.assertEqual(row[7], 85)
        wb.close()

    def test_save_and_reload(self):
        """Guarda y recarga: verifica integridad."""
        wb, ws, audit_date, _ = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        ws.cell(2, 4).value = '{"page": "test"}'
        save_workbook(wb, self.output_path)
        wb.close()

        # Recargar
        import openpyxl
        wb2 = openpyxl.load_workbook(self.output_path)
        self.assertIn(audit_date, wb2.sheetnames)
        self.assertIn("_control", wb2.sheetnames)
        ws2 = wb2[audit_date]
        self.assertEqual(ws2.cell(2, 4).value, '{"page": "test"}')
        wb2.close()

    def test_save_workbook_fallback(self):
        """save_workbook con archivo bloqueado → fallback."""
        wb, ws, _, _ = setup_multisheet(
            self.output_path, self.urls_path, resume=False
        )
        # Mock: primer save falla, segundo save (fallback) funciona
        real_save = wb.save
        call_count = [0]

        def mock_save(path):
            call_count[0] += 1
            if call_count[0] == 1:
                raise PermissionError("Archivo bloqueado")
            # segundo intento (fallback) → ejecuta real
            real_save(path)

        with patch.object(wb, "save", side_effect=mock_save):
            result = save_workbook(wb, self.output_path)
        self.assertNotEqual(result, self.output_path)
        self.assertTrue(result.endswith("_browser.xlsx"))
        wb.close()


# ═══════════════════════════════════════════════════════════════════════════
# TESTS: Config loading
# ═══════════════════════════════════════════════════════════════════════════

class TestConfigLoading(unittest.TestCase):
    """Tests de carga de configuración desde JSON."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.config_path = os.path.join(self.tmpdir, "audit.json")
        self.config = {
            "workers": 5,
            "retry": 3,
            "timeout": 45000,
            "discard_cookies": True,
            "progress": True,
            "headed": True,
            "verbose": True,
        }
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(self.config, f)

    def tearDown(self):
        os.remove(self.config_path)
        os.rmdir(self.tmpdir)

    def test_parse_config(self):
        """Simula la carga de config del amain()."""
        config_file = self.config_path
        if not os.path.exists(config_file):
            self.fail("Config file not created")

        with open(config_file, encoding="utf-8") as f:
            cfg = json.load(f)

        # Simular argparse defaults exactos (None/False para no-especificados)
        class Args:
            pass

        args = Args()
        # Asignar defaults como argparse: None o False
        for k in cfg:
            if not hasattr(args, k):
                setattr(args, k, None if k not in ("headed", "discard_cookies", "progress", "verbose") else False)

        # Aplicar config: solo sobreescribe valores falsy/None
        for k, v in cfg.items():
            if not getattr(args, k, None):
                setattr(args, k, v)

        self.assertEqual(args.workers, 5)
        self.assertEqual(args.retry, 3)
        self.assertEqual(args.timeout, 45000)
        self.assertTrue(args.discard_cookies)
        self.assertTrue(args.progress)
        self.assertTrue(args.headed)
        self.assertTrue(args.verbose)

    def test_config_arg_priority(self):
        """Flag CLI (truthy) debe tener prioridad sobre config."""
        with open(self.config_path, encoding="utf-8") as f:
            cfg = json.load(f)

        class Args:
            pass

        args = Args()
        # Simular flags CLI explícitos (valores truthy → no sobreescribir)
        args.workers = 1     # explícito → NO se sobreescribe
        args.timeout = 35000 # explícito → NO se sobreescribe
        # No seteamos retry → se toma del config

        for k, v in cfg.items():
            if not getattr(args, k, None):
                setattr(args, k, v)

        # workers=1 CLI explícito → se mantiene
        self.assertEqual(args.workers, 1)
        # timeout explícito → se mantiene
        self.assertEqual(args.timeout, 35000)
        # retry no especificado en CLI → toma config
        self.assertEqual(args.retry, 3)


# ═══════════════════════════════════════════════════════════════════════════
# TESTS: validate_url + error classification integration
# ═══════════════════════════════════════════════════════════════════════════

class TestValidationPipeline(unittest.TestCase):
    """Pipeline de validación: validate_url → error code → classify."""

    def test_url_to_classify_flow(self):
        """URL inválida → código → clasificación."""
        test_cases = [
            ("", "URL vacía o inválida", "UNKNOWN", None),
            ("ftp://ford.com/file", "Scheme", "NETWORK_ERROR", None),
            ("http://localhost:8080", "SSRF", "NETWORK_ERROR", None),
            ("https://evil.com/phish", "whitelist", "NETWORK_ERROR", None),
        ]

        for url, expected_msg, expected_code, _ in test_cases:
            err = validate_url(url)
            self.assertIsNotNone(err, f"URL '{url}' debería fallar")
            if expected_msg:
                self.assertIn(expected_msg, err or "", f"URL '{url}' debería contener '{expected_msg}'")

            # Simular el mismo flujo que amain()
            code = _error_code_from_detail(err or "")
            self.assertIsInstance(code, str)

            # Clasificar
            cats = classify_errors([{"row": 2, "code": code, "error": err}])
            # URL_INVALID se omite en classify (ya reportado)
            if code == "URL_INVALID":
                self.assertNotIn("URL_INVALID", str(cats))
            else:
                self.assertTrue(len(cats) > 0, f"URL '{url}' debería tener categoría")

    def test_valid_url_no_classify(self):
        """URL válida → sin error → sin clasificación."""
        urls = [
            "https://www.ford.com/es/mustang",
            "https://brandpr.ford.com/preview/mach-e",
        ]
        for url in urls:
            err = validate_url(url)
            self.assertIsNone(err, f"URL '{url}' debería ser válida")

    def test_mixed_urls_batch(self):
        """Lote mixto de URLs: algunas válidas, otras no."""
        urls = [
            (3, "https://www.ford.com/home", None),
            (5, "http://localhost:8080", "SSRF"),
            (7, "https://brandpr.ford.com/preview", None),
            (9, "https://evil.com", "whitelist"),
        ]

        errors_detail = []
        valid_count = 0
        for row, url, expected_err in urls:
            err = validate_url(url)
            if err:
                code = _error_code_from_detail(err)
                errors_detail.append({"row": row, "code": code, "error": err})
            else:
                valid_count += 1

        self.assertEqual(valid_count, 2)
        self.assertEqual(len(errors_detail), 2)
        categories = classify_errors(errors_detail)
        # Al menos una categoría de error de red
        self.assertTrue(len(categories) > 0)


# ═══════════════════════════════════════════════════════════════════════════
# TESTS: write_result (Excel lock es el integrador)
# ═══════════════════════════════════════════════════════════════════════════

class TestWriteResultIntegration(unittest.TestCase):
    """write_result con Excel real + metrics dict."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.output_path = os.path.join(self.tmpdir, "output.xlsx")
        # Crear workbook con datos
        import openpyxl
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        for c, h in enumerate(SHEET_HEADERS, start=1):
            self.ws.cell(1, c).value = h
        self.ws.cell(2, 1).value = "Ford Home"
        self.ws.cell(2, 2).value = "https://www.ford.com/home"

    def tearDown(self):
        self.wb.close()
        for f in os.listdir(self.tmpdir):
            try:
                os.remove(os.path.join(self.tmpdir, f))
            except PermissionError:
                pass
        os.rmdir(self.tmpdir)

    def _run_write_result(self, result: dict) -> dict:
        """Ejecuta write_result con mocks async."""
        metrics = {
            "total": 1, "ok_aa": 0, "ok_dd": 0,
            "errors": 0, "retries": 0, "total_beacons": 0,
            "times": [], "errores_detalle": [],
        }
        lock = asyncio.Lock()
        saved_count = [0]

        async def run():
            from extract_browser import write_result
            await write_result(
                self.ws, result, metrics, lock,
                self.output_path, saved_count,
                show_progress=False, total_urls=1, workers=1,
            )

        asyncio.run(run())
        return metrics

    def test_write_successful_result(self):
        """Resultado exitoso escribe datos en celdas correctas."""
        result = {
            "row": 2, "url": "https://www.ford.com/home",
            "status": 200, "error": None, "code": None,
            "title": "Ford Home Page", "digitaldata": {"page": "home"},
            "digitaldata_auto": {"page": "home"},
            "aa_parsed": {"events": ["event1"]},
            "aa_source": "beacon", "extra_beacons": [{"events": ["event2"]}],
            "elapsed_s": 3.5, "retries_used": 0,
        }
        metrics = self._run_write_result(result)

        # Verificar metrics actualizados
        self.assertEqual(metrics["ok_aa"], 1)
        self.assertEqual(metrics["ok_dd"], 1)
        self.assertEqual(metrics["errors"], 0)
        self.assertEqual(metrics["total_beacons"], 2)  # 1 principal + 1 extra

        # Verificar celdas — nuevo layout:
        #   C (3) = digitaldata manual, D (4) = digitaldata auto
        #   E (5) = AA analytics, G (7) = metadata
        dd_manual = self.ws.cell(2, 3).value  # col C = digitaldata (manual/primary)
        self.assertIn("home", dd_manual or "")
        dd_auto = self.ws.cell(2, 4).value  # col D = digitaldata (automatica)
        self.assertIn("home", dd_auto or "")
        aa_val = self.ws.cell(2, 5).value  # col E = AA analytics
        self.assertIn("event1", aa_val or "")
        meta_val = self.ws.cell(2, 7).value  # col G = metadata
        self.assertIn("score", meta_val or "")
        self.assertIn("beacon", meta_val or "")

    def test_write_error_result(self):
        """Resultado con error → cols de AA/dd muestran error, se registra en metrics."""
        result = {
            "row": 2, "url": "https://www.ford.com/home",
            "status": -1, "error": "timeout", "code": "TIMEOUT",
            "title": "", "digitaldata": None,
            "digitaldata_auto": None,
            "aa_parsed": None,
            "aa_source": None, "extra_beacons": [],
            "elapsed_s": 35.0, "retries_used": 2,
        }
        metrics = self._run_write_result(result)

        self.assertEqual(metrics["ok_aa"], 0)
        self.assertEqual(metrics["ok_dd"], 0)
        self.assertEqual(metrics["errors"], 1)
        self.assertEqual(metrics["total_beacons"], 1)  # 1 default
        self.assertEqual(len(metrics["errores_detalle"]), 1)

        # Verificar col E = error en AA
        aa_val = self.ws.cell(2, 5).value
        self.assertIn("timeout", aa_val or "")
        self.assertIn("TIMEOUT", aa_val or "")

        # Verificar col C (manual) y D (auto) = error en JSON
        dd_manual = self.ws.cell(2, 3).value
        self.assertIn("no digitaldata", dd_manual or "")
        dd_auto = self.ws.cell(2, 4).value
        self.assertIn("no digitaldata", dd_auto or "")

        # Verificar metadata incluye el error (col G)
        meta_val = self.ws.cell(2, 7).value
        self.assertIn("TIMEOUT", meta_val or "")

    def test_write_partial_result(self):
        """Solo digitaldata (sin AA) → métricas parciales."""
        result = {
            "row": 2, "url": "https://www.ford.com/home",
            "status": 200, "error": "no AA data captured", "code": "NO_AA_DATA",
            "title": "Ford Page", "digitaldata": {"page": "home"},
            "digitaldata_auto": {"page": "home"},
            "aa_parsed": None,
            "aa_source": None, "extra_beacons": [],
            "elapsed_s": 8.0, "retries_used": 0,
        }
        metrics = self._run_write_result(result)

        self.assertEqual(metrics["ok_dd"], 1)
        self.assertEqual(metrics["ok_aa"], 0)
        self.assertEqual(metrics["errors"], 1)  # por aa_parsed=None

        # col C (manual) = digitaldata
        dd_manual = self.ws.cell(2, 3).value
        self.assertIn("home", dd_manual or "")

        # col D (auto) = digitaldata
        dd_auto = self.ws.cell(2, 4).value
        self.assertIn("home", dd_auto or "")

        # col E = error en JSON (no "parentizado")
        aa_val = self.ws.cell(2, 5).value
        self.assertIn("no AA data captured", aa_val or "")
        self.assertIn("NO_AA_DATA", aa_val or "")


# ═══════════════════════════════════════════════════════════════════════════
# TESTS: compute_score integrado con metrics reales
# ═══════════════════════════════════════════════════════════════════════════

class TestScoreIntegration(unittest.TestCase):
    """compute_score con métricas generadas por write_result."""

    def test_score_from_write_results(self):
        """Simula múltiples write_result y verifica score final."""
        metrics = {
            "total": 0, "ok_aa": 0, "ok_dd": 0,
            "errors": 0, "retries": 0, "total_beacons": 0,
            "times": [], "errores_detalle": [],
        }

        # Simular 3 URLs exitosas
        for i in range(3):
            metrics["total"] += 1
            metrics["ok_aa"] += 1
            metrics["ok_dd"] += 1
            metrics["total_beacons"] += 2
            metrics["times"].append(4.0)

        score = compute_score(metrics)
        self.assertGreater(score, 80)
        self.assertLessEqual(score, 100)

    def test_score_degrades_with_errors(self):
        """Más errores → score más bajo."""
        good = {
            "total": 10, "ok_aa": 9, "ok_dd": 8,
            "errors": 1, "retries": 0, "total_beacons": 25,
            "times": [5]*10, "errores_detalle": [],
        }
        bad = {
            "total": 10, "ok_aa": 3, "ok_dd": 5,
            "errors": 7, "retries": 5, "total_beacons": 5,
            "times": [30]*10, "errores_detalle": [],
        }
        self.assertGreater(compute_score(good), compute_score(bad))


# ═══════════════════════════════════════════════════════════════════════════
# TESTS: Error classification pipeline
# ═══════════════════════════════════════════════════════════════════════════

class TestClassifyPipeline(unittest.TestCase):
    """classify_errors con casos reales compuestos."""

    def test_classify_mixed_errors_no_duplicates(self):
        """Múltiples errores del mismo tipo → una categoría."""
        errors = [
            {"row": 2, "code": "TIMEOUT"},
            {"row": 5, "code": "TIMEOUT"},
            {"row": 8, "code": "NETWORK_ERROR"},
            {"row": 11, "code": "NETWORK_ERROR"},
            {"row": 14, "code": "HTTP_403", "error": "403 Forbidden"},
        ]
        cats = classify_errors(errors)
        self.assertEqual(cats.get("Timeout"), [2, 5])
        self.assertEqual(cats.get("Error de red/conexión"), [8, 11])
        self.assertEqual(cats.get("HTTP 403 (acceso denegado)"), [14])

    def test_classify_empty_and_missing_codes(self):
        """Errores sin código explícito se deducen del texto (como en el flujo real)."""
        # En el flujo real, write_result NO incluye "code" key en errores_detalle
        errors = [
            {"row": 3, "error": "timeout after 30s"},
            {"row": 7, "error": "connection refused"},
            {"row": 9, "error": "No AA beacon found"},
        ]
        cats = classify_errors(errors)
        self.assertIn("Timeout", cats)
        self.assertIn("Error de red/conexión", cats)
        self.assertIn("Sin dato AA (no beacon)", cats)


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    unittest.main(verbosity=2)

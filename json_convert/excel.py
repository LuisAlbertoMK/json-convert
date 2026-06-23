"""
json_convert/excel.py — Helpers de Excel y escritura de resultados.

Contiene toda la lógica de generación de reports Excel:
formatos, sheets multi-fecha, control sheets, split AA,
colores condicionales y guardado con fallback.
"""

from __future__ import annotations

import json
import logging
import os
import time as _time
from copy import copy
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, PatternFill

# ── Constantes ──
INPUT_FILE = "RevisionManual.xlsx"
SAVE_EVERY_N = 20

SHEET_HEADERS = [
    "nombre pagina auditada",
    "pagina auditada (URL)",
    "digitaldata (manual)",
    "digitaldata (automatica)",
    "AA analytics (manual)",
    "AA analytics (automatico)",
    "AA analytics (estructurado)",
    "metadata / extra beacons",
]

CONTROL_HEADERS = [
    "Fecha", "Source", "URLs", "AA OK", "DD OK", "Errores",
    "Reintentos", "Score", "Tiempo (s)", "Workers",
]

HEADER_FILLS = {
    "A": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "B": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "C": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),  # manual DD — azul claro
    "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # auto DD — rojo claro
    "E": PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid"),  # manual AA — lavanda
    "F": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # AA auto — amarillo
    "G": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # AA struct — verde
}

DATA_FILLS = {
    "G": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}


# ═══════════════════════════════════════════════════════════════════════════
# FORMATO
# ═══════════════════════════════════════════════════════════════════════════

def _pretty_json(obj: object) -> str:
    try:
        return json.dumps(obj, indent=2, ensure_ascii=False)
    except (ValueError, TypeError, OverflowError):
        # Fallback: safe serialize con str() para no-perder datos
        return json.dumps(_safe_serialize(obj), indent=2, ensure_ascii=False)


def _safe_serialize(obj: object, depth: int = 0) -> object:
    """Convierte objetos con potenciales ciclos a tipos JSON-safe."""
    if depth > 20:
        _log_truncation("max depth reached", str(obj)[:60])
        return str(obj)[:200]
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            try:
                out[str(k)] = _safe_serialize(v, depth + 1)
            except Exception:
                logging.debug("_safe_serialize fallback for key %s", str(k)[:60])
                _log_truncation(f"key {str(k)[:60]}", str(v)[:60])
                out[str(k)] = str(v)[:200]
        return out
    if isinstance(obj, (list, tuple)):
        truncated = [_safe_serialize(item, depth + 1) for item in obj[:50]]
        if len(obj) > 50:
            _log_truncation("list truncated", f"{len(obj)} items -> 50")
        return truncated
    if isinstance(obj, (str, int, float, bool)) or obj is None:
        return obj
    _log_truncation(f"type {type(obj).__name__}", str(obj)[:60])
    return str(obj)[:200]


def _log_truncation(reason: str, preview: str) -> None:
    """Log warning when data is truncated during serialization."""
    logging.warning("Data truncated in _safe_serialize: %s — preview: %s", reason, preview)


def _set_col_widths(ws: Any) -> None:
    for col, w in [("A", 15), ("B", 15), ("C", 80), ("D", 80), ("E", 60), ("F", 100), ("G", 80), ("H", 60)]:
        ws.column_dimensions[col].width = w


def _auto_row_height(ws: Any) -> None:
    """Ajusta alto de filas segun el maximo de lineas JSON en cols 3-4-5-6-7-8."""
    JSON_COLS = [3, 4, 5, 6, 7, 8]
    LINE_HEIGHT = 15
    MAX_HEIGHT = 409
    for row in range(2, ws.max_row + 1):
        max_lines = 1
        for col in JSON_COLS:
            val = ws.cell(row, col).value
            if val:
                lines = str(val).count("\n") + 1
                max_lines = max(max_lines, lines)
        height = min(max_lines * LINE_HEIGHT, MAX_HEIGHT)
        current = ws.row_dimensions[row].height
        if current is None or height > current:
            ws.row_dimensions[row].height = height


def _write_cell(ws: Any, row: int, col: int, value: object, wrap: bool = True) -> None:
    cell = ws.cell(row, col)
    cell.value = value
    cell.number_format = "@"
    cell.alignment = Alignment(wrap_text=True, vertical="top")


# ═══════════════════════════════════════════════════════════════════════════
# VALIDACION DE SHEET
# ═══════════════════════════════════════════════════════════════════════════

def validate_sheet(ws: Any) -> list:
    errores = []
    header = str(ws.cell(1, 2).value or "").strip().lower()
    if "pagina" not in header:
        errores.append("Col B debe tener header 'pagina auditada'")
    has_urls = any(ws.cell(row, 2).value for row in range(2, ws.max_row + 1))
    if not has_urls:
        errores.append("No hay URLs en col B")
    return errores


# ═══════════════════════════════════════════════════════════════════════════
# PERSISTENCIA
# ═══════════════════════════════════════════════════════════════════════════

def save_workbook(wb: Any, path: str) -> str:
    """Guarda con fallback si el archivo está bloqueado."""
    try:
        wb.save(path)
        return path
    except PermissionError:
        name, ext = os.path.splitext(path)
        fallback = f"{name}_browser{ext}"
        try:
            wb.save(fallback)
            logging.warning("Archivo bloqueado, guardado como %s", fallback)
            return fallback
        except PermissionError:
            for attempt in range(3):
                try:
                    _time.sleep(2)
                    wb.save(path)
                    logging.info("Recuperado tras %ds de espera", (attempt + 1) * 2)
                    return path
                except PermissionError:
                    continue
            fallback2 = f"{name}_browser{int(_time.time())}{ext}"
            wb.save(fallback2)
            logging.warning("Lock persistente, guardado como %s", fallback2)
            return fallback2


# ═══════════════════════════════════════════════════════════════════════════
# DATA FILLS (colores condicionales)
# ═══════════════════════════════════════════════════════════════════════════

def _is_json_error(val: str) -> bool:
    try:
        parsed = json.loads(val)
        return isinstance(parsed, dict) and "error" in parsed and "code" in parsed
    except (json.JSONDecodeError, TypeError):
        return False


def _has_json_data(val: str) -> bool:
    try:
        parsed = json.loads(val)
        return isinstance(parsed, dict) and "error" not in parsed
    except (json.JSONDecodeError, TypeError):
        return False


def apply_data_fills(ws: Any) -> None:
    """Aplica colores de fondo a celdas de datos segun su contenido (JSON-aware)."""
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        c = ws.cell(row, 3).value  # digitaldata (manual)
        if c and isinstance(c, str) and _is_json_error(c):
            ws.cell(row, 3).fill = RED
        d = ws.cell(row, 4).value  # digitaldata (automatica)
        if d and isinstance(d, str) and _is_json_error(d):
            ws.cell(row, 4).fill = RED
        f = ws.cell(row, 6).value  # AA analytics (automatico)
        if f and isinstance(f, str) and _has_json_data(f):
            ws.cell(row, 6).fill = YELLOW
        g = ws.cell(row, 7).value  # AA analytics (estructurado)
        if g and isinstance(g, str) and _has_json_data(g):
            ws.cell(row, 7).fill = DATA_FILLS["G"]


# ═══════════════════════════════════════════════════════════════════════════
# SPLIT AA
# ═══════════════════════════════════════════════════════════════════════════

def _has_digitaldata(col_d: object) -> bool:
    """Chequea si col D tiene digitalData válido (no error/empty)."""
    if not col_d or not isinstance(col_d, str) or col_d.strip() in ("", "-", "N/A"):
        return False
    if "(no digitaldata)" in col_d:
        return False
    try:
        dd = json.loads(col_d)
        if isinstance(dd, dict) and dd.get("error") == "no digitaldata":
            return False
        return True
    except (json.JSONDecodeError, ValueError):
        return False


def _has_aa(col_e: object) -> bool:
    """Chequea si col E tiene datos AA reales (no error)."""
    if not col_e or not isinstance(col_e, str) or col_e.strip() in ("", "-", "N/A"):
        return False
    try:
        parsed = json.loads(col_e)
        return bool(isinstance(parsed, dict) and "solution" in parsed and "error" not in parsed)
    except (json.JSONDecodeError, TypeError):
        return False


def split_aa_workbooks(wb: Any, audit_date: str, output_dir: str) -> None:
    """Crea con_aa.xlsx (tiene AA o digitalData) y sin_aa.xlsx (no tiene nada).

    Ahora considera AMBAS columnas: AA (col E) y digitalData (col D).
    """
    ws = wb[audit_date]
    con_rows, sin_rows = [], []
    for row in range(2, ws.max_row + 1):
        col_d = ws.cell(row, 4).value  # digitaldata (automatica)
        col_f = ws.cell(row, 6).value  # AA analytics (automatico)
        has_data = _has_aa(col_f) or _has_digitaldata(col_d)
        if has_data:
            con_rows.append(row)
        else:
            sin_rows.append(row)

    for suffix, rows in [("con_aa", con_rows), ("sin_aa", sin_rows)]:
        path = os.path.join(output_dir, f"{suffix}.xlsx")
        swb = openpyxl.Workbook()
        swb.remove(swb.active)
        sws = swb.create_sheet(audit_date)
        sws.append(SHEET_HEADERS)
        for col_letter, fill in HEADER_FILLS.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            sws.cell(1, col_idx).fill = fill
        for row_num in rows:
            for col in range(1, 9):
                src = ws.cell(row_num, col)
                dst = sws.cell(row_num, col)
                dst.value = src.value
                if src.has_style:
                    dst.fill = copy(src.fill) if src.fill else None
        _set_col_widths(sws)
        apply_data_fills(sws)
        _auto_row_height(sws)
        save_workbook(swb, path)
        swb.close()
        logging.info("Split %s: %d filas -> %s", suffix, len(rows), path)


# ═══════════════════════════════════════════════════════════════════════════
# MULTI-SHEET HISTORIAL
# ═══════════════════════════════════════════════════════════════════════════

def setup_multisheet(output_path: str, urls_source: str, resume: bool) -> tuple:
    """Carga o crea workbook con sheets _control + fecha actual.
    Retorna (wb, ws, audit_date, skipped)."""
    if os.path.exists(output_path) and resume:
        wb = openpyxl.load_workbook(output_path)
        audit_date = datetime.now().strftime("%Y-%m-%d")
        if audit_date in wb.sheetnames:
            ws = wb[audit_date]
            # Limpiar filas existentes (sin borrar header)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)
            logging.info("Resumed: clearing existing sheet %s", audit_date)
            return wb, ws, audit_date, True
    elif os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        audit_date = datetime.now().strftime("%Y-%m-%d")
        if audit_date in wb.sheetnames:
            del wb[audit_date]
    else:
        wb = openpyxl.Workbook()
        audit_date = datetime.now().strftime("%Y-%m-%d")
        ws_ctrl = wb.create_sheet("_control")
        ws_ctrl.append(CONTROL_HEADERS)
        ws_ctrl.column_dimensions["A"].width = 14
        ws_ctrl.column_dimensions["B"].width = 30

    ws = wb.create_sheet(audit_date)
    ws.append(SHEET_HEADERS)
    for col_letter, fill in HEADER_FILLS.items():
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws.cell(1, col_idx).fill = fill
    _set_col_widths(ws)
    _auto_row_height(ws)
    return wb, ws, audit_date, False


def update_control(wb: Any, audit_date: str, source: str, total: int,
                   ok_aa: int, ok_dd: int, errors: int, retries: int,
                   score: int, elapsed_s: float, workers: int) -> None:
    """Agrega fila al sheet _control."""
    if "_control" not in wb.sheetnames:
        cs = wb.create_sheet("_control")
        cs.append(CONTROL_HEADERS)
        _set_col_widths(cs)
    else:
        cs = wb["_control"]
    cs.append([
        audit_date, source, total, ok_aa, ok_dd,
        errors, retries, score, round(elapsed_s, 1), workers,
    ])


def update_vars_sheet(wb: Any, audit_date: str, rows_aa: list[tuple[int, dict]]) -> None:
    """Crea/actualiza sheet _vars con las eVars y props detectadas."""
    sheet_name = "_vars"
    if sheet_name in wb.sheetnames:
        vs = wb[sheet_name]
        vs.delete_rows(2, vs.max_row)
    else:
        vs = wb.create_sheet(sheet_name)
        vs.append(["variable", "descripcion"])
        _set_col_widths(vs)

    seen = set()
    for _row_num, parsed in rows_aa:
        for key in parsed.get("eVars", {}):
            if key not in seen:
                seen.add(key)
                vs.append([key, ""])
        for key in parsed.get("props", {}):
            if key not in seen:
                seen.add(key)
                vs.append([key, ""])


# ═══════════════════════════════════════════════════════════════════════════
# PROGRESS
# ═══════════════════════════════════════════════════════════════════════════

def print_progress(done: int, total: int, errors: int, workers: int,
                   start_time: float | None = None) -> None:
    """Muestra barra de progreso en consola con elapsed + ETA."""
    import time as _time
    pct = done / max(total, 1) * 100
    bar_len = 30
    filled = int(bar_len * done / max(total, 1))
    bar = "#" * filled + "." * (bar_len - filled)
    elapsed_str = ""
    eta_str = ""
    if start_time is not None and done > 0:
        elapsed = _time.perf_counter() - start_time
        rate = elapsed / done
        remaining = total - done
        eta = rate * remaining
        elapsed_str = f"  Elapsed: {elapsed:.0f}s"
        eta_str = f"  ETA: {eta:.0f}s"
    print(f"  [{bar}] {done}/{total} ({pct:.0f}%)  Err: {errors}{elapsed_str}{eta_str}")

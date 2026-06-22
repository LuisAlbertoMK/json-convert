"""
styles.py — Estilos compartidos de openpyxl para reportes Excel.

Reduce duplicación entre audit_report.py, generate_migration_catalog.py,
match_prod_preview.py y extract_aa.py.

Cada módulo puede definir sus estilos adicionales localmente.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Header ──
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

# ── Data fonts ──
FONT_DATA = Font(name="Consolas", size=9)
FONT_PARAM = Font(name="Calibri", size=10)

# ── Alignments ──
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Borders ──
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# ── Status fills (verde/ambar/rojo — gama usada por migrate+match) ──
FILL_OK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_WARN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_FAIL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

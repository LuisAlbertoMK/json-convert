"""
extract_browser.py — CLI entry point: navega URLs, extrae AA beacons, escribe Excel.

Delegates logic to json_convert package modules.
Uso: python extract_browser.py [--urls urls.json] [--workers N] ...
"""

from __future__ import annotations

import asyncio
import logging
import os
import re
import signal
import subprocess
import sys
import time
from datetime import datetime, timedelta

from json_convert import (  # noqa: F401 — re-export for backwards compat
    VALID_URL_SCHEMES,
    ALLOWED_HOSTNAME_SUFFIXES,
    AA_DOMAINS,
    DATA_LAYER_NAMES,
    INPUT_FILE,
    SAVE_EVERY_N,
    SHEET_HEADERS,
    CONTROL_HEADERS,
    HEADER_FILLS,
    DATA_FILLS,
    ERROR_CODES,
    validate_url,
    sanitize_url_for_log,
    parse_aa_beacon,
    build_aa_from_s,
    extract_s_object,
    extract_digital_data,
    extract_title,
    try_dismiss_cookie_consent,
    validate_sheet,
    save_workbook,
    _pretty_json,
    _set_col_widths,
    _auto_row_height,
    _write_cell,
    _is_json_error,
    _has_json_data,
    apply_data_fills,
    split_aa_workbooks,
    setup_multisheet,
    update_control,
    update_vars_sheet,
    print_progress,
    _error_code_from_detail,
    classify_errors,
    compute_score,
    compute_url_score,
)

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import TimeoutError as PwTimeout
from playwright.async_api import async_playwright

# ── Graceful shutdown ──
_shutdown_flag = False
"""Global flag: set by signal handler to request graceful shutdown."""


def _request_shutdown(signum: int | None = None, frame: object | None = None) -> None:
    global _shutdown_flag
    if not _shutdown_flag:
        _shutdown_flag = True
        logging.warning("Señal %s recibida. Terminando gracefulmente...", signum)


# ═══════════════════════════════════════════════════════════════════════════
# WRITE RESULT — pipeline: escribe una URL en Excel con métricas + auto-save
# ═══════════════════════════════════════════════════════════════════════════

async def write_result(
    ws: object, result: dict, metrics: dict,
    excel_lock: asyncio.Lock, output_path: str,
    saved_count: list,  # [int] mutable para closure
    show_progress: bool = False,
    total_urls: int = 0,
    workers: int = 1,
    start_time: float | None = None,
) -> None:
    """
    Escribe el resultado de una URL en el Excel.
    Usa lock para evitar corrupción entre workers.
    """
    async with excel_lock:
        row = result["row"]
        n_beacons = 1 + len(result.get("extra_beacons", []))

        # digitaldata → col C
        dd_val = result.get("digitaldata")
        if dd_val is not None:
            _write_cell(ws, row, 3, _pretty_json(dd_val))
            metrics["ok_dd"] += 1
        else:
            _write_cell(ws, row, 3, _pretty_json({"error": "no digitaldata", "code": "DD_MISSING"}))

        # AA → col D
        if result.get("aa_parsed"):
            _write_cell(ws, row, 4, _pretty_json(result["aa_parsed"]))
            metrics["ok_aa"] += 1
        else:
            err_code = _error_code_from_detail(result.get("error", "no AA"))
            _write_cell(ws, row, 4, _pretty_json({"error": result.get("error", "no AA"), "code": err_code}))

        # Score por URL (0-100)
        url_score = compute_url_score(result)

        # Metadata → col F
        meta = {"score": url_score, "status": result.get("status", 0),
                "aa_source": result.get("aa_source"),
                "beacons": n_beacons, "title": result.get("title", ""),
                "error": result.get("error"), "code": result.get("code"),
                "elapsed_s": result.get("elapsed_s", 0),
                "url": result.get("url", "")[:120]}
        if result.get("extra_beacons"):
            meta["extra_beacons"] = result["extra_beacons"]
        _write_cell(ws, row, 6, _pretty_json(meta))

        if result.get("error") or not result["aa_parsed"]:
            metrics["errors"] += 1
            metrics["errores_detalle"].append({"row": row, "error": result.get("error", "no AA")})

        metrics["total_beacons"] += n_beacons
        metrics["times"].append(result["elapsed_s"])
        metrics["retries"] += result.get("retries_used", 0)

        # Guardado incremental
        saved_count[0] += 1
        if saved_count[0] % SAVE_EVERY_N == 0:
            save_workbook(ws.parent, output_path)
            logging.info("  Guardado incremental (#%d)", saved_count[0])
        if show_progress:
            print_progress(saved_count[0], total_urls, metrics["errors"], workers,
                           start_time=start_time)


# ═══════════════════════════════════════════════════════════════════════════
# WORKER — procesa UNA URL, devuelve resultado
# ═══════════════════════════════════════════════════════════════════════════

BEACON_REGEX = re.compile(
    r'https?://[^"\'\s]*(?:'
    + "|".join(re.escape(d) for d in AA_DOMAINS)
    + r')[^"\'\s]*',
    re.IGNORECASE,
)


async def process_url(
    page: object, row: int, url: str,
    wait_after: int = 4,
    timeout_ms: int = 35000,
    max_retry: int = 1,
) -> dict:
    """
    Navega a una URL, captura beacons + data layer.
    Devuelve dict con resultados.
    """
    if _shutdown_flag:
        return {"url": url, "row": row, "error": "Shutdown requested"}

    start = time.perf_counter()
    err = validate_url(url)
    if err:
        return {"url": url, "row": row, "error": err}

    all_beacons = []
    last_dd = None
    last_s = None
    last_title = ""
    navigation_error = ""
    parsed_aa = []
    extra_aa = []

    async def _on_response(response: object) -> None:
        if _shutdown_flag:
            return
        url_lower = response.url.lower()
        if any(domain in url_lower for domain in AA_DOMAINS):
            try:
                body = await response.body()
                beacon = response.url
                if "?" in beacon and len(beacon) > 50:
                    all_beacons.append(beacon)
            except Exception:
                pass

    for attempt in range(1 + max(0, max_retry)):
        if _shutdown_flag:
            break
        try:
            async with page.context.expect_page(timeout=timeout_ms) as popup_info:
                await page.goto(url, wait_until="domcontentloaded",
                                timeout=timeout_ms)
            try:
                popup = await popup_info.value
                await popup.close()
                logging.debug("Popup cerrado: %s", url)
            except Exception:
                pass

            await page.wait_for_timeout(wait_after * 1000)

            last_title = await extract_title(page)
            last_dd = await extract_digital_data(page)
            last_s = await extract_s_object(page)
            await try_dismiss_cookie_consent(page)
            await page.wait_for_timeout(1000)

            if not _shutdown_flag:
                navigation_error = ""
                break

        except PwTimeout:
            navigation_error = "Timeout al navegar"
            logging.debug("Timeout en intento %d: %s", attempt + 1, url)
            if attempt < max_retry:
                # ⚠ page puede estar cerrada si el browser perdió conexión
                try:
                    await page.wait_for_timeout(2000)
                except Exception:
                    break
        except asyncio.CancelledError:
            navigation_error = "Tarea cancelada"
            logging.debug("Cancelled en intento %d: %s", attempt + 1, url)
            raise  # No reintentar — propagar cancelación
        except Exception as e:
            navigation_error = f"Error al navegar: {e}"
            logging.debug("Error en intento %d: %s", attempt + 1, e)
            if attempt < max_retry:
                # ⚠ page puede estar cerrada si el browser perdió conexión
                try:
                    await page.wait_for_timeout(2000)
                except Exception:
                    break

    # Parsear beacons
    extra_count = 0
    for beacon in all_beacons:
        if extra_count < 5:
            try:
                parsed = parse_aa_beacon(beacon, last_title)
                if parsed.get("hit", {}).get("type") in ("pageView", None):
                    parsed_aa.append(parsed)
                else:
                    extra_aa.append(parsed)
                    extra_count += 1
            except Exception:
                continue

    elapsed = time.perf_counter() - start
    result = {
        "url": url, "row": row,
        "page_name": _page_name_from_url(url),
        "digitaldata": last_dd,
        "raw_beacons": all_beacons,
        "aa_parsed": parsed_aa[0] if parsed_aa else None,
        "extra_beacons": extra_aa,
        "metadata": {
            "title": last_title,
            "s_object": last_s,
            "elapsed_s": round(elapsed, 1),
            "beacon_count": len(all_beacons),
        },
        "elapsed_s": round(elapsed, 1),
        "status": 0,  # página cargada, código HTTP no disponible con domcontentloaded
        "aa_source": None,  # se actualiza en write_result si hay AA
        "title": last_title,
        "code": None,
        "retries_used": attempt,
    }
    if navigation_error:
        result["error"] = navigation_error
    return result


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _page_name_from_url(url: str) -> str:
    """Deriva un nombre legible desde la URL."""
    try:
        parts = url.rstrip("/").split("/")
        for i, p in enumerate(parts):
            if "ford" in p and i + 2 < len(parts):
                return parts[-1].replace(".html", "").replace("-", " ").title()
        return parts[-1].replace(".html", "").replace("-", " ").title()
    except Exception:
        return url[:60]


# ═══════════════════════════════════════════════════════════════════════════
# MAIN (async)
# ═══════════════════════════════════════════════════════════════════════════

async def amain() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    _setup_logging(args.verbose)

    # ── Fuente de URLs ──
    urls, url_source = _resolve_urls(args)
    output_path = args.output or _resolve_output(url_source, args.market)

    # Mostrar conteo antes de empezar
    market_label = args.market.upper() if args.market else "TODOS"
    env_label = {"preview": "PREVIEW", "produccion": "PRODUCCION", "ambas": "TODOS"}.get(args.entorno, args.entorno)
    print(f"\n  Auditando {len(urls)} URLs [{env_label}] para mercado {market_label}...\n")

    # Crear directorio de output si no existe (ej: PR/)
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # ── Excel (crear ANTES del pipeline) ──
    wb, ws, audit_date, _ = setup_multisheet(output_path, url_source, resume=False)

    # ── Playwright ──
    async with async_playwright() as pw:
        # Usar Chrome real (evita bloqueo Akamai WAF); fallback a bundled Chromium
        launch_kwargs = dict(
            headless=not args.headed,
            args=["--disable-blink-features=AutomationControlled"],
            proxy={"server": args.proxy} if args.proxy else None,
        )
        try:
            browser = await pw.chromium.launch(channel="chrome", **launch_kwargs)
        except Exception:
            logging.info("Chrome no disponible, usando Chromium bundled")
            browser = await pw.chromium.launch(**launch_kwargs)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
            locale="es_PR" if "es" in url_source else "en_US",
            viewport={"width": 1920, "height": 1080},
            ignore_https_errors=True if args.proxy else False,
        )
        context.set_default_timeout(args.timeout * 1000 if args.timeout else 35000)
        await context.route("**/*", lambda route, req: _route_beacons(route, req))

        # ── Pipeline (procesa URLs + escribe Excel) ──
        results, errors_detail, metrics = await _run_pipeline(
            context, urls, args, ws, output_path,
        )

        # ── Post-pipeline ──
        _auto_row_height(ws)
        apply_data_fills(ws)

        update_control(
            wb, audit_date, url_source,
            metrics["total"], metrics["ok_aa"], metrics["ok_dd"],
            metrics["errors"], metrics["retries"],
            compute_score(metrics), metrics.get("total_time", 0), args.workers or 3,
        )

        aa_rows = [(r["row"], r["aa_parsed"]) for r in results if r.get("aa_parsed")]
        if aa_rows:
            update_vars_sheet(wb, audit_date, aa_rows)

        out = save_workbook(wb, output_path)
        wb.close()

        split = getattr(args, "split_aa", False)
        if split and audit_date:
            output_dir = os.path.dirname(output_path) or "."
            split_aa_workbooks(wb, audit_date, output_dir)

        logging.info("Guardado: %s", out)
        _print_metrics(metrics, args, output_path)

        await browser.close()


def _build_parser() -> argparse.ArgumentParser:
    import argparse
    p = argparse.ArgumentParser(description="Extrae Adobe Analytics de URLs Ford")
    p.add_argument("--urls", help="Archivo JSON con URLs")
    p.add_argument("--input", default=INPUT_FILE, help="Excel de entrada")
    p.add_argument("--output", help="Excel de salida")
    p.add_argument("--market", help="Filtrar por mercado en urls.json")
    p.add_argument("--entorno", default="preview", choices=("preview", "produccion", "ambas"),
                   help="Entorno a auditar: preview (default), produccion, o ambas")
    p.add_argument("--row", type=int, help="Fila inicial en Excel clasico")
    p.add_argument("--workers", type=int, default=3, help="Concurrencia")
    p.add_argument("--timeout", type=int, default=35, help="Timeout por URL (s)")
    p.add_argument("--wait-after", type=int, default=4, help="Espera post-carga (s)")
    p.add_argument("--max-retry", type=int, default=1, help="Reintentos por URL")
    p.add_argument("--headless", action="store_true", help="Forzar headless")
    p.add_argument("--headed", action="store_true", help="Forzar headed (con UI)")
    p.add_argument("--proxy", help="Proxy HTTP (ej: http://proxy:8080)")
    p.add_argument("--resume", action="store_true", help="Reanudar corrida existente")
    p.add_argument("--split-aa", action="store_true", help="Dividir con/sin AA")
    p.add_argument("--progress", action="store_true", help="Mostrar barra de progreso")
    p.add_argument("--log-file", help="Archivo de log")
    p.add_argument("--verbose", action="store_true", help="Logging debug")
    p.add_argument("--run-clean", action="store_true", help="Forzar clean run")
    return p


def _setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s [%(levelname)s] %(message)s",
                        datefmt="%H:%M:%S")


def _resolve_urls(args: argparse.Namespace) -> tuple[list[tuple[int, str]], str]:
    """Determina fuente de URLs: JSON o Excel clasico."""
    if args.urls:
        import json
        with open(args.urls, encoding="utf-8") as f:
            all_entries = json.load(f)
        market = args.market
        env = args.entorno or "preview"
        if env == "ambas":
            if market:
                entries = [e for e in all_entries
                           if e.get("market", "").upper() == market.upper()]
            else:
                entries = all_entries
        else:
            if market:
                entries = [e for e in all_entries
                           if e.get("market", "").upper() == market.upper()
                           and e.get("entorno", "preview") == env]
            else:
                entries = [e for e in all_entries if e.get("entorno", "preview") == env]
        urls = [(i + 2, e["url"]) for i, e in enumerate(entries)]
        source = args.market or "multi"
        return urls, source

    # Modo clasico: Excel plano
    input_path = args.input or INPUT_FILE
    if not os.path.exists(input_path):
        print(f"[ERR] No se encuentra {input_path}")
        sys.exit(1)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    errs = validate_sheet(ws)
    if errs:
        for e in errs:
            print(f"[ERR] {e}")
        sys.exit(1)
    row_start = args.row or 2
    urls = [(r, str(ws.cell(r, 2).value or "").strip())
            for r in range(row_start, ws.max_row + 1) if ws.cell(r, 2).value]
    wb.close()
    source = "clasico"
    return urls, source


def _resolve_output(url_source: str, market: str = None) -> str:
    """Determina nombre de output segun fuente."""
    if url_source == "clasico":
        return "resultado.xlsx"
    if market:
        return os.path.join(market.upper(), "historial.xlsx")
    return "historial.xlsx"


async def _run_pipeline(
    context: object,
    urls: list[tuple[int, str]],
    args: argparse.Namespace,
    ws: object,
    output_path: str,
) -> tuple[list[dict], list[dict], dict]:
    """Ejecuta pipeline de URLs con Playwright, escribe Excel vía write_result."""
    semaphore = asyncio.Semaphore(args.workers or 3)
    excel_lock = asyncio.Lock()
    saved_count = [0]
    results = []
    errors_detail = []
    metrics = {
        "total": len(urls), "ok_aa": 0, "ok_dd": 0,
        "errors": 0, "retries": 0,
        "total_beacons": 0, "times": [],
        "errores_detalle": [],
    }

    async def _process_one(row: int, url: str) -> dict:
        async with semaphore:
            page = await context.new_page()
            try:
                beacons: list[str] = []

                async def _capture(response: object) -> None:
                    url_lower = response.url.lower()
                    if any(d in url_lower for d in AA_DOMAINS):
                        try:
                            body = await response.body()
                            beacons.append(response.url)
                        except Exception:
                            pass

                page.on("response", _capture)
                result = await process_url(
                    page, row, url,
                    wait_after=args.wait_after or 4,
                    timeout_ms=(args.timeout or 35) * 1000,
                    max_retry=args.max_retry or 1,
                )
                result["raw_beacons"] = beacons
                return result
            finally:
                await page.close()

    pipeline_start = time.perf_counter()

    async def _worker(row: int, url: str) -> None:
        if _shutdown_flag:
            return
        try:
            result = await _process_one(row, url)
        except Exception as e:
            # Si process_url no pudo manejar el error, registrar como fallo
            logging.error("[%s] Fallo grave: %s", url[:60], e)
            result = {"url": url, "row": row, "error": str(e),
                      "aa_parsed": None, "digitaldata": None,
                      "status": -1, "aa_source": None, "title": "",
                      "code": "FATAL", "elapsed_s": 0, "retries_used": 0}
            metrics["errors"] += 1
        results.append(result)

        # Escribir en Excel + actualizar métricas via pipeline write_result
        try:
            await write_result(
                ws, result, metrics, excel_lock, output_path, saved_count,
                show_progress=bool(args.progress),
                total_urls=metrics["total"],
                workers=args.workers or 3,
                start_time=pipeline_start,
            )
        except Exception as e:
            logging.error("[%s] Error escribiendo resultado: %s", url[:60], e)

        if result.get("error"):
            errors_detail.append({"row": row, "error": result["error"]})

        if not args.progress:
            url_index = row - 1  # 1-based URL index
            status = "OK" if not result.get("error") else "ERR"
            elapsed = time.perf_counter() - pipeline_start
            logging.info("[URL %d/%d] %s %s  (%ds)", url_index, metrics["total"],
                         status, url[:60], int(elapsed))

    tasks = [_worker(row, url) for row, url in urls]
    await asyncio.gather(*tasks, return_exceptions=True)

    metrics["total_time"] = sum(metrics["times"])
    metrics["errores_detalle"] = errors_detail

    # Guardado final
    save_workbook(ws.parent, output_path)
    logging.info("Guardado final (%d URLs)", len(results))

    return results, errors_detail, metrics


async def _route_beacons(route: object, request: object) -> None:
    """Intercepta requests a dominios AA."""
    url_lower = request.url.lower()
    if any(domain in url_lower for domain in AA_DOMAINS):
        await route.continue_()
    else:
        await route.continue_()


def _print_metrics(metrics: dict, args: argparse.Namespace, output_path: str) -> int:
    """Muestra metricas finales en consola."""
    success_rate = (metrics["ok_aa"] / max(metrics["total"], 1)) * 100
    dd_rate = (metrics["ok_dd"] / max(metrics["total"], 1)) * 100
    avg_time = sum(metrics["times"]) / max(len(metrics["times"]), 1)
    beacons_per_url = metrics["total_beacons"] / max(metrics["total"], 1)
    total_time = metrics.get("total_time", 0)
    score = compute_score(metrics)

    print(f"""
{"=" * 55}
  METRICAS Y SCORE
{"=" * 55}
  Config:            {args.workers} worker(s) concurrente(s)
  URLs procesadas:   {metrics['ok_aa']}/{metrics['total']}
  AA capturados:     {success_rate:.0f}%
  digitaldata:       {dd_rate:.0f}%
  Beacons totales:   {metrics['total_beacons']} ({beacons_per_url:.1f}/URL)
  Reintentos:        {metrics['retries']}
  Errores:           {metrics['errors']}
{"-" * 55}
  Tiempo total:      {timedelta(seconds=int(total_time))}
  Promedio/URL:      {avg_time:.1f}s
  Guardados incr.:   cada {SAVE_EVERY_N} URLs
{"-" * 55}
  SCORE GLOBAL:      {score}/100
{"-" * 55}""")

    if metrics.get("errores_detalle"):
        print(f"  DETALLE ERRORES ({len(metrics['errores_detalle'])}):")
        categories = classify_errors(metrics["errores_detalle"])
        for cat, rows in categories.items():
            rows_str = ", ".join(str(r) for r in rows)
            print(f"    * {cat}: Fila(s) {rows_str}")

    if score < 60:
        print("  Sugerencias:")
        if success_rate < 50:
            print("    * Verificar VPN / acceso a las URLs")
        if dd_rate < 50:
            print("    * El data layer podria no llamarse window.digitaldata")
        if avg_time > 30:
            print("    * Paginas lentas. Aumentar --timeout o verificar SPAs")

    if os.path.exists(output_path):
        print(f"\n{'=' * 55}")
        print("  Limpiando AA en Excel...")
        print(f"{'=' * 55}")
        clean_script = os.path.join(os.path.dirname(__file__) or ".", "extract_aa.py")
        cmd = [sys.executable, clean_script, "--input", output_path]
        r = subprocess.run(cmd, capture_output=True, text=True)
        print(r.stdout)

    return score


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    try:
        signal.signal(signal.SIGINT, _request_shutdown)
        signal.signal(signal.SIGTERM, _request_shutdown)
    except (ValueError, AttributeError):
        pass
    try:
        asyncio.run(amain())
    except KeyboardInterrupt:
        _request_shutdown()
        logging.warning("Interrumpido por usuario.")
        sys.exit(130)


if __name__ == "__main__":
    main()

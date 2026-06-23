"""
extract_aa.py — Extrae campos seleccionables del JSON de Adobe Analytics en Excel.

Lee col F (AA analytics automatico), extrae campos elegidos, escribe en col G (estructurado).

Header-aware: detecta columnas por nombre (backwards compat con formato viejo de 6 cols).

Uso:
  python extract_aa.py                                          # valores por defecto
  python extract_aa.py --keep page,request,props,evars          # solo estos 4
  python extract_aa.py --keep events,visitor,products           # campos adicionales
  python extract_aa.py --keep all                               # TODO el JSON (solo pretty-print)
  python extract_aa.py --input historial.xlsx                   # otro archivo
  python extract_aa.py --score                                  # métricas detalladas
  python extract_aa.py --verbose                               # logging debug

Maneja 2 formatos detectados en col E:
  - Grupo 1 (filas 2-10): keys "eVars" + "props" con prop1/eVar1
  - Grupo 2 (filas 11-15): keys "evars" + "props" con c1/v1
"""

import json
import logging
import os
import subprocess
import sys

import openpyxl
from openpyxl.styles import Alignment, PatternFill

from json_convert.excel import _auto_row_height, save_workbook

DEFAULT_KEEP = ["page", "request", "props", "evars"]
ALL_FIELDS = [
    "solution", "page", "request", "visitor", "hit", "page_data",
    "browser", "events", "eVars", "evars", "props",
    "products", "frame", "adobe", "pageName", "channel", "page_attributes", "environment",
]


def extract_fields(data: dict, keep: list[str]) -> dict:
    """
    Extrae del JSON original solo los campos en `keep`.
    Maneja eVars/evars indistintamente.
    """
    result = {}
    for field in keep:
        if field == "evars":
            # Unifica eVars (Grupo 1) y evars (Grupo 2)
            val = data.get("eVars") or data.get("evars")
            if val is not None:
                result["evars"] = val
        elif field == "eVars":
            val = data.get("eVars") or data.get("evars")
            if val is not None:
                result["eVars"] = val
        else:
            val = data.get(field)
            if val is not None:
                result[field] = val
    return result


def count_values(obj) -> dict:
    """Cuenta elementos de cada campo en el JSON extraído (para métricas)."""
    stats = {}
    for key, val in obj.items():
        if isinstance(val, dict):
            stats[key] = f"{len(val)} keys"
        elif isinstance(val, list):
            stats[key] = f"{len(val)} items"
        elif isinstance(val, str):
            stats[key] = f"{len(val)} chars" if len(val) > 50 else val[:50]
        else:
            stats[key] = str(val)[:50]
    return stats


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Extrae campos del JSON de Adobe Analytics en Excel")
    parser.add_argument("--urls", help="urls.json para auto-generar input si no existe")
    parser.add_argument("--input", default="RevisionManual.xlsx", help="Archivo Excel")
    parser.add_argument("--market", help="Mercado específico (para auto-generación)")
    parser.add_argument("--sheet", help="Nombre del sheet (default: auto-detecta el mas reciente)")
    parser.add_argument("--keep", default=",".join(DEFAULT_KEEP),
                        help=f"Campos a conservar (separados por coma). Opciones: {','.join(ALL_FIELDS)}. Usar 'all' para todo.")
    parser.add_argument("--score", action="store_true", help="Mostrar metricas detalladas por fila")
    parser.add_argument("--verbose", action="store_true", help="Logging detallado (debug)")
    args = parser.parse_args()

    # ── Logging ──
    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(level=log_level, format="%(asctime)s | %(levelname)s | %(message)s",
                        handlers=[logging.StreamHandler(sys.stderr)])

    # Parsear keep
    if args.keep == "all":
        keep = ALL_FIELDS
    else:
        keep = [k.strip() for k in args.keep.split(",") if k.strip()]
        # Validar
        unknown = [k for k in keep if k not in ALL_FIELDS and k != "all"]
        if unknown:
            logging.error("Campos desconocidos: %s. Opciones validas: %s o 'all'",
                          unknown, ",".join(ALL_FIELDS))
            sys.exit(1)

    # ── Auto-bootstrap desde urls.json si no existe el input ──
    if not os.path.exists(args.input):
        urls_path = args.urls
        if not urls_path:
            # Buscar urls.json en CWD o junto al script
            for candidate in [os.path.join(os.getcwd(), "urls.json"),
                              os.path.join(os.path.dirname(__file__), "urls.json")]:
                if os.path.exists(candidate):
                    urls_path = candidate
                    break
        if urls_path and os.path.exists(urls_path):
            # Estimar cantidad de URLs
            _n_urls = 0
            try:
                with open(urls_path) as _f:
                    _all_urls = json.load(_f)
                if args.market:
                    _n_urls = len([e for e in _all_urls
                                   if e.get("market", "").upper() == args.market.upper()
                                   and e.get("entorno", "preview") == "preview"])
                else:
                    _n_urls = len([e for e in _all_urls if e.get("entorno", "preview") == "preview"])
            except Exception:
                pass
            urls_label = f" ({_n_urls} URLs)" if _n_urls else ""
            logging.warning("No se encuentra '%s'. Generando%s desde %s...",
                            args.input, urls_label, urls_path)
            extract_browser = os.path.join(os.path.dirname(__file__), "extract_browser.py")
            cmd = [sys.executable, extract_browser, "--urls", urls_path, "--split-aa"]
            if args.market:
                cmd.extend(["--market", args.market])
            elif args.input != "RevisionManual.xlsx":
                # Inferir mercado desde el path del input (ej: PR/historial.xlsx → PR)
                input_dir = os.path.dirname(args.input)
                if input_dir and input_dir not in (".", "") and input_dir.isalpha() and input_dir.isupper():
                    cmd.extend(["--market", input_dir])
            try:
                r = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
                if r.returncode != 0:
                    logging.error("extract_browser falló (exit %d): %s", r.returncode, r.stderr[:200])
                    sys.exit(1)
                logging.info("Historial generado correctamente.")
            except subprocess.TimeoutExpired:
                logging.exception("extract_browser excedió el tiempo de espera (600s)")
                sys.exit(1)
            except FileNotFoundError:
                logging.exception("No se encuentra extract_browser.py junto a extract_aa.py")
                sys.exit(1)
        else:
            logging.error("No se encuentra '%s'. Usá --urls <urls.json> para auto-generar.", args.input)
            sys.exit(1)

    wb = openpyxl.load_workbook(args.input)

    # Seleccionar sheet
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            logging.error("Sheet '%s' no encontrado", args.sheet)
            sys.exit(1)
        ws = wb[args.sheet]
    else:
        # Auto-detect: el sheet de datos mas reciente (excluye _control)
        candidates = [s for s in wb.sheetnames if s != "_control"]
        candidates.sort(key=lambda s: s if "-" in s else "0", reverse=True)
        if not candidates:
            logging.error("No hay sheets de datos en el archivo")
            sys.exit(1)
        ws = wb[candidates[0]]
        logging.info("Sheet auto-detectado: '%s'", candidates[0])

    # Header-aware column detection (backwards compat: old=6cols, new=7cols)
    hdr = {str(ws.cell(1, c).value or "").strip().lower(): c for c in range(1, ws.max_column + 1)}
    aa_src_col = hdr.get("aa analytics (automatico)", 6)   # read from: col F
    aa_dst_col = hdr.get("aa analytics (estructurado)", 7)  # write to: col G

    src_col_letter = openpyxl.utils.get_column_letter(aa_src_col)
    if "analytics" not in str(ws.cell(1, aa_src_col).value or "").lower():
        logging.warning("Col %s header no esperado: '%s'",
                        src_col_letter, ws.cell(1, aa_src_col).value)

    total = 0
    errores = []
    stats_rows = []

    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row, aa_src_col).value  # AA analytics (col E nueva / col D vieja)
        if not raw:
            errores.append((row, "COL_E_EMPTY", "col E vacia"))
            continue

        raw_str = str(raw).strip()
        if not raw_str or raw_str.startswith("(no") or raw_str.startswith("(error"):
            code = "NO_AA_DATA"
            errores.append((row, code, f"col E sin datos validos: {raw_str[:60]}"))
            continue

        try:
            data = json.loads(raw_str)
        except json.JSONDecodeError as e:
            errores.append((row, "JSON_INVALID", f"JSON invalido: {e}"))
            continue

        extracted = extract_fields(data, keep)

        # Si no se extrajo nada, avisar
        if not extracted:
            errores.append((row, "NO_FIELDS_MATCHED", "ningun campo coincidio en el JSON origen"))
            continue

        pretty = json.dumps(extracted, indent=2, ensure_ascii=False)
        cell = ws.cell(row, aa_dst_col)
        cell.value = pretty
        cell.number_format = "@"  # texto plano — evita que Excel interprete JSON como numero/fecha
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        total += 1

        if args.score:
            stats_rows.append({"row": row, "fields": list(extracted.keys()), **count_values(extracted)})

    # Ancho col destino
    dst_letter = openpyxl.utils.get_column_letter(aa_dst_col)
    ws.column_dimensions[dst_letter].width = 80

    # Auto-ajuste alto de filas segun contenido JSON
    _auto_row_height(ws)

    # Guardar
    out = save_workbook(wb, args.input)
    wb.close()

    logging.info("Guardado: %s", out)
    logging.info("OK Procesadas: %d filas", total)
    logging.info("Campos extraidos: %s", keep)

    if args.score and stats_rows:
        print(f"\n{'='*55}", file=sys.stderr)
        print("  MÉTRICAS POR FILA (--score)", file=sys.stderr)
        print(f"{'='*55}", file=sys.stderr)
        for s in stats_rows:
            print(f"  Fila {s['row']}: campos={s['fields']}", file=sys.stderr)
            for k, v in s.items():
                if k not in ("row", "fields"):
                    print(f"    {k}: {v}", file=sys.stderr)
            print(file=sys.stderr)

    if errores:
        logging.warning("Errores: %d", len(errores))
        for r, code, msg in errores:
            logging.debug("  Fila %d [%s]: %s", r, code, msg)
    else:
        logging.info("Sin errores")


if __name__ == "__main__":
    main()

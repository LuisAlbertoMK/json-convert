#!/usr/bin/env python3
"""
generate_migration_catalog.py — Genera catálogo de migración AEM.

Compara los valores REALES de digitalData (capturados por extract_browser.py)
contra el estándar ESPERADO (definido en expected.json) y genera un catálogo
con JSON pretty-printed por URL y acciones detalladas.

Uso:
  # 1. Generar template de url-mapping desde RevisionManual.xlsx
  python generate_migration_catalog.py --gen-template

  # 2. Generar catálogo (después de llenar url-mapping.json)
  python generate_migration_catalog.py \
      --historial PR/historial.xlsx \
      --mapping url-mapping.json \
      --expected expected.json \
      --market PR

Flujo completo:
  RevisionManual.xlsx
    → _gen_urls.py → urls.json
    → extract_browser.py → {market}/historial.xlsx
    → generate_migration_catalog.py → {market}/catalogo-migracion.xlsx
"""

import argparse
import json
import os
import re
import sys
from datetime import date

# Force UTF-8 stdout for Windows cp1252 terminal (safe in Python >=3.7)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

from json_convert.styles import (
    ALIGN_CENTER,
    ALIGN_WRAP,
    FILL_FAIL,
    FILL_HEADER,
    FILL_OK,
    FILL_WARN,
    FONT_DATA,
    FONT_HEADER,
    FONT_PARAM,
    THIN_BORDER,
)

# ── Constantes ──
PARAMS_ORDER = ["pageName", "siteSection", "pageNameNoVehicle",
                "client", "site", "variantName", "pageType"]

COLS = [
    "URL de Producción",
    "Path de AEM Migrado",
    "Parámetro (digitalData.page)",
    "Valor en Producción (Actual)",
    "Valor Esperado en AEM (Nuevo Componente)",
    "Estado / Acción para Authoring",
]
COL_WIDTHS = [45, 45, 30, 50, 50, 60]

# ── Estilos locales únicos ──
THICK_TOP = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="medium", color="4472C4"),
    bottom=Side(style="thin", color="D9D9D9"),
)
FILL_REMOVE = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_JSON_BG = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# ── Emojis (solo para Excel, no para print) ──
EMOJI = {"ok": "✅", "warn": "⚠️", "create": "❌", "remove": "🗑️"}


# ════════════════════════════════════════════
#  CORE
# ════════════════════════════════════════════

from json_convert.utils import load_json


def find_digitaldata_in_historial(historial_path: str, preview_url: str) -> dict:
    """Busca el digitalData de una URL en el historial.

    Header-aware: busca "digitaldata (automatica)" o fallback a col 3.
    """
    wb = openpyxl.load_workbook(historial_path, data_only=True)
    data_sheet = None
    for sn in wb.sheetnames:
        if sn not in ("_control", "_vars", "Sheet"):
            data_sheet = wb[sn]
            break
    if not data_sheet and "Sheet" in wb.sheetnames:
        data_sheet = wb["Sheet"]
    if not data_sheet:
        wb.close()
        return {}

    # Header-aware column detection
    dd_col = 3  # default: old format
    for c in range(1, data_sheet.max_column + 1):
        hv = data_sheet.cell(1, c).value
        if hv and "digitaldata (automatica)" in str(hv).strip().lower():
            dd_col = c
            break

    for row in range(2, data_sheet.max_row + 1):
        url_cell = data_sheet.cell(row, 2).value
        if url_cell and preview_url in str(url_cell):
            dd_raw = data_sheet.cell(row, dd_col).value
            wb.close()
            if dd_raw and isinstance(dd_raw, str):
                try:
                    return json.loads(dd_raw)
                except json.JSONDecodeError:
                    return {}
            return {}
    wb.close()
    return {}


def resolve_expected(param_name: str, page_key: str, cfg: dict) -> tuple:
    """
    Resuelve el valor esperado para un parámetro según el page_key.

    Retorna (expected_value, action, detail).
    action: ok | warn | create | remove
    """
    param_cfg = cfg.get("params", {}).get(param_name, {})
    rule = param_cfg.get("rule", "pattern")

    # No tenemos actual acá, solo resolvemos el expected
    if rule == "deprecated":
        return (None, "remove", param_cfg.get("note", "Eliminar"))

    if rule == "fixed":
        return (param_cfg.get("value", ""), "warn",
                param_cfg.get("note", "Valor fijo requerido"))

    if rule == "mirror":
        return ("(ver pageName)", "warn",
                param_cfg.get("note", "Alinear con pageName"))

    if rule == "required":
        val = param_cfg.get("mapping", {}).get(page_key, "")
        if not val:
            return ("", "create", "Sin regla")
        return (val, "create", param_cfg.get("note", "Nuevo parámetro"))

    if rule == "mapping":
        val = param_cfg.get("mapping", {}).get(page_key, "")
        if not val:
            return ("", "create", "Sin regla")
        return (val, "warn", "Mapear sección")

    # pattern (default)
    val = param_cfg.get("patterns", {}).get(page_key, "")
    if not val:
        return ("", "create", "Sin regla")
    return (val, "warn", param_cfg.get("default_note", "Alinear nomenclatura"))


def evaluate_param(param_name: str, actual: str | None,
                   page_key: str, cfg: dict) -> tuple:
    """
    Evalúa un parámetro individual contra el estándar.

    Retorna (expected_value, action_emoji, action_detail).
    """
    expected, base_action, detail = resolve_expected(param_name, page_key, cfg)

    if base_action == "remove":
        if actual is not None:
            return ("🗑️ Eliminar", "🗑️", detail)
        return ("—", "✅", "No aplica (ya eliminado)")

    if actual is None:
        return (f"❌ Crear: {expected}", "❌",
                f"Crear: {detail}")

    if str(actual) == str(expected):
        return (f"✅ {actual}", "✅", "Componente alineado")

    # Diferente: determinar detalle específico
    if param_name == "pageName":
        prefix = cfg.get("prefix", "")
        if not str(actual).startswith(prefix):
            return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                    f"Agregar prefijo '{prefix}'")
        return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                f"Alinear: {detail}")

    if param_name == "siteSection":
        return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                f"Sección legacy '{actual}' → '{expected}'")

    if param_name == "site":
        return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                f"Unificar sitio '{actual}' → '{expected}'")

    return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
            f"Modificar: {detail}")


def build_expected_page(page_key: str, cfg: dict) -> dict:
    """Construye el dict digitalData.page esperado según las reglas."""
    expected = {}
    for param in PARAMS_ORDER:
        val, action, _ = resolve_expected(param, page_key, cfg)
        if action == "remove":
            continue  # no incluir deprecated
        if val is not None and val != "":
            expected[param] = val
    return expected


def compare_page(actual_page: dict, expected_page: dict, cfg: dict,
                 page_key: str) -> list[dict]:
    """Compara parámetro por parámetro y genera acciones."""
    results = []
    for param in PARAMS_ORDER:
        actual_val = actual_page.get(param)
        actual_str = str(actual_val) if actual_val is not None else None

        expected_val, base_action, detail = resolve_expected(param, page_key, cfg)

        action_emoji = "✅"
        action_text = "Componente alineado"

        if base_action == "remove":
            if actual_val is not None:
                action_emoji = "🗑️"
                action_text = f"Eliminar: {detail}"
                expected_str = "—"
            else:
                action_emoji = "✅"
                action_text = f"No aplica ({detail})"
                expected_str = "—"
            results.append({
                "param": param,
                "actual": actual_str,
                "expected": expected_str,
                "action_emoji": action_emoji,
                "action_text": action_text,
            })
            continue

        expected_str = str(expected_val) if expected_val is not None else ""

        if actual_str is None:
            action_emoji = "❌"
            action_text = f"Crear: {detail}"
        else:
            param_cfg = cfg.get("params", {}).get(param, {})
            rule = param_cfg.get("rule", "pattern")
            # Para pattern rules: usar startswith si expected es un prefijo
            if rule == "pattern" and actual_str.startswith(expected_str):
                action_emoji = "✅"
                action_text = "Componente alineado"
            elif actual_str == expected_str:
                action_emoji = "✅"
                action_text = "Componente alineado"
            else:
                action_emoji = "⚠️"
                if param == "pageName":
                    prefix = cfg.get("prefix", "")
                    if not actual_str.startswith(prefix):
                        action_text = f"Agregar prefijo '{prefix}'"
                    else:
                        action_text = "Alinear nomenclatura"
                elif param == "siteSection":
                    action_text = f"Sección legacy → '{expected_str}'"
                elif param == "site":
                    action_text = f"Unificar sitio → '{expected_str}'"
                else:
                    action_text = f"Alinear: {detail}"

        results.append({
            "param": param,
            "actual": actual_str,
            "expected": expected_str,
            "action_emoji": action_emoji,
            "action_text": action_text,
        })
    return results


# ════════════════════════════════════════════
#  OUTPUT
# ════════════════════════════════════════════

def pp_json(obj: dict) -> str:
    """Pretty-print JSON con indentación para celdas Excel."""
    if not obj:
        return "{ }"
    return json.dumps(obj, indent=2, ensure_ascii=False)


def generate_catalog(historial_path: str, mapping_path: str,
                     expected_path: str, market: str, output_path: str):
    """Genera el catálogo de migración completo."""
    expected_cfg = load_json(expected_path)
    mappings = load_json(mapping_path)
    market_cfg = expected_cfg.get("markets", {}).get(market, {})

    if not market_cfg:
        print(f"[ERR] Mercado '{market}' no encontrado en {expected_path}")
        sys.exit(1)

    # Crear workbook
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = f"Catálogo Migración {market}"

    # Headers (fila 1)
    for col_idx, header in enumerate(COLS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ─── Datos ───
    row_idx = 2
    all_url_results = []  # ← acumula resultados para el resumen

    for mapping in mappings:
        preview_url = mapping.get("preview_url", "")
        production_url = mapping.get("production_url", preview_url)
        aem_path = mapping.get("aem_path", "")
        page_key = mapping.get("page_key", "")
        nombre = mapping.get("nombre", "")

        if not page_key:
            print(f"  [!] Saltando URL sin page_key: {preview_url}")
            continue

        # Obtener digitalData real — probar preview_url primero, luego production_url
        dd = find_digitaldata_in_historial(historial_path, preview_url)
        if not dd and production_url and production_url != preview_url:
            dd = find_digitaldata_in_historial(historial_path, production_url)
        actual_page = dd.get("page", {}) if dd else {}

        if dd:
            print(f"  [+] {preview_url[:55]}... ({nombre})")
        else:
            print(f"  [!] Sin digitalData: {preview_url[:55]}... ({nombre})")

        # Construir expected
        expected_page = build_expected_page(page_key, market_cfg)

        # Comparar
        param_results = compare_page(actual_page, expected_page,
                                     market_cfg, page_key)

        # Acumular para resumen
        all_url_results.append({
            "preview_url": preview_url,
            "production_url": production_url,
            "aem_path": aem_path,
            "page_key": page_key,
            "nombre": nombre,
            "dd_found": bool(dd),
            "actual_page": actual_page,
            "expected_page": expected_page,
            "param_results": param_results,
        })

        # --- FILA 1: URL + JSON actual ---
        actual_json = pp_json(actual_page) if actual_page else "{ }"
        expected_json = pp_json(expected_page)

        # Escribir celdas de la URL (cols 1-2, se combinan para esta URL)
        ws.cell(row_idx, 1, production_url).font = FONT_DATA
        ws.cell(row_idx, 1).alignment = ALIGN_WRAP
        ws.cell(row_idx, 1).border = THICK_TOP

        ws.cell(row_idx, 2, aem_path).font = FONT_DATA
        ws.cell(row_idx, 2).alignment = ALIGN_WRAP
        ws.cell(row_idx, 2).border = THICK_TOP

        ws.cell(row_idx, 3, "📦 digitalData.page (completo)").font = Font(
            name="Calibri", size=10, bold=True)
        ws.cell(row_idx, 3).alignment = ALIGN_WRAP
        ws.cell(row_idx, 3).border = THICK_TOP

        # Actual JSON
        cell_actual = ws.cell(row_idx, 4, actual_json)
        cell_actual.font = Font(name="Consolas", size=9)
        cell_actual.alignment = ALIGN_WRAP
        cell_actual.fill = FILL_JSON_BG
        cell_actual.border = THICK_TOP

        # Expected JSON
        cell_expected = ws.cell(row_idx, 5, expected_json)
        cell_expected.font = Font(name="Consolas", size=9)
        cell_expected.alignment = ALIGN_WRAP
        cell_expected.fill = FILL_JSON_BG
        cell_expected.border = THICK_TOP

        ws.cell(row_idx, 6, "").border = THICK_TOP

        # Alto de fila generoso para el JSON
        max_lines = max(len(actual_json.split("\n")),
                        len(expected_json.split("\n")))
        ws.row_dimensions[row_idx].height = max(60, min(max_lines * 15, 400))

        row_idx += 1

        # --- FILAS SIGUIENTES: un parámetro por fila ---
        for pr in param_results:
            ws.cell(row_idx, 1, "").border = THIN_BORDER
            ws.cell(row_idx, 2, "").border = THIN_BORDER

            cell_param = ws.cell(row_idx, 3, pr["param"])
            cell_param.font = FONT_PARAM
            cell_param.alignment = ALIGN_WRAP
            cell_param.border = THIN_BORDER

            cell_act = ws.cell(row_idx, 4, pr["actual"] if pr["actual"] else "No existe")
            cell_act.font = FONT_PARAM
            cell_act.alignment = ALIGN_WRAP
            cell_act.border = THIN_BORDER

            cell_exp = ws.cell(row_idx, 5, pr["expected"])
            cell_exp.font = FONT_PARAM
            cell_exp.alignment = ALIGN_WRAP
            cell_exp.border = THIN_BORDER

            action_text = f"{pr['action_emoji']} {pr['action_text']}"
            cell_action = ws.cell(row_idx, 6, action_text)
            cell_action.font = FONT_PARAM
            cell_action.alignment = ALIGN_WRAP
            cell_action.border = THIN_BORDER

            # Color según acción
            fill = FILL_OK
            if pr["action_emoji"] == "⚠️":
                fill = FILL_WARN
            elif pr["action_emoji"] == "❌":
                fill = FILL_FAIL
            elif pr["action_emoji"] == "🗑️":
                fill = FILL_REMOVE

            cell_action.fill = fill
            ws.row_dimensions[row_idx].height = 22

            row_idx += 1

        # Fila separadora entre URLs
        row_idx += 1

    # ─── Fila de firma ───
    row_idx += 1
    ws.cell(row_idx, 1,
            f"Generado: {date.today().isoformat()} | "
            f"Mercado: {market} | "
            f"URLs: {len(mappings)}").font = Font(italic=True, color="999999")

    # Freeze pane
    ws.freeze_panes = "A2"

    # Auto-filter en headers
    ws.auto_filter.ref = f"A1:F{row_idx - 1}"

    # Guardar
    out_wb.save(output_path)
    print(f"\n[OK] Catalogo generado: {output_path}")
    print(f"     {len(mappings)} URLs × {len(PARAMS_ORDER)} parámetros")

    # Auto-generar resumen .md + .html
    _auto_summary(all_url_results, market, os.path.dirname(output_path), market_cfg)


# ════════════════════════════════════════════
#  AUTO-SUMMARY (.md + .html)
# ════════════════════════════════════════════

def _auto_summary(all_url_results: list, market: str, output_dir: str,
                  market_cfg: dict):
    """Genera resumen .md y .html del catálogo (sin footer ni próximos pasos).

    all_url_results: lista de dicts generada en generate_catalog()
    market_cfg: config del mercado (ej: expected.json → markets["PR"])
    """

    # ── Contar estados ──
    total_params = 0
    ok_count = 0
    warn_count = 0
    create_count = 0
    remove_count = 0

    for ent in all_url_results:
        for pr in ent["param_results"]:
            total_params += 1
            emoji = pr["action_emoji"]
            if emoji == "✅":
                ok_count += 1
            elif emoji == "⚠️":
                warn_count += 1
            elif emoji == "❌":
                create_count += 1
            elif emoji == "🗑️":
                remove_count += 1

    ok_pct = round(ok_count / total_params * 100, 1) if total_params else 0
    warn_pct = round(warn_count / total_params * 100, 1) if total_params else 0
    create_pct = round(create_count / total_params * 100, 1) if total_params else 0
    remove_pct = round(remove_count / total_params * 100, 1) if total_params else 0

    today = date.today().isoformat()
    total_urls = len(all_url_results)
    prefix = market_cfg.get("prefix", "")

    # ── Detectar patrones comunes ──
    # Recolectar todos los page_types de páginas que no tienen pageType aún
    pages_sin_pageType = []
    missing_prefix = []
    legacy_sections = set()
    for ent in all_url_results:
        if not ent["dd_found"]:
            continue
        for pr in ent["param_results"]:
            if pr["param"] == "pageType" and pr["action_emoji"] == "❌":
                pages_sin_pageType.append(ent["nombre"] or ent["production_url"])
            if pr["param"] == "pageName" and pr["action_emoji"] == "⚠️" and prefix:
                actual = pr.get("actual", "")
                if actual and not actual.startswith(prefix):
                    missing_prefix.append(ent["nombre"] or ent["production_url"])
            if pr["param"] == "siteSection" and pr["action_emoji"] == "⚠️":
                legacy_sections.add(f'{pr["actual"]} → {pr["expected"]}')

    # ═══════════════════════════════════════
    #  .md
    # ═══════════════════════════════════════
    md_lines = [
        f"# Resumen Catálogo de Migración — {market}",
        "",
        f"**Fecha**: {today}",
        f"**URLs analizadas**: {total_urls}",
        f"**Parámetros por URL**: {len(PARAMS_ORDER)}",
        f"**Total parámetros**: {total_params}",
        "",
        "## Resumen General",
        "",
        "| Estado | Cantidad | Porcentaje |",
        "|--------|----------|------------|",
        f"| ✅ Alineados | {ok_count} | {ok_pct}% |",
        f"| ⚠️ Requiere cambios | {warn_count} | {warn_pct}% |",
        f"| ❌ No existe (crear) | {create_count} | {create_pct}% |",
        f"| 🗑️ Deprecado | {remove_count} | {remove_pct}% |",
        "",
    ]

    # Patrones detectados
    if pages_sin_pageType or missing_prefix or legacy_sections:
        md_lines.append("## Patrones Detectados")
        md_lines.append("")
        if pages_sin_pageType:
            md_lines.append(f"- **❌ pageType faltante**: {len(pages_sin_pageType)} páginas — nuevo parámetro del componente EUA")
        if missing_prefix:
            md_lines.append(f"- **⚠️ Prefijo faltante** '{prefix}': {len(missing_prefix)} páginas sin el prefijo en pageName")
        for old_new in sorted(legacy_sections):
            md_lines.append(f"- **⚠️ Sección legacy**: {old_new}")
        md_lines.append("")

    # Detalle por URL
    md_lines.append("## Detalle por URL")
    md_lines.append("")
    all_ok = all(pr["action_emoji"] == "✅" for ent in all_url_results for pr in ent["param_results"])
    if all_ok:
        md_lines.append("_Todas las URLs están alineadas — no se requieren cambios._")
        md_lines.append("")
    else:
        for ent in all_url_results:
            url = ent["production_url"] or ent["preview_url"]
            problems = [pr for pr in ent["param_results"] if pr["action_emoji"] != "✅"]
            if not problems:
                continue

            md_lines.append(f"### {ent['nombre'] or url}")
            md_lines.append(f"- **URL**: {url}")
            if ent["aem_path"]:
                md_lines.append(f"- **AEM**: {ent['aem_path']}")
            md_lines.append("")
            md_lines.append("| Parámetro | Actual → Esperado | Acción |")
            md_lines.append("|-----------|-------------------|--------|")
            for pr in problems:
                actual = pr.get("actual", "") or "(no existe)"
                expected = pr.get("expected", "") or "(sin regla)"
                md_lines.append(f"| `{pr['param']}` | {actual} → {expected} | {pr['action_text']} |")
            md_lines.append("")

    md_path = os.path.join(output_dir, "resumen-catalogo-migracion.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))
    print(f"[OK] Resumen .md: {md_path}")

    # ═══════════════════════════════════════
    #  .html (inline CSS, portable)
    # ═══════════════════════════════════════
    def _html_emoji(e: str) -> str:
        """HTML-safe emoji."""
        return {
            "✅": "&#9989;",
            "⚠️": "&#9888;&#65039;",
            "❌": "&#10060;",
            "🗑️": "&#128465;&#65039;",
        }.get(e, e)

    # Clasificar color
    def _color(e: str) -> str:
        return {"✅": "green", "⚠️": "orange", "❌": "red", "🗑️": "gray"}.get(e, "black")

    rows = []
    for ent in all_url_results:
        url = ent["production_url"] or ent["preview_url"]
        problems = [pr for pr in ent["param_results"] if pr["action_emoji"] != "✅"]
        if not problems:
            continue
        rows.append(f"<tr class='url-sep'><td colspan='3'><strong>{_html_escape(ent['nombre'] or url)}</strong><br><small>{_html_escape(url)}</small></td></tr>")
        for pr in problems:
            actual = _html_escape(pr.get("actual", "") or "(no existe)")
            expected = _html_escape(pr.get("expected", "") or "(sin regla)")
            emoji = _html_emoji(pr["action_emoji"])
            c = _color(pr["action_emoji"])
            rows.append(f"<tr><td><code>{pr['param']}</code></td><td>{actual} → {expected}</td><td style='color:{c}'>{emoji} {_html_escape(pr['action_text'])}</td></tr>")

    # Pct a color
    health_color = "green" if ok_pct >= 80 else ("orange" if ok_pct >= 50 else "red")

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Resumen Catálogo Migración — {_html_escape(market)}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#f8f9fa; color:#333; padding:20px; }}
  h1 {{ font-size:1.5em; margin-bottom:4px; }}
  .meta {{ color:#666; font-size:0.9em; margin-bottom:16px; }}
  .meta span {{ margin-right:16px; }}
  .summary {{ display:flex; gap:12px; margin:16px 0; flex-wrap:wrap; }}
  .card {{ background:#fff; border-radius:8px; padding:16px 20px; flex:1; min-width:120px; box-shadow:0 1px 3px rgba(0,0,0,.08); text-align:center; }}
  .card .num {{ font-size:1.8em; font-weight:700; }}
  .card .label {{ font-size:0.8em; color:#666; }}
  .health {{ font-size:1.1em; margin:12px 0 20px; padding:10px 16px; border-radius:6px; color:#fff; font-weight:600; text-align:center; }}
  .health.green {{ background:#28a745; }}
  .health.orange {{ background:#fd7e14; }}
  .health.red {{ background:#dc3545; }}
  table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
  th {{ background:#4472C4; color:#fff; padding:10px 12px; text-align:left; font-size:0.85em; }}
  td {{ padding:8px 12px; border-bottom:1px solid #eee; font-size:0.85em; vertical-align:top; }}
  tr.url-sep td {{ background:#eef; font-size:0.95em; border-top:2px solid #4472C4; }}
  code {{ background:#f0f0f0; padding:1px 4px; border-radius:3px; font-size:0.9em; }}
  .patrones {{ background:#fff; border-radius:8px; padding:16px 20px; margin:16px 0; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
  .patrones ul {{ margin:8px 0 0 20px; }}
  .patrones li {{ margin:4px 0; }}
</style>
</head>
<body>
<h1>Resumen Catálogo de Migración — {_html_escape(market)}</h1>
<div class="meta">
  <span>📅 {today}</span>
  <span>🌐 {total_urls} URLs</span>
  <span>📊 {total_params} parámetros</span>
</div>

<div class="health {health_color}">
  Salud del catálogo: {ok_pct}% alineado
</div>

<div class="summary">
  <div class="card"><div class="num" style="color:#28a745">{ok_count}</div><div class="label">✅ Alineados</div></div>
  <div class="card"><div class="num" style="color:#fd7e14">{warn_count}</div><div class="label">⚠️ Requiere cambios</div></div>
  <div class="card"><div class="num" style="color:#dc3545">{create_count}</div><div class="label">❌ No existe (crear)</div></div>
  <div class="card"><div class="num" style="color:#6c757d">{remove_count}</div><div class="label">🗑️ Deprecado</div></div>
</div>
"""

    # Patrones
    if pages_sin_pageType or missing_prefix or legacy_sections:
        html += '<div class="patrones"><strong>Patrones Detectados</strong><ul>'
        if pages_sin_pageType:
            html += f"<li>❌ <strong>pageType faltante</strong>: {len(pages_sin_pageType)} páginas — nuevo parámetro del componente EUA</li>"
        if missing_prefix:
            html += f"<li>⚠️ <strong>Prefijo faltante</strong> '{_html_escape(prefix)}': {len(missing_prefix)} páginas sin el prefijo en pageName</li>"
        for old_new in sorted(legacy_sections):
            safe = _html_escape(old_new)
            html += f"<li>⚠️ <strong>Sección legacy</strong>: {safe}</li>"
        html += '</ul></div>'

    # Tabla
    if rows:
        html += "<table><thead><tr><th>Parámetro</th><th>Actual → Esperado</th><th>Acción</th></tr></thead><tbody>"
        html += "\n".join(rows)
        html += "</tbody></table>"
    else:
        html += "<p style='color:green;font-size:1.1em'>✅ Todas las URLs están alineadas — no se requieren cambios.</p>"

    html += "\n</body>\n</html>"

    html_path = os.path.join(output_dir, "resumen-catalogo-migracion.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Resumen .html: {html_path}")


def _html_escape(s: str) -> str:
    """Escapa caracteres HTML."""
    return (s.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;"))


# ════════════════════════════════════════════
#  CLI
# ════════════════════════════════════════════

def generate_template(historial_path: str, output: str):
    """Genera url-mapping.json template desde RevisionManual.xlsx."""
    wb = openpyxl.load_workbook(historial_path, data_only=True)
    ws = wb.active

    mappings = []
    for row in range(2, ws.max_row + 1):
        nombre = ws.cell(row, 1).value
        url = ws.cell(row, 2).value
        if not url:
            continue
        url = str(url).strip()
        if not url:
            continue

        name_str = str(nombre).strip() if nombre else ""
        suggested_key = re.sub(r"[^a-z0-9]+", "-", name_str.lower()).strip("-")
        if not suggested_key:
            suggested_key = "page"

        mappings.append({
            "preview_url": url,
            "production_url": "",
            "aem_path": "",
            "page_key": suggested_key,
            "nombre": name_str
        })

    wb.close()

    with open(output, "w", encoding="utf-8") as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)

    print(f"[OK] Template generado: {output}")
    print(f"     {len(mappings)} URLs. Editar production_url, aem_path y page_key.")
    return mappings


def main():
    parser = argparse.ArgumentParser(
        description="Genera catálogo de migración AEM",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)

    parser.add_argument("--historial", default="PR/historial.xlsx",
                        help="Ruta al historial.xlsx del pipeline")
    parser.add_argument("--mapping", default="data/url-mapping.json",
                        help="JSON mapeo preview→producción")
    parser.add_argument("--expected", default="data/expected.json",
                        help="JSON con valores esperados por mercado")
    parser.add_argument("--market", default="PR",
                        help="Código de mercado (PR, MX...)")
    parser.add_argument("--output", default=None,
                        help="Ruta de salida (default: {market}/catalogo-migracion.xlsx)")
    parser.add_argument("--gen-template", action="store_true",
                        help="Generar url-mapping.json template")
    parser.add_argument("--input", default="RevisionManual.xlsx",
                        help="Excel para --gen-template")

    args = parser.parse_args()

    if args.gen_template:
        generate_template(args.input, args.mapping)
        return

    # Validar archivos
    for fpath, fname in [(args.historial, "historial"),
                         (args.mapping, "mapping"),
                         (args.expected, "expected")]:
        if not os.path.exists(fpath):
            print(f"[ERR] Archivo {fname} no encontrado: {fpath}")
            print("  Usar --gen-template para crear el mapping")
            sys.exit(1)

    # Output default: {market}/catalogo-migracion.xlsx
    if args.output is None:
        market_dir = args.market.upper()
        os.makedirs(market_dir, exist_ok=True)
        output_path = os.path.join(market_dir, "catalogo-migracion.xlsx")
    else:
        output_path = args.output
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    generate_catalog(args.historial, args.mapping,
                     args.expected, args.market, output_path)


if __name__ == "__main__":
    main()

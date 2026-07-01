#!/usr/bin/env python3
"""
generate_validation_matrix.py — Genera matriz de validación estilo v2.

Compara valores REALES de digitalData (de historial.xlsx) contra el estándar
ESPERADO (expected.json) y genera un Excel con el formato de la "Matriz Maestra
de Validación de Datos" de Ford PR.

Columnas (formato v2):
  1. Página / URL
  2. Propiedad del Data Layer (page.*)
  3. Valor Actual
  4. Valor Esperado (PR Adaptado)
  5. Estado
  6. Responsable
  7. Acción Requerida en AEM / Especificación

Uso:
  # Preview
  python src/generate_validation_matrix.py --market PR --entorno preview

  # Producción
  python src/generate_validation_matrix.py --market PR --entorno produccion

  # Ambos entornos (columnas lado a lado)
  python src/generate_validation_matrix.py --market PR --entorno ambas

  # Output custom
  python src/generate_validation_matrix.py --market PR --entorno preview --output PR/preview/mi-matriz.xlsx
"""

import argparse
import json
import os
import re
import sys
from datetime import date

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from json_convert.utils import load_json
from urllib.parse import urlparse

# ── Constants ──
PARAMS_ORDER = [
    "pageName", "siteSection", "hierarchy",
    "pageNameNoVehicle", "client", "site",
    "variantName", "pageType",
]

SHEET_NAME = "Matriz Maestra de Validación"

COLS_SINGLE = [
    "URL Auditada",
    "Página / URL",
    "Propiedad del Data Layer (page.*)",
    "Valor Actual",
    "Valor Esperado (PR Adaptado)",
    "Estado",
    "Responsable",
    "Acción Requerida en AEM / Especificación",
    "Observación / Nota",
]

COLS_SIMPLE = [
    "URL Auditada",
    "Página / URL",
    "Propiedad del Data Layer (page.*)",
    "Valor Actual",
    "Valor Esperado (PR Adaptado)",
]

COLS_DUAL = [
    "URL Auditada",
    "Página / URL",
    "Propiedad del Data Layer (page.*)",
    "Valor Actual (Preview)",
    "Valor Actual (Producción)",
    "Valor Esperado (PR Adaptado)",
    "Estado Preview",
    "Estado Producción",
    "Responsable",
    "Acción Requerida en AEM / Especificación",
    "Observación / Nota",
]

DEFAULT_OUTPUT = "matriz-validacion.xlsx"

# ── Styles ──
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

FILL_OK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_WARN = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
FILL_FAIL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

FONT_DATA = Font(size=10, name="Calibri")
FONT_PARAM = Font(size=10, name="Consolas")
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ════════════════════════════════════════════
#  CORE
# ════════════════════════════════════════════


def _get_historial_sheet(wb):
    """Encuentra la hoja de datos MÁS RECIENTE en un workbook de historial.

    El historial acumula sheets por fecha (2026-06-29, 2026-06-29_2, ...).
    Retorna el último sheet de datos (el más reciente), ignorando sheets
    de control (_control, _vars, Sheet legacy).
    """
    candidates = []
    for sn in wb.sheetnames:
        if sn not in ("_control", "_vars", "Sheet"):
            candidates.append(sn)
    if not candidates:
        return None
    # Tomar el último cronológicamente
    latest = sorted(candidates)[-1]
    return wb[latest]


def _find_dd_col(ws) -> int:
    """Encuentra la columna de digitaldata automatica."""
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(1, c).value
        if hv and "digitaldata (automatica)" in str(hv).strip().lower():
            return c
    return 3  # fallback


def _urls_match(url_a: str, url_b: str) -> bool:
    """Compara dos URLs por path exacto (no substring).

    Retorna True si comparten el mismo path (ignorando trailing slash).
    Previene falsos positivos donde '/' matchea TODO o
    una URL es substring de otra (ej: / vs /camiones/).
    """
    from urllib.parse import urlparse
    if not url_a or not url_b:
        return False
    a = url_a.lower().rstrip("/")
    b = url_b.lower().rstrip("/")
    if a == b:
        return True
    # Comparar paths (mismo netloc + path exacto)
    try:
        pa = urlparse(a)
        pb = urlparse(b)
        # Mismo netloc y path (ignorando trailing slash)?
        if pa.netloc == pb.netloc and pa.path.rstrip("/") == pb.path.rstrip("/"):
            return True
    except Exception:
        pass
    return False


def _build_historial_index(historial_path: str) -> dict | None:
    """Construye un indice URL->digitalData del historial.

    Retorna dict {url_normalizada: data_dict} o None si no existe.
    """
    if not os.path.exists(historial_path):
        return None
    wb = openpyxl.load_workbook(historial_path, data_only=True)
    ws = _get_historial_sheet(wb)
    if not ws:
        wb.close()
        return None
    dd_col = _find_dd_col(ws)
    index = {}
    for row in range(2, ws.max_row + 1):
        url_cell = ws.cell(row, 2).value
        dd_raw = ws.cell(row, dd_col).value
        if url_cell and dd_raw and isinstance(dd_raw, str) and dd_raw.strip():
            try:
                data = json.loads(dd_raw)
                if data:  # solo si se parseo correctamente
                    index[str(url_cell).lower().rstrip("/")] = data
            except json.JSONDecodeError:
                pass
    wb.close()
    return index if index else None


def find_digitaldata_in_historial(historial_path: str, urls: list[str]) -> dict:
    """Busca el digitalData de una lista de URLs en el historial.

    Construye un indice de todo el historial y busca matching flexible
    contra cada URL proporcionada.
    """
    index = _build_historial_index(historial_path)
    if not index:
        return {}
    for url in urls:
        if not url:
            continue
        for hist_url, data in index.items():
            if _urls_match(url, hist_url):
                return data
    return {}


def resolve_expected(param_name: str, page_key: str, cfg: dict) -> tuple:
    """
    Resuelve el valor esperado para un parámetro según el page_key.

    Retorna (expected_value, action_code, action_detail).
    action_code: "ok" | "warn" | "create" | "remove"
    """
    param_cfg = cfg.get("params", {}).get(param_name, {})
    rule = param_cfg.get("rule", "pattern")

    if rule == "deprecated":
        return (None, "remove", param_cfg.get("note", "Eliminar"))

    if rule == "fixed":
        return (param_cfg.get("value", ""), "warn",
                param_cfg.get("note", "Valor fijo requerido"))

    if rule == "mirror":
        page_name_val = (cfg.get("params", {})
                         .get("pageName", {})
                         .get("patterns", {})
                         .get(page_key, ""))
        if not page_name_val:
            return ("(sin pageName)", "warn",
                    param_cfg.get("note", "Alinear con pageName"))
        return (page_name_val, "warn",
                param_cfg.get("note", "Alinear con pageName"))

    if rule == "required":
        val = param_cfg.get("mapping", {}).get(page_key, "")
        if not val:
            inferred = _infer_expected(param_name, page_key, cfg)
            return (inferred, "create",
                    f"⚠️ NO DEFINIDO en expected.json — agregar entrada para '{page_key}'")
        return (val, "create", param_cfg.get("note", "Nuevo parámetro"))

    if rule == "mapping":
        val = param_cfg.get("mapping", {}).get(page_key, "")
        if not val:
            inferred = _infer_expected(param_name, page_key, cfg)
            return (inferred, "warn",
                    f"⚠️ NO DEFINIDO en expected.json — agregar mapping para '{page_key}'")
        return (val, "warn", "Mapear sección")

    # pattern (default)
    val = param_cfg.get("patterns", {}).get(page_key, "")
    if not val:
        inferred = _infer_expected(param_name, page_key, cfg)
        return (inferred, "create",
                f"⚠️ NO DEFINIDO en expected.json — agregar pattern para '{page_key}'")
    return (val, "warn", param_cfg.get("default_note", "Alinear nomenclatura"))


def evaluate_one(param_name: str, actual: str | None, page_key: str,
                 cfg: dict) -> tuple:
    """
    Evalúa un parámetro individual y retorna (estado_emoji, accion_texto).

    estado_emoji: ✅ | ⚠️ | ❌
    accion_texto: texto descriptivo en español como la v2
    """
    expected, base_action, detail = resolve_expected(param_name, page_key, cfg)

    # ── Deprecated ──
    if base_action == "remove":
        if actual is not None:
            return ("⚠️", f"Eliminar: {detail}")
        return ("✅", "No aplica (ya eliminado)")

    # ── No existe en actual ──
    if actual is None:
        if base_action == "create":
            return ("❌", f"CREAR: {expected}")
        return ("❌", f"Crear: {expected}")

    expected_str = str(expected) if expected is not None else ""

    # ── Coinciden exacto ──
    if str(actual) == expected_str:
        return ("✅", "Correcto. Se mantiene el valor adaptado actual.")

    # ── Pattern rule: startswith ──
    param_cfg = cfg.get("params", {}).get(param_name, {})
    rule = param_cfg.get("rule", "pattern")
    if rule == "pattern" and expected_str and str(actual).startswith(expected_str):
        return ("✅", "Correcto. Se mantiene el valor adaptado actual.")

    # ── No coinciden ──
    actual_lower = str(actual).lower().strip() if actual else ""
    if actual_lower in ("error page", "error_page", "error-page", "errorpage", ""):
        site_val = cfg.get("params", {}).get(param_name, {}).get("site_value", "")
        if site_val:
            return ("⚠️", f"Del sitio: '{site_val}' → '{expected_str}'")
    return ("⚠️", f"Cambiar '{actual}' → '{expected_str}'")

    if param_name == "client":
        return ("⚠️", f"Estandarizar a '{expected_str}' (con espacio) en AEM.")

    if param_name == "pageName":
        prefix = cfg.get("prefix", "")
        if not str(actual).startswith(prefix):
            return ("⚠️", f"Agregar prefijo '{prefix}': {str(actual)} → {expected_str}")
        return ("⚠️", f"Alinear nomenclatura: {str(actual)} → {expected_str}")

    if param_name == "siteSection":
        return ("⚠️", f"Cambiar sección '{str(actual)}' → '{expected_str}'")

    if param_name == "site":
        return ("⚠️", f"Unificar sitio '{str(actual)}' → '{expected_str}'")

    if param_name == "pageType":
        if not actual:
            return ("❌", f"CREAR: {expected_str}")
        return ("⚠️", f"Alinear: {str(actual)} → {expected_str}")

    return ("⚠️", f"Cambiar '{str(actual)}' → '{expected_str}'")


# ════════════════════════════════════════════
#  OUTPUT
# ════════════════════════════════════════════


def _fill_header(ws, headers: list[str]):
    """Escribe headers con estilo."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _write_row(ws, row: int, values: list, status_col: int | None = None):
    """Escribe una fila de datos con bordes y formato."""
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.font = FONT_DATA if c <= 2 or c == status_col else FONT_PARAM
        cell.alignment = ALIGN_WRAP
        cell.border = THIN_BORDER

    # Color status cell
    if status_col and values[status_col - 1]:
        status_val = str(values[status_col - 1]).strip()
        target = ws.cell(row, status_col)
        if status_val == "✅":
            target.fill = FILL_OK
        elif status_val == "⚠️":
            target.fill = FILL_WARN
        elif status_val == "❌":
            target.fill = FILL_FAIL


def _infer_page_key(url: str) -> str:
    """Infiere page_key desde una URL.

    Toma el último segmento significativo del path (salta códigos de idioma).
    Si el path es vacío o raíz, retorna 'home'.
    """
    if not url:
        return "unknown"
    path = urlparse(url).path.rstrip("/")
    if not path or path == "/":
        return "home"
    segments = [s for s in path.split("/")
                if s and s not in ("esp", "en", "content", "na",
                                    "es_pr", "en_pr", "es_mx", "en_mx")]
    if not segments:
        return "home"
    return segments[-1]

def _infer_expected(param_name: str, page_key: str, cfg: dict) -> str:
    """Valor esperado para page_key NO definido en expected.json.

    ANTES inventaba valores como `fpr:{page_key}` — eso GENERABA DATA FALSA
    que parecía legítima en la matriz. Ahora retorna un marcador CLARO
    para que el usuario agregue manualmente el valor real.

    Ver: _validate_expected_coverage() para detección temprana.
    """
    return "⚠️ SIN DEFINIR — agregar a expected.json"


def _sanitize_sheet_name(name: str) -> str:
    """Sanitiza nombre de hoja (max 31 chars, sin []:?*/)."""
    clean = re.sub(r'[\[\]:?*\\/]', '', name)
    return clean[:31]


def _build_page_name(mapping: dict) -> str:
    """Construye nombre descriptivo de página."""
    nombre = mapping.get("nombre", "")
    page_key = mapping.get("page_key", "")
    production_url = mapping.get("production_url", "")
    if nombre:
        return nombre
    if page_key:
        return page_key.replace("-", " ").title()
    # Fallback: último segmento de URL
    return production_url.rstrip("/").split("/")[-1].replace("-", " ").title()


def _build_note(param_name: str, actual_str: str | None, entorno: str) -> str:
    """Genera nota explicativa cuando el valor actual es un fallback del sitio.

    Solo agrega nota cuando el valor real del sitio es "error page" / "errorPage"
    — indica que la página no tiene digitalData configurado en PROD.
    """
    if actual_str is None:
        return ""
    actual_lower = str(actual_str).lower()
    is_error_page = any(
        kw in actual_lower
        for kw in ("error page", "errorpage", "error-page")
    )
    if not is_error_page:
        return ""
    if entorno == "produccion":
        return (
            "El sitio de producción no tiene digitalData configurado "
            "para esta URL. El valor 'error page' es un fallback del servidor."
        )
    return (
        "El sitio no tiene digitalData configurado para esta URL. "
        "El valor 'error page' es un fallback del servidor."
    )


# ════════════════════════════════════════════
#  HELPERS (extraídas de generate_matrix)
# ════════════════════════════════════════════


def _setup_workbook_sheet(
    existing_wb: openpyxl.Workbook | None,
    sheet_name: str | None,
    headers: list[str],
    simple: bool,
) -> tuple:
    """Crea o reusa un workbook y su hoja activa.

    Modo normal: crea workbook nuevo + hoja.
    Modo split: crea una hoja adicional en el workbook existente.

    Returns:
        (wb, ws) — el workbook y la hoja lista para escribir.
    """
    if existing_wb:
        wb = existing_wb
        safe = _sanitize_sheet_name(sheet_name or SHEET_NAME)
        if safe in wb.sheetnames:
            idx = 2
            while f"{safe[:28]}-{idx}" in wb.sheetnames:
                idx += 1
            safe = f"{safe[:28]}-{idx}"
        ws = wb.create_sheet(title=safe)
        print(f"  [{safe}] Generando matriz...")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        sheet_title = "Sheet1" if simple else _sanitize_sheet_name(SHEET_NAME)
        ws.title = sheet_title

    _fill_header(ws, headers)
    return wb, ws


def _set_column_widths(ws, headers: list[str], is_dual: bool, simple: bool):
    """Define anchos de columna según el modo."""
    cols = len(headers)
    if simple:
        widths = [10, 40, 28, 45, 45]
    else:
        widths = (
            [10, 40, 28, 45, 45, 12, 14, 14, 18, 55, 55]
            if is_dual
            else [10, 40, 28, 45, 45, 12, 18, 55, 55]
        )
    for i, w in enumerate(widths[:cols], 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w


def _get_page_data(
    mapping: dict,
    entorno: str,
    catalogo_docs: dict | None,
    catalogo: dict | None,
    catalogo_prod: dict | None,
    historial_preview: str | None,
    historial_prod: str | None,
    display_url: str | None,
) -> dict:
    """Obtiene datos actuales de una página desde 3 fuentes con prioridad.

    Orden de prioridad: catalogo_docs > catalogo pipeline > historial.
    Respeta la división preview/producción según entorno.

    Returns:
        dict con: preview_data, prod_data, preview_source, prod_source,
        url_auditada, page_name, page_key, has_preview, has_prod.
    """
    is_dual = entorno == "ambas"
    preview_url = mapping.get("preview_url", "")
    production_url = mapping.get("production_url", preview_url)
    page_key = mapping.get("page_key", "")
    page_name = _build_page_name(mapping)

    candidate_urls = [u for u in (preview_url, production_url) if u]
    actual_page_preview: dict = {}
    actual_page_prod: dict = {}
    preview_source = None
    prod_source = None

    # 1. Catalogo docs (match exacto)
    if is_dual or entorno == "preview":
        for cu in candidate_urls:
            cdata = _catalog_lookup(catalogo_docs or {}, cu, exact_only=True)
            if cdata:
                actual_page_preview = cdata
                preview_source = "catalogo_docs"
                break
    if is_dual or entorno == "produccion":
        for cu in candidate_urls:
            cdata = _catalog_lookup(catalogo_docs or {}, cu, exact_only=True)
            if cdata:
                actual_page_prod = cdata
                prod_source = "catalogo_docs"
                break

    # 2. Catalogo pipeline (fallback)
    if not actual_page_preview and (is_dual or entorno == "preview"):
        for cu in candidate_urls:
            cdata = _catalog_lookup(catalogo or {}, cu)
            if cdata:
                actual_page_preview = cdata
                preview_source = "catalogo"
                break
    if not actual_page_prod and (is_dual or entorno == "produccion"):
        for cu in candidate_urls:
            cdata = _catalog_lookup(catalogo_prod or catalogo or {}, cu)
            if cdata:
                actual_page_prod = cdata
                prod_source = "catalogo"
                break

    # 3. Historial (último fallback)
    if not actual_page_preview and (is_dual or entorno == "preview") and historial_preview:
        dd_preview = find_digitaldata_in_historial(historial_preview, candidate_urls)
        if dd_preview:
            actual_page_preview = dd_preview.get("page", {})
            preview_source = "historial"
    if not actual_page_prod and (is_dual or entorno == "produccion") and historial_prod:
        dd_prod = find_digitaldata_in_historial(historial_prod, candidate_urls)
        if dd_prod:
            actual_page_prod = dd_prod.get("page", {})
            prod_source = "historial"

    has_preview = bool(actual_page_preview)
    has_prod = bool(actual_page_prod)
    url_auditada = production_url or preview_url or (display_url or "")

    if has_preview or has_prod:
        src = preview_source or prod_source or ""
        print(f"  [+] {page_name} ({src})")
    else:
        print(f"  [!] Sin datos: {page_name}")

    return {
        "preview_data": actual_page_preview,
        "prod_data": actual_page_prod,
        "preview_source": preview_source,
        "prod_source": prod_source,
        "url_auditada": url_auditada,
        "page_name": page_name,
        "page_key": page_key,
        "has_preview": has_preview,
        "has_prod": has_prod,
    }


def _write_page_rows(
    ws, start_row: int, page_data: dict,
    market_cfg: dict, is_dual: bool, is_split: bool,
    first_param_row: bool, entorno: str,
    display_url: str | None = None,
    simple: bool = False,
) -> tuple:
    """Escribe las filas de parámetros para una página en la hoja.

    Returns:
        (next_row, row_count, warns, fails, first_param_row).
    """
    preview_data = page_data["preview_data"]
    prod_data = page_data["prod_data"]
    url_auditada = page_data["url_auditada"]
    page_name = page_data["page_name"]
    page_key = page_data["page_key"]

    if is_dual:
        actual_page_preview = preview_data
        actual_page_prod = prod_data
    else:
        actual_page = preview_data if entorno == "preview" else prod_data

    row = start_row
    row_count = 0
    warns = 0
    fails = 0
    fp_row = first_param_row

    for param in PARAMS_ORDER:
        param_cfg = market_cfg.get("params", {}).get(param, {})
        if not param_cfg and param != "hierarchy":
            continue

        if is_dual:
            # Modalidad dual: Preview vs Producción lado a lado
            actual_val_preview = actual_page_preview.get(param)
            actual_val_prod = actual_page_prod.get(param)

            actual_str_preview = str(actual_val_preview) if actual_val_preview is not None else None
            actual_str_prod = str(actual_val_prod) if actual_val_prod is not None else None

            status_preview, action_preview = evaluate_one(param, actual_str_preview, page_key, market_cfg)
            status_prod, action_prod = evaluate_one(param, actual_str_prod, page_key, market_cfg)

            expected_val, _, _ = resolve_expected(param, page_key, market_cfg)
            expected_str = str(expected_val) if expected_val is not None else "\u2014"

            note = _build_note(param, actual_str_prod, "produccion")
            note_preview = _build_note(param, actual_str_preview, "preview")
            full_note = note or note_preview or ""

            page_or_url = display_url if (is_split and fp_row) else ("" if is_split else page_name)
            url_row = url_auditada if (is_split and fp_row) else ("" if is_split else url_auditada)
            fp_row = False
            values = [
                url_row,
                page_or_url,
                param,
                actual_str_preview if actual_str_preview else "\u2014",
                actual_str_prod if actual_str_prod else "\u2014",
                expected_str,
                status_preview,
                status_prod,
                "Authoring AEM",
                action_preview if status_preview != "\u2705" else action_prod,
                full_note,
            ]
            _write_row(ws, row, values, status_col=7)
            if status_preview == "\u26a0\ufe0f" or status_prod == "\u26a0\ufe0f":
                warns += 1
            if status_preview == "\u274c" or status_prod == "\u274c":
                fails += 1
        else:
            # Modalidad single: un entorno
            actual_val = actual_page.get(param)
            actual_str = str(actual_val) if actual_val is not None else None

            # Si el scraper da "error page" pero sabemos el valor real
            # del sitio (site_value), usarlo en vez de "error page"
            if actual_str and actual_str.lower().strip() in ("error page", "error_page", "error-page", "errorpage", ""):
                site_val = market_cfg.get("params", {}).get(param, {}).get("site_value", "")
                if site_val:
                    actual_str = site_val
                else:
                    actual_str = "\u2014"

            status, action_text = evaluate_one(param, actual_str, page_key, market_cfg)

            expected_val, _, _ = resolve_expected(param, page_key, market_cfg)
            expected_str = str(expected_val) if expected_val is not None else "\u2014"

            note = _build_note(param, actual_str, entorno)
            page_or_url = display_url if (is_split and fp_row) else ("" if is_split else page_name)
            url_row = url_auditada if (is_split and fp_row) else ("" if is_split else url_auditada)
            fp_row = False
            if simple:
                values = [
                    url_row,
                    page_or_url,
                    param,
                    actual_str if actual_str else "\u2014",
                    expected_str,
                ]
                _write_row(ws, row, values)
            else:
                values = [
                    url_row,
                    page_or_url,
                    param,
                    actual_str if actual_str else "\u2014",
                    expected_str,
                    status,
                    "Authoring AEM",
                    action_text,
                    note,
                ]
                _write_row(ws, row, values, status_col=6)
                if status == "\u26a0\ufe0f":
                    warns += 1
                elif status == "\u274c":
                    fails += 1

        row += 1
        row_count += 1

    return row, row_count, warns, fails, fp_row


def _get_headers(is_dual: bool, simple: bool) -> tuple[list[str], int]:
    """Resuelve headers y cantidad de columnas según modo."""
    if simple:
        headers = COLS_SIMPLE
    else:
        headers = COLS_DUAL if is_dual else COLS_SINGLE
    return headers, len(headers)


def _write_orphan_rows(
    ws, start_row: int,
    historial_path: str | None, market_cfg: dict,
    entorno: str, is_dual: bool, mappings: list,
) -> tuple:
    """Procesa entradas del historial que no tienen mapping (huérfanas).

    Returns: (next_row, row_count).
    """
    if not historial_path:
        return start_row, 0

    index = _build_historial_index(historial_path)
    if not index:
        return start_row, 0

    # URLs del mapping ya cubiertas
    covered = set()
    for m in mappings:
        for u in (m.get("preview_url", ""), m.get("production_url", "")):
            if u:
                covered.add(u.lower().rstrip("/"))

    row = start_row
    row_count = 0
    orphan_count = 0

    for hist_url, data in sorted(index.items()):
        is_covered = any(cu in hist_url or hist_url in cu for cu in covered)
        if is_covered:
            continue
        page = data.get("page", {})
        has_data = any(page.get(p) for p in PARAMS_ORDER if page.get(p))
        if not has_data:
            continue

        orphan_count += 1
        page_name = hist_url.rstrip("/").split("/")[-1].replace("-", " ").title()
        print(f"  [+] (historial) {page_name}")

        for param in PARAMS_ORDER:
            param_cfg = market_cfg.get("params", {}).get(param, {})
            if not param_cfg and param != "hierarchy":
                continue
            actual_val = page.get(param)
            actual_str = str(actual_val) if actual_val is not None else None
            note = _build_note(param, actual_str, entorno)

            if is_dual:
                values = [
                    hist_url,
                    page_name + " [historial]",
                    param,
                    actual_str if actual_str else "\u2014",
                    "\u2014",
                    "\u2014",
                    "\U0001f4cb",
                    "\u2014",
                    "\u2014",
                    "Agregar a url-mapping.json y expected.json",
                    note,
                ]
            else:
                values = [
                    hist_url,
                    page_name + " [historial]",
                    param,
                    actual_str if actual_str else "\u2014",
                    "\u2014",
                    "\U0001f4cb",
                    "\u2014",
                    "Agregar a url-mapping.json y expected.json",
                    note,
                ]
            _write_row(ws, row, values)
            row += 1
            row_count += 1
        row += 1  # separador entre páginas

    if orphan_count:
        print(f"     ({orphan_count} páginas adicionales del historial sin mapping)")

    return row, row_count


def _finalize_sheet(
    ws, row: int, headers: list[str],
    market: str, entorno: str, extra: str,
    row_count: int, ok_count: int, warns: int, fails: int,
):
    """Escribe fila de firma, freeze panes y auto-filter."""
    cols = len(headers)

    # ── Fila de firma ──
    ws.cell(row, 1,
            f"Generado: {date.today().isoformat()} | "
            f"Mercado: {market} | Entorno: {entorno} | "
            f"Parámetros: {len(PARAMS_ORDER)} | "
            f"Páginas: {row_count // len(PARAMS_ORDER)}{extra}").font = Font(italic=True, color="999999")

    # Freeze + auto-filter
    ws.freeze_panes = "A2"
    last_col = openpyxl.utils.get_column_letter(cols)
    ws.auto_filter.ref = f"A1:{last_col}{row - 1}"

    print(f"     {row_count} filas")
    print(f"     \u2705 {ok_count}  \u26a0\ufe0f {warns}  \u274c {fails}")


def generate_matrix(market: str, entorno: str, historial_preview: str | None,
                    historial_prod: str | None, mappings: list,
                    expected_cfg: dict, output_path: str,
                    catalogo: dict | None = None,
                    catalogo_prod: dict | None = None,
                    catalogo_docs: dict | None = None,
                    skip_orphans: bool = False,
                    existing_wb: openpyxl.Workbook | None = None,
                    sheet_name: str | None = None,
                    display_url: str | None = None,
                    simple: bool = False):
    """Genera la matriz de validación Excel.

    Args:
        skip_orphans: Si True, omite la sección de entradas huérfanas del historial.
        existing_wb: Workbook existente para modo split (multi-sheet).
                     Cuando se provee, NO guarda el archivo (lo hace el caller).
        sheet_name: Nombre de la hoja dentro del workbook existente.
        display_url: URL a mostrar en la primera fila de la hoja (split mode).
    """
    market_cfg = expected_cfg.get("markets", {}).get(market, {})
    if not market_cfg:
        print(f"[ERR] Mercado '{market}' no encontrado en expected.json")
        sys.exit(1)

    is_dual = entorno == "ambas"
    is_split = existing_wb is not None
    headers, cols = _get_headers(is_dual, simple)

    # ── Setup workbook ──
    wb, ws = _setup_workbook_sheet(existing_wb, sheet_name, headers, simple)
    _set_column_widths(ws, headers, is_dual, simple)

    row = 2
    total_row_count = 0
    total_warns = 0
    total_fails = 0
    first_param_row = True

    for mapping in mappings:
        page_key = mapping.get("page_key", "")
        if not page_key:
            print(f"  [!] Saltando URL sin page_key: {mapping.get('preview_url', '')}")
            continue

        # ── Lookup datos de la página ──
        page_data = _get_page_data(
            mapping, entorno,
            catalogo_docs, catalogo, catalogo_prod,
            historial_preview, historial_prod,
            display_url,
        )

        # ── Escribir filas de parámetros ──
        row, rc, w, f, first_param_row = _write_page_rows(
            ws, row, page_data,
            market_cfg, is_dual, is_split,
            first_param_row, entorno,
            display_url=display_url,
            simple=simple,
        )
        total_row_count += rc
        total_warns += w
        total_fails += f

        # Fila separadora entre páginas
        row += 1

    # ── Orphans: URLs del historial sin mapping ──
    historial_path = historial_preview if entorno in ("preview", "ambas") else historial_prod
    if not simple and not skip_orphans and historial_path:
        orphan_row, orphan_count = _write_orphan_rows(
            ws, row, historial_path, market_cfg, entorno, is_dual, mappings,
        )
        row = orphan_row
        total_row_count += orphan_count

    ok_count = total_row_count - total_warns - total_fails

    if simple:
        wb.save(output_path)
        print(f"     {len(mappings)} páginas × {len(PARAMS_ORDER)} parámetros = {total_row_count} filas")
        print(f"     \u2705 {ok_count}  \u26a0\ufe0f {total_warns}  \u274c {total_fails}")
        return

    # ── Fila de firma + freeze ──
    row += 1
    extra = "" if not skip_orphans else " (split)"
    _finalize_sheet(ws, row, headers, market, entorno, extra,
                    total_row_count, ok_count, total_warns, total_fails)

    if is_split:
        return

    wb.save(output_path)
    print(f"\n[OK] Matriz generada: {output_path}")
    print(f"     {len(mappings)} páginas \u00d7 {len(PARAMS_ORDER)} par\u00e1metros = {total_row_count} filas")


def load_mappings(mapping_path: str) -> list:
    """Carga y filtra mappings del url-mapping.json."""
    data = load_json(mapping_path)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return data.get("mappings", [])
    return []


def load_catalogo(catalogo_path: str) -> dict:
    """Carga un catalogo-migracion.xlsx y retorna {url: {param: valor_actual}}.

    Columnas esperadas:
      1: URL de Produccion (puede estar vacia en filas de continuation)
      3: Parametro (digitalData.page) — o "📦 digitalData.page (completo)" con JSON
      4: Valor en Produccion (Actual)
    """
    if not os.path.exists(catalogo_path):
        return {}
    wb = openpyxl.load_workbook(catalogo_path, data_only=True)
    ws = wb.active
    catalog: dict[str, dict[str, str]] = {}
    current_url: str | None = None

    for row in range(2, ws.max_row + 1):
        url = ws.cell(row, 1).value
        param = ws.cell(row, 3).value
        actual = ws.cell(row, 4).value

        # Track current URL (continuation rows have None)
        if url:
            current_url = str(url).strip().lower().rstrip("/")
        if not current_url or not param:
            continue

        param_s = str(param).strip()
        actual_s = str(actual).strip() if actual is not None else ""

        if current_url not in catalog:
            catalog[current_url] = {}

        # "📦 (completo)" row: parse embedded JSON
        if "completo" in param_s.lower():
            if actual_s.startswith("{"):
                try:
                    parsed = json.loads(actual_s)
                    if isinstance(parsed, dict):
                        for k, v in parsed.items():
                            if v is not None and str(v).strip() not in ("", "\u2014", "—"):
                                catalog[current_url][k] = str(v).strip()
                except json.JSONDecodeError:
                    pass
            continue

        # Individual param row
        if param_s.startswith("page."):
            param_s = param_s[5:]
        if actual_s and actual_s not in ("", "\u2014", "—", "{}") and "no existe" not in actual_s.lower():
            catalog[current_url][param_s] = actual_s

    wb.close()
    return catalog


def _catalog_lookup(catalog: dict, url: str, exact_only: bool = False) -> dict | None:
    """Busca una URL en el catalogo.

    Args:
        catalog: dict {url_normalizada: {param: valor}}
        url: URL a buscar
        exact_only: True = solo match exacto (docs catalog),
                    False = path-prefix flexible (pipeline catalogo)

    Previene falsos positivos donde '/' matchea TODO o '/esp/'
    matchea '/especificaciones/' por substring.
    """
    from urllib.parse import urlparse

    def _paths_match(short: str, long: str) -> bool:
        """True si short es prefijo de long en frontera de path component."""
        if short == long:
            return True
        if short == "/":
            return False  # root solo matchea root
        if long.startswith(short) and len(long) > len(short) and long[len(short)] == "/":
            return True
        if short.startswith(long) and len(short) > len(long) and short[len(long)] == "/":
            return True
        return False

    url_norm = url.lower().rstrip("/")
    if url_norm in catalog:
        return catalog[url_norm]

    if exact_only:
        return None

    parsed = urlparse(url_norm)
    for cat_url, data in catalog.items():
        cat_norm = cat_url.lower().rstrip("/")
        cat_parsed = urlparse(cat_norm)
        # Mismo netloc y scheme
        if (parsed.scheme, parsed.netloc) != (cat_parsed.scheme, cat_parsed.netloc):
            continue
        # Path matching por componente (no substring)
        if _paths_match(parsed.path or "/", cat_parsed.path or "/"):
            return data
    return None


def _safe_filename(page_key: str, seen: dict[str, int]) -> str:
    """Genera un nombre de archivo seguro a partir de un page_key.

    Si el page_key ya fue usado, agrega un sufijo numérico (-1, -2...).
    Ejemplo: 'ev' → 'ev', segunda ocurrencia → 'ev-1'
    """
    safe = re.sub(r'[^a-zA-Z0-9_-]', '_', page_key).strip('_-') or "pagina"
    rank = seen.get(page_key, 0)
    seen[page_key] = rank + 1
    if rank > 0:
        safe = f"{safe}-{rank}"
    return safe


def _safe(val) -> str:
    """Convierte a string seguro para print."""
    s = str(val) if val is not None else ""
    return s.encode('ascii', errors='replace').decode('ascii')


def main():
    parser = argparse.ArgumentParser(
        description="Genera matriz de validación estilo v2 (Ford PR)"
    )
    parser.add_argument("--market", default="PR",
                        help="Código de mercado (PR, MX...)")
    parser.add_argument("--entorno", default="produccion",
                        choices=["preview", "produccion", "ambas"],
                        help="Entorno a analizar")
    parser.add_argument("--historial-preview",
                        help="Path a historial preview (default: {market}/preview/historial.xlsx)")
    parser.add_argument("--historial-produccion",
                        help="Path a historial produccion (default: {market}/produccion/historial.xlsx)")
    parser.add_argument("--mapping", default="data/url-mapping.json",
                        help="JSON mapeo preview→producción")
    parser.add_argument("--expected", default="data/expected.json",
                        help="JSON con valores esperados por mercado")
    parser.add_argument("--catalogo",
                        help="Path a catalogo-migracion.xlsx (default: auto-detect)")
    parser.add_argument("--catalogo-docs",
                        help="Path al catalogo de docs/ con valores reales de produccion "
                             "(ej: docs/ford-pr-catalogo-valores-pre-preview.xlsx)")
    parser.add_argument("--output", default=None,
                        help="Ruta de salida (default: {market}/matriz-validacion-{entorno}.xlsx)")
    parser.add_argument("--split", action="store_true",
                        help="Genera un archivo por página en {market}/split/")
    parser.add_argument("--simple", action="store_true",
                        help="Formato simple: 5 columnas, Sheet1, sin estilos (formato ejemplo GTBEMEAPUB)")
    args = parser.parse_args()

    # ── Resolver paths ──
    market = args.market.upper()
    base_dir = os.getcwd()

    if args.historial_preview:
        h_preview = args.historial_preview
    else:
        h_preview = os.path.join(base_dir, market, "preview", "historial.xlsx")

    if args.historial_produccion:
        h_prod = args.historial_produccion
    else:
        h_prod = os.path.join(base_dir, market, "produccion", "historial.xlsx")

    mapping_path = args.mapping if os.path.isabs(args.mapping) else os.path.join(base_dir, args.mapping)
    expected_path = args.expected if os.path.isabs(args.expected) else os.path.join(base_dir, args.expected)

    # ── Validar archivos según entorno ──
    # Verificar si hay fuentes alternativas de datos antes de exigir historial
    has_catalogo_docs = bool(args.catalogo_docs) or os.path.exists("docs/ford-pr-catalogo-valores-pre-preview.xlsx")
    has_pipeline_catalogo = os.path.exists(
        os.path.join(base_dir, market, args.entorno, "catalogo-migracion.xlsx")
    ) if not args.catalogo else os.path.exists(args.catalogo)

    if args.entorno in ("preview", "ambas") and not os.path.exists(h_preview):
        print(f"[WARN] No se encuentra historial preview: {h_preview}")
        if args.entorno == "preview" and not has_catalogo_docs and not has_pipeline_catalogo:
            print("  Ejecuta primero una auditoría (opción 2) con --entorno preview")
            sys.exit(1)

    if args.entorno in ("produccion", "ambas") and not os.path.exists(h_prod):
        print(f"[WARN] No se encuentra historial produccion: {h_prod}")
        if args.entorno == "produccion" and not has_catalogo_docs and not has_pipeline_catalogo:
            print("  Ejecuta primero una auditoría (opción 2) con --entorno produccion")
            sys.exit(1)

    for fpath, fname in [(mapping_path, "mapping"), (expected_path, "expected")]:
        if not os.path.exists(fpath):
            print(f"[ERR] Archivo {fname} no encontrado: {fpath}")
            sys.exit(1)

    # ── Output ──
    if args.output is None:
        output_name = f"matriz-validacion-{args.entorno}.xlsx"
        output_path = os.path.join(base_dir, market, output_name)
    else:
        output_path = args.output
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    # ── Cargar datos ──
    expected_cfg = load_json(expected_path)
    mappings = load_mappings(mapping_path)

    # ── Pre-validación: todas las page_keys existen en expected ──
    if mappings and market in expected_cfg.get("markets", {}):
        mercado_exp = expected_cfg["markets"][market].get("expected", {})
        if mercado_exp:
            ausentes = [m["page_key"] for m in mappings
                        if m.get("page_key") and m["page_key"] not in mercado_exp]
            if ausentes:
                print(f"\n{'='*60}")
                print(f"  [WARN] {len(ausentes)} page_keys SIN definir en expected.json:")
                for pk in sorted(ausentes):
                    print(f"    ⚠️  {pk}")
                print(f"  Se marcarán como '⚠️ SIN DEFINIR' en la matriz")
                print(f"{'='*60}\n")

    if not mappings:
        # Auto-construir mappings desde urls.json
        urls_path = os.path.join(base_dir, "data", "urls.json")
        if os.path.exists(urls_path):
            with open(urls_path, encoding="utf-8") as f:
                urls_data = json.load(f)
            for entry in urls_data:
                url_market = entry.get("market", "").upper()
                if url_market != market:
                    continue
                url = entry.get("url", "")
                if not url:
                    continue
                pk = entry.get("page_key") or _infer_page_key(url)
                mappings.append({
                    "page_key": pk,
                    "production_url": url,
                    "preview_url": url,
                    "nombre": pk.replace("-", " ").title(),
                })
            if mappings:
                print(f"[OK] Mappings auto-generados desde urls.json ({len(mappings)} URLs)")
            else:
                print(f"[ERR] No hay URLs para mercado '{market}' en urls.json")
                sys.exit(1)
        else:
            print(f"[ERR] No se encontraron mappings en {mapping_path}")
            print(f"      Y no existe {urls_path} para auto-generar")
            sys.exit(1)

    if market not in expected_cfg.get("markets", {}):
        print(f"[ERR] Mercado '{market}' no encontrado en {expected_path}")
        print(f"  Mercados disponibles: {', '.join(expected_cfg.get('markets', {}).keys())}")
        sys.exit(1)

    # ── Cargar catalogo (auto-detect) ──
    default_catalogo = os.path.join(base_dir, market, "produccion", "catalogo-migracion.xlsx")
    catalogo_path = args.catalogo if args.catalogo else default_catalogo
    catalogo_prod_path = os.path.join(base_dir, market, "produccion", "catalogo-migracion.xlsx")
    catalogo_preview_path = os.path.join(base_dir, market, "preview", "catalogo-migracion.xlsx")

    catalogo = load_catalogo(catalogo_path)
    catalogo_prod = load_catalogo(catalogo_prod_path)

    if catalogo:
        print(f"[OK] Catalogo cargado: {catalogo_path} ({len(catalogo)} URLs)")

    # ── Cargar catalogo docs (valores reales de produccion) ──
    catalogo_docs = None
    if args.catalogo_docs:
        catalogo_docs = load_catalogo(args.catalogo_docs)
        if catalogo_docs:
            print(f"[OK] Catalogo docs cargado: {args.catalogo_docs} ({len(catalogo_docs)} URLs)")
    elif os.path.exists("docs/ford-pr-catalogo-valores-pre-preview.xlsx"):
        catalogo_docs = load_catalogo("docs/ford-pr-catalogo-valores-pre-preview.xlsx")
        if catalogo_docs:
            print(f"[OK] Catalogo docs (auto): docs/ford-pr-catalogo-valores-pre-preview.xlsx ({len(catalogo_docs)} URLs)")

    # ── Generar ──
    h_preview_actual = h_preview if os.path.exists(h_preview) else None
    h_prod_actual = h_prod if os.path.exists(h_prod) else None

    if args.simple:
        # Modo simple: UN archivo individual por URL (formato ejemplos GTBEMEAPUB)
        simple_dir = os.path.join(base_dir, market, "simple")
        os.makedirs(simple_dir, exist_ok=True)
        for mapping in mappings:
            display_url = mapping.get("production_url") or mapping.get("preview_url", "")
            page_key = mapping.get("page_key", "")
            if not page_key:
                print(f"  [!] Saltando URL sin page_key")
                continue
            # Si se especifico --output y solo hay 1 mapping, usarlo directamente
            if args.output and len(mappings) == 1:
                simple_path = args.output
            else:
                page_name = _build_page_name(mapping).replace(" ", "")
                simple_path = os.path.join(simple_dir, f"{page_name}.xlsx")
            generate_matrix(
                market=market,
                entorno=args.entorno,
                historial_preview=h_preview_actual,
                historial_prod=h_prod_actual,
                mappings=[mapping],
                expected_cfg=expected_cfg,
                output_path=simple_path,
                catalogo=catalogo,
                catalogo_prod=catalogo_prod,
                catalogo_docs=catalogo_docs,
                skip_orphans=True,
                simple=True,
                display_url=display_url,
            )
    elif args.split:
        # Modo split: UN archivo, UNA hoja (sheet) por URL
        split_dir = os.path.join(base_dir, market, "split")
        os.makedirs(split_dir, exist_ok=True)
        entorno_name = args.entorno
        split_path = os.path.join(split_dir, f"matriz-split-{entorno_name}.xlsx")
        print(f"\n[Split] Generando matriz multi-sheet: {split_path}")
        split_wb = openpyxl.Workbook()
        # Eliminar hoja por defecto (la recreamos con la 1ra URL)
        split_wb.remove(split_wb.active)

        total_ok = 0
        total_warn = 0
        total_fail = 0
        for mapping in mappings:
            sheet_name = _build_page_name(mapping)
            display_url = mapping.get("production_url") or mapping.get("preview_url", "")
            page_key = mapping.get("page_key", "")
            if not page_key:
                print(f"  [!] Saltando URL sin page_key")
                continue
            generate_matrix(
                market=market,
                entorno=args.entorno,
                historial_preview=h_preview_actual,
                historial_prod=h_prod_actual,
                mappings=[mapping],
                expected_cfg=expected_cfg,
                output_path=split_path,
                catalogo=catalogo,
                catalogo_prod=catalogo_prod,
                catalogo_docs=catalogo_docs,
                skip_orphans=True,
                existing_wb=split_wb,
                sheet_name=sheet_name,
                display_url=display_url,
            )

        # Guardar el workbook completo
        split_wb.save(split_path)
        print(f"\n[OK] Split generado: {split_path}")
        print(f"     {len(mappings)} hojas (1 por URL)")
        # No podemos sumar los contadores porque generate_matrix
        # los imprime por hoja individual. El resumen está arriba.
    else:
        generate_matrix(
            market=market,
            entorno=args.entorno,
            historial_preview=h_preview_actual,
            historial_prod=h_prod_actual,
            mappings=mappings,
            expected_cfg=expected_cfg,
            output_path=output_path,
            catalogo=catalogo,
            catalogo_prod=catalogo_prod,
            catalogo_docs=catalogo_docs,
        )


if __name__ == "__main__":
    main()

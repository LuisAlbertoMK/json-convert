"""
extract_browser.py — CLI entry point: navega URLs, extrae AA beacons, escribe Excel.

Delegates logic to json_convert package modules.
Uso: python extract_browser.py [--urls urls.json] [--workers N] ...
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import os
import re
import signal
import subprocess
import sys
import time
from datetime import datetime, timedelta

import openpyxl
from playwright.async_api import Page, async_playwright
from playwright.async_api import TimeoutError as PwTimeout

from json_convert import (
    AA_DOMAINS,
    INPUT_FILE,
    SAVE_EVERY_N,
    apply_data_fills,
    classify_errors,
    compute_score,
    parse_aa_beacon,
    sanitize_url_for_log,
    save_workbook,
    setup_multisheet,
    split_aa_workbooks,
    update_control,
    update_vars_sheet,
    validate_sheet,
)
from json_convert.browser import _shutdown_flag, process_url
from json_convert.cache import UrlCache
from json_convert.excel import _auto_row_height
from json_convert.pipeline import (
    run_pipeline,
)


def _request_shutdown(signum: int | None = None, frame: object | None = None) -> None:
    global _shutdown_flag
    if not _shutdown_flag:
        _shutdown_flag = True
        logging.warning("Señal %s recibida. Terminando gracefulmente...", signum)




# ── BEACON_REGEX (compilado acá por conveniencia del entry point) ──
BEACON_REGEX = re.compile(
    r'https?://[^"\'\s]*(?:'
    + "|".join(re.escape(d) for d in AA_DOMAINS)
    + r')[^"\'\s]*',
    re.IGNORECASE,
)

# ═══════════════════════════════════════════════════════════════════════════
# MAIN (async)
# ═══════════════════════════════════════════════════════════════════════════

async def amain() -> None:
    parser = _build_parser()
    args = parser.parse_args()

    log_path = _setup_logging(args.verbose, args.log_file)
    print(f"\n  Log: {log_path}")

    # ── Caché de navegación ──
    cache = UrlCache(ttl=args.cache_ttl) if not args.no_cache else None

    if args.clear_cache and cache:
        cleared = cache.clear()
        print(f"\n  Caché limpiada: {cleared} archivo(s) eliminado(s)")
        return

    # ── Fuente de URLs ──
    urls, url_source = _resolve_urls(args)
    output_path = args.output or _resolve_output(url_source, args.market, args.entorno)

    # Mostrar conteo antes de empezar
    market_label = args.market.upper() if args.market else "TODOS"
    env_label = {"preview": "PREVIEW", "produccion": "PRODUCCION", "ambas": "TODOS"}.get(args.entorno, args.entorno)
    print(f"\n  Auditando {len(urls)} URLs [{env_label}] para mercado {market_label}...\n")

    # Crear directorio de output si no existe (ej: PR/)
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # ── Excel (crear ANTES del pipeline) ──
    wb, ws, audit_date, _ = setup_multisheet(output_path, url_source, resume=False)

    # ── Playwright ──
    async with async_playwright() as pw:
        browser_type = args.browser or "chromium"
        launch_kwargs: dict = dict(
            headless=not args.headed,
            proxy={"server": args.proxy} if args.proxy else None,
        )

        if browser_type == "firefox":
            print(f"  Usando Firefox ({browser_type})")
            browser = await pw.firefox.launch(
                firefox_user_prefs={
                    # Desactivar tracking protection para capturar beacons AA
                    "privacy.trackingprotection.enabled": False,
                    "privacy.trackingprotection.pbmode.enabled": False,
                    "privacy.trackingprotection.emailtracking.enabled": False,
                },
                **launch_kwargs,
            )
        else:
            launch_kwargs["args"] = ["--disable-blink-features=AutomationControlled"]
            try:
                browser = await pw.chromium.launch(channel="chrome", **launch_kwargs)
            except Exception:
                logging.info("Chrome no disponible, usando Chromium bundled")
                browser = await pw.chromium.launch(**launch_kwargs)

        # ── Context ──
        ua_firefox = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:150.0) "
            "Gecko/20100101 Firefox/150.0"
        )
        ua_chromium = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
        )
        context_kwargs: dict = dict(
            user_agent=ua_firefox if browser_type == "firefox" else ua_chromium,
            locale="es-PR" if "es" in url_source else "en-US",
            viewport={"width": 1920, "height": 1080},
            ignore_https_errors=True if args.proxy else False,
        )
        # NO compartir contexto — cada página crea el suyo para evitar
        # "Target page, context or browser has been closed" en cascada.

        # ── Crear process_func que cada worker usa para procesar una URL ──
        semaphore = asyncio.Semaphore(args.workers or 3)

        async def _process_one(row: int, url: str) -> dict:
            """Crea page, procesa URL, captura beacons, cierra page.
            Usa caché si está disponible (--no-cache para desactivar)."""
            # Verificar caché antes de navegar
            if cache is not None:
                cached = cache.get(url)
                if cached is not None:
                    # Actualizar row + elapsed_s para reflejar la corrida actual
                    cached = dict(cached)
                    cached["row"] = row
                    cached["_from_cache"] = True
                    # Si el cache tiene digitaldata_manual, preservarlo
                    if "digitaldata_manual" not in cached:
                        cached["digitaldata_manual"] = cached.get("digitaldata")
                    # Parsear beacons del caché si existen pero no se parsearon antes
                    _raw = cached.get("raw_beacons", [])
                    if _raw and not cached.get("aa_parsed"):
                        _parsed_aa = []
                        _extra_aa = []
                        for _b in _raw[:6]:
                            try:
                                _p = parse_aa_beacon(_b, cached.get("title", ""))
                                if _p.get("hit", {}).get("type") in ("pageView", None):
                                    _parsed_aa.append(_p)
                                else:
                                    _extra_aa.append(_p)
                            except Exception:
                                continue
                        if _parsed_aa:
                            cached["aa_parsed"] = _parsed_aa[0]
                        if _extra_aa:
                            cached["extra_beacons"] = _extra_aa
                    logging.info("Cache hit: %s", url[:80])
                    return cached

            async with semaphore:
                # Contexto propio por página para evitar "Target closed" en cascada.
                ctx = await browser.new_context(**context_kwargs)
                ctx.set_default_timeout(args.timeout * 1000 if args.timeout else 60000)
                if browser_type == "firefox":
                    await ctx.add_init_script("""
                        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                    """)
                else:
                    await ctx.add_init_script("""
                        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                        Object.defineProperty(navigator, 'plugins', { get: () => [1,2,3,4,5] });
                        Object.defineProperty(navigator, 'languages', { get: () => ['es-PR', 'en-US'] });
                        window.chrome = { runtime: {} };
                    """)
                page = await ctx.new_page()

                try:
                    beacons: list[str] = []

                    async def _capture(response: object) -> None:
                        url_lower = response.url.lower()
                        if any(d in url_lower for d in AA_DOMAINS):
                            try:
                                await response.body()
                                beacons.append(response.url)
                            except Exception:
                                pass

                    page.on("response", _capture)
                    result = await process_url(
                        page, row, url,
                        wait_after=args.wait_after or 2,
                        timeout_ms=(args.timeout or 35) * 1000,
                        max_retry=args.max_retry or 1,
                    )
                    result["raw_beacons"] = beacons
                    # Parsear beacons capturados → aa_parsed estructurado
                    parsed_aa = []
                    extra_aa = []
                    for _b in beacons[:6]:
                        try:
                            _p = parse_aa_beacon(_b, result.get("title", ""))
                            if _p.get("hit", {}).get("type") in ("pageView", None):
                                parsed_aa.append(_p)
                            else:
                                extra_aa.append(_p)
                        except Exception:
                            continue
                    if parsed_aa:
                        result["aa_parsed"] = parsed_aa[0]
                    if extra_aa:
                        result["extra_beacons"] = extra_aa
                    # Guardar en caché (sin row específico ni elapsed)
                    if cache is not None:
                        cache.set(url, result)
                    return result
                finally:
                    await page.close()
                    try:
                        await ctx.close()
                    except Exception as e:
                        logging.debug("Error cerrando contexto de página: %s", e)

        # ── Pipeline (procesa URLs + escribe Excel) ──
        results, _errors_detail, metrics = await run_pipeline(
            _process_one, urls,
            workers=args.workers or 3,
            ws=ws,
            output_path=output_path,
            show_progress=bool(args.progress),
        )

        # ── Post-pipeline ──
        _auto_row_height(ws)
        apply_data_fills(ws)

        update_control(
            wb, audit_date, url_source,
            metrics["total"], metrics["ok_aa"], metrics["ok_dd"],
            metrics["errors"], metrics["retries"],
            compute_score(metrics), metrics.get("total_time", 0), args.workers or 3,
        )

        aa_rows = [(r["row"], r["aa_parsed"]) for r in results if r.get("aa_parsed")]
        if aa_rows:
            update_vars_sheet(wb, audit_date, aa_rows)

        out = save_workbook(wb, output_path)

        split = getattr(args, "split_aa", False)
        if split and audit_date:
            output_dir = os.path.dirname(output_path) or "."
            split_aa_workbooks(wb, audit_date, output_dir)

        wb.close()
        logging.info("Guardado: %s", out)

        # Limpieza AA en Excel (post-procesamiento)
        _run_aa_cleanup(output_path)

        score = _display_metrics(metrics, args)
        await browser.close()


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Extrae Adobe Analytics de URLs Ford")
    p.add_argument("--urls", help="Archivo JSON con URLs")
    p.add_argument("--input", default=INPUT_FILE, help="Excel de entrada")
    p.add_argument("--output", help="Excel de salida")
    p.add_argument("--market", help="Filtrar por mercado en urls.json")
    p.add_argument("--entorno", default="preview", choices=("preview", "produccion", "ambas"),
                   help="Entorno a auditar: preview (default), produccion, o ambas")
    p.add_argument("--row", type=int, help="Fila inicial en Excel clasico")
    p.add_argument("--workers", type=int, default=3, help="Concurrencia")
    p.add_argument("--timeout", type=int, default=35, help="Timeout por URL (s)")
    p.add_argument("--wait-after", type=int, default=2, help="Espera post-carga (s) — smart wait dinámico")
    p.add_argument("--max-retry", type=int, default=1, help="Reintentos por URL")
    p.add_argument("--headless", action="store_true", help="Forzar headless")
    p.add_argument("--headed", action="store_true", help="Forzar headed (con UI)")
    p.add_argument("--proxy", help="Proxy HTTP (ej: http://proxy:8080)")
    p.add_argument("--resume", action="store_true", help="Reanudar corrida existente")
    p.add_argument("--split-aa", action="store_true", help="Dividir con/sin AA")
    p.add_argument("--progress", action="store_true", help="Mostrar barra de progreso")
    p.add_argument("--log-file", help="Archivo de log")
    p.add_argument("--verbose", action="store_true", help="Logging debug")
    p.add_argument("--run-clean", action="store_true", help="Forzar clean run")
    p.add_argument("--no-cache", action="store_true", help="Desactivar caché de navegación")
    p.add_argument("--cache-ttl", type=int, default=86400,
                   help="TTL de caché en segundos (default: 86400 = 24h)")
    p.add_argument("--clear-cache", action="store_true", help="Limpiar caché y salir")
    p.add_argument("--browser", default="firefox", choices=("chromium", "firefox"),
                   help="Browser engine: firefox (default, evita ERR_ABORTED) o chromium (para produccion con Chrome real)")
    return p


def _setup_logging(verbose: bool, log_file: str | None = None) -> str:
    """Configura logging a consola + archivo.

    Si no se pasa log_file explícito, genera uno automático en logs/
    con timestamp. Retorna la ruta del archivo de log.
    """
    level = logging.DEBUG if verbose else logging.INFO
    log_path = log_file or _auto_log_path()
    os.makedirs(os.path.dirname(log_path) or ".", exist_ok=True)

    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_path, encoding="utf-8"),
        ],
    )
    # Forzar stdout a UTF-8 para evitar UnicodeEncodeError en Windows
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    return log_path


def _auto_log_path() -> str:
    """Genera ruta tipo logs/audit-2026-06-18_171500.log."""
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return os.path.join("logs", f"audit-{ts}.log")


def _resolve_urls(args: argparse.Namespace) -> tuple[list[tuple[int, str]], str]:
    """Determina fuente de URLs: JSON o Excel clasico."""
    if args.urls:
        import json
        with open(args.urls, encoding="utf-8") as f:
            all_entries = json.load(f)
        market = args.market
        env = args.entorno or "preview"
        if env == "ambas":
            if market:
                entries = [e for e in all_entries
                           if e.get("market", "").upper() == market.upper()]
            else:
                entries = all_entries
        else:
            if market:
                entries = [e for e in all_entries
                           if e.get("market", "").upper() == market.upper()
                           and e.get("entorno", "preview") == env]
            else:
                entries = [e for e in all_entries if e.get("entorno", "preview") == env]
        urls = [(i + 2, e["url"]) for i, e in enumerate(entries)]
        source = args.market or "multi"
        return urls, source

    # Modo clasico: Excel plano
    input_path = args.input or INPUT_FILE
    if not os.path.exists(input_path):
        print(f"[ERR] No se encuentra {input_path}")
        sys.exit(1)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    errs = validate_sheet(ws)
    if errs:
        for e in errs:
            print(f"[ERR] {e}")
        sys.exit(1)
    row_start = args.row or 2
    urls = [(r, str(ws.cell(r, 2).value or "").strip())
            for r in range(row_start, ws.max_row + 1) if ws.cell(r, 2).value]
    wb.close()
    source = "clasico"
    return urls, source


def _resolve_output(url_source: str, market: str | None = None,
                    entorno: str | None = None) -> str:
    """Determina nombre de output segun fuente y entorno."""
    if url_source == "clasico":
        return "resultado.xlsx"
    if market:
        market_dir = market.upper()
        if entorno and entorno not in ("ambas", None):
            return os.path.join(market_dir, entorno, "historial.xlsx")
        return os.path.join(market_dir, "historial.xlsx")
    return "historial.xlsx"




_ACCEPTANCE_TARGETS: dict[str, int] = {
    "PR": 80,
    "MX": 90,
}


def _resolve_target(market: str | None) -> int:
    """Retorna el target de aceptacion para el mercado dado."""
    if market and market.upper() in _ACCEPTANCE_TARGETS:
        return _ACCEPTANCE_TARGETS[market.upper()]
    return 80


def _display_metrics(metrics: dict, args: argparse.Namespace) -> int:
    """Muestra metricas finales en consola. Puro print — sin side-effects."""
    success_rate = (metrics["ok_aa"] / max(metrics["total"], 1)) * 100
    dd_rate = (metrics["ok_dd"] / max(metrics["total"], 1)) * 100
    avg_time = sum(metrics["times"]) / max(len(metrics["times"]), 1)
    beacons_per_url = metrics["total_beacons"] / max(metrics["total"], 1)
    total_time = metrics.get("total_time", 0)
    score = compute_score(metrics)
    market_label = (args.market or "GLOBAL").upper()
    target = _resolve_target(args.market)
    passed = score >= target

    print(f"""
{"=" * 55}
  METRICAS Y SCORE — [{market_label}]
{"=" * 55}
  Config:            {args.workers} worker(s) concurrente(s)
  URLs procesadas:   {metrics['ok_aa']}/{metrics['total']}
  AA capturados:     {success_rate:.0f}%
  digitaldata:       {dd_rate:.0f}%
  Beacons totales:   {metrics['total_beacons']} ({beacons_per_url:.1f}/URL)
  Reintentos:        {metrics['retries']}
  Errores:           {metrics['errors']}
{"-" * 55}
  Tiempo total:      {timedelta(seconds=int(total_time))}
  Promedio/URL:      {avg_time:.1f}s
  Guardados incr.:   cada {SAVE_EVERY_N} URLs
{"-" * 55}
  SCORE [{market_label}]:  {score}/100
  TARGET ACEPTACION:   {target}/100  {"✅ PASA" if passed else "❌ NO PASA"}
{"-" * 55}""")

    if not passed:
        print(f"  ⚠  Score {score} no alcanza el target {target} para {market_label}.")
        if success_rate < 80:
            print("    * Baja tasa de captura AA — verificar conectividad / bloqueos")
        if dd_rate < 70:
            print("    * Baja tasa de digitalData — posible cambio en data layer")
        if avg_time > 20:
            print("    * Promedio alto por URL — considerar aumentar workers o reducir timeout")

    if metrics.get("errores_detalle"):
        print(f"  DETALLE ERRORES ({len(metrics['errores_detalle'])}):")
        categories = classify_errors(metrics["errores_detalle"])
        for cat, rows in categories.items():
            rows_str = ", ".join(str(r) for r in rows)
            print(f"    * {cat}: Fila(s) {rows_str}")

    return score


def _run_aa_cleanup(output_path: str) -> None:
    """Ejecuta extract_aa.py post-procesamiento sobre el Excel generado."""
    if not os.path.exists(output_path):
        return
    clean_script = os.path.join(os.path.dirname(__file__) or ".", "extract_aa.py")
    cmd = [sys.executable, clean_script, "--input", output_path]
    r = subprocess.run(cmd, capture_output=True, text=True)
    print(r.stdout)


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    try:
        signal.signal(signal.SIGINT, _request_shutdown)
        signal.signal(signal.SIGTERM, _request_shutdown)
    except (ValueError, AttributeError):
        pass
    try:
        asyncio.run(amain())
    except KeyboardInterrupt:
        _request_shutdown()
        logging.warning("Interrumpido por usuario.")
        sys.exit(130)


if __name__ == "__main__":
    main()

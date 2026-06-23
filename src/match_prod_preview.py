#!/usr/bin/env python3
"""
match_prod_preview.py — Compara digitalData entre US Expected, Preview y Producción.

Modos de operación:
  1. 3‑vías (--mode 3way): US Expected vs Preview (real) vs Producción (real)
     Requiere: {market}/preview/historial.xlsx + {market}/produccion/historial.xlsx
  2. 2‑way  (--mode 2way): Prometido (Preview/Expected) vs Entregado (Producción)
     Modo clásico original.

Uso:
  # 3‑vías (nuevo)
  python match_prod_preview.py --market PR --mode 3way

  # 2‑way con preview real
  python match_prod_preview.py --market PR --mode 2way --preview PR/preview/historial.xlsx

  # 2‑way usando expected.json como prometido (fallback)
  python match_prod_preview.py --market PR --mode 2way

  # Auto-detect: 3way si existen ambos historiales, sino 2way
  python match_prod_preview.py --market PR

Output: {market}/match/match-3way.{xlsx,md,html} o match-prod-vs-preview.{xlsx,md,html}
"""

import argparse
import json
import os
import sys
from datetime import date

# Force UTF-8 stdout for Windows cp1252 terminal (safe in Python >=3.7)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import PatternFill

from json_convert.styles import (
    ALIGN_CENTER,
    ALIGN_WRAP,
    FILL_FAIL,
    FILL_HEADER,
    FILL_OK,
    FILL_WARN,
    FONT_DATA,
    FONT_HEADER,
    FONT_PARAM,
    THIN_BORDER,
)
from json_convert.utils import load_json

# ── Constantes ──
PARAMS_ORDER = ["pageName", "siteSection", "pageNameNoVehicle",
                "client", "site", "variantName", "pageType"]

# 2‑way (original)
MATCH_COLS = [
    "URL Preview (Prometido)",
    "URL Producción (Entregado)",
    "Nombre",
    "Parámetro",
    "Valor Prometido",
    "Valor Entregado",
    "Match",
]
MATCH_COL_WIDTHS = [45, 45, 25, 20, 50, 50, 12]

# 3‑way: US Expected vs Preview vs Producción
MATCH3_COLS = [
    "Nombre / URL",
    "Parámetro",
    "Valor US Expected",
    "Valor Preview (Real)",
    "Valor Producción (Real)",
    "US vs Preview",
    "US vs Production",
    "Preview vs Production",
]
MATCH3_COL_WIDTHS = [50, 20, 40, 40, 40, 14, 14, 14]

# FILL_FAIL → FILL_FAIL (mismo color FCE4EC)




def find_digitaldata_in_historial(historial_path: str, url_fragment: str) -> dict:
    """Busca digitalData por URL en un historial."""
    if not os.path.exists(historial_path):
        return {}
    wb = openpyxl.load_workbook(historial_path, data_only=True)
    data_sheet = None
    for sn in wb.sheetnames:
        if sn not in ("_control", "_vars", "Sheet"):
            data_sheet = wb[sn]
            break
    if not data_sheet and "Sheet" in wb.sheetnames:
        data_sheet = wb["Sheet"]
    if not data_sheet:
        wb.close()
        return {}

    dd_col = 3
    for c in range(1, data_sheet.max_column + 1):
        hv = data_sheet.cell(1, c).value
        if hv and "digitaldata (automatica)" in str(hv).strip().lower():
            dd_col = c
            break

    for row in range(2, data_sheet.max_row + 1):
        url_cell = data_sheet.cell(row, 2).value
        if url_cell and url_fragment in str(url_cell):
            dd_raw = data_sheet.cell(row, dd_col).value
            wb.close()
            if dd_raw and isinstance(dd_raw, str):
                try:
                    return json.loads(dd_raw)
                except json.JSONDecodeError:
                    return {}
            return {}
    wb.close()
    return {}


def build_expected_page(page_key: str, cfg: dict) -> dict:
    """Construye dict digitalData.page según expected.json."""
    expected = {}
    for param in PARAMS_ORDER:
        param_cfg = cfg.get("params", {}).get(param, {})
        rule = param_cfg.get("rule", "pattern")
        if rule == "deprecated":
            continue
        if rule == "fixed":
            val = param_cfg.get("value", "")
        elif rule == "mirror":
            continue  # se resuelve con pageName
        elif rule in ("mapping", "required"):
            val = param_cfg.get("mapping", {}).get(page_key, "")
        else:  # pattern
            val = param_cfg.get("patterns", {}).get(page_key, "")
        if val:
            expected[param] = val
    return expected


def compare_params(promised: dict, delivered: dict,
                   param_list: list[str]) -> list[dict]:
    """Compara dos dicts digitalData.page parámetro por parámetro (2‑way)."""
    results = []
    for param in param_list:
        p_val = promised.get(param)
        d_val = delivered.get(param)
        p_str = str(p_val) if p_val is not None else ""
        d_str = str(d_val) if d_val is not None else ""

        if not p_str and not d_str:
            results.append({"param": param, "promised": "—", "delivered": "—",
                            "match": "⚠️", "match_text": "Sin datos en ambos"})
        elif not p_str:
            results.append({"param": param, "promised": "—", "delivered": d_str,
                            "match": "❌", "match_text": "Sin dato prometido"})
        elif not d_str:
            results.append({"param": param, "promised": p_str, "delivered": "—",
                            "match": "❌", "match_text": "Sin dato entregado"})
        elif p_str == d_str:
            results.append({"param": param, "promised": p_str, "delivered": d_str,
                            "match": "✅", "match_text": "Match"})
        else:
            results.append({"param": param, "promised": p_str, "delivered": d_str,
                            "match": "⚠️", "match_text": "Diferente"})
    return results


def _match_label(a: str, b: str) -> str:
    """Retorna emoji match entre dos valores."""
    if not a and not b:
        return "⚠️"
    if not a or not b:
        return "❌"
    return "✅" if a == b else "⚠️"


def compare_params_3way(expected: dict, preview: dict,
                        production: dict,
                        param_list: list[str]) -> list[dict]:
    """Compara US Expected vs Preview vs Producción (3‑way)."""
    results = []
    for param in param_list:
        e_val = str(expected.get(param, "")) if expected.get(param) else ""
        p_val = str(preview.get(param, "")) if preview.get(param) else ""
        d_val = str(production.get(param, "")) if production.get(param) else ""

        results.append({
            "param": param,
            "expected": e_val or "—",
            "preview": p_val or "—",
            "production": d_val or "—",
            "match_ep": _match_label(e_val, p_val),
            "match_ed": _match_label(e_val, d_val),
            "match_pd": _match_label(p_val, d_val),
        })
    return results


# ════════════════════════════════════════════
#  OUTPUT
# ════════════════════════════════════════════

def generate_report(mappings: list, all_results: list,
                    market: str, output_dir: str, mode: str):
    """Genera .xlsx, .md y .html con los resultados de la comparación."""
    if mode == "3way":
        return _generate_report_3way(mappings, all_results, market, output_dir)

    today = date.today().isoformat()

    # ── Contar estadísticas ──
    total_params = sum(len(r["params"]) for r in all_results)
    match_ok = sum(1 for r in all_results for p in r["params"] if p["match"] == "✅")
    match_warn = sum(1 for r in all_results for p in r["params"] if p["match"] == "⚠️")
    match_fail = sum(1 for r in all_results for p in r["params"] if p["match"] == "❌")
    ok_pct = round(match_ok / total_params * 100, 1) if total_params else 0

    # ═══════════════════  .xlsx  ═══════════════════
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Match {market}"

    for col_idx, header in enumerate(MATCH_COLS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for col_idx, width in enumerate(MATCH_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    row_idx = 2
    for ent in all_results:
        for pr in ent["params"]:
            ws.cell(row_idx, 1, ent["preview_url"]).font = FONT_DATA
            ws.cell(row_idx, 1).alignment = ALIGN_WRAP
            ws.cell(row_idx, 1).border = THIN_BORDER

            ws.cell(row_idx, 2, ent["production_url"]).font = FONT_DATA
            ws.cell(row_idx, 2).alignment = ALIGN_WRAP
            ws.cell(row_idx, 2).border = THIN_BORDER

            ws.cell(row_idx, 3, ent["nombre"]).font = FONT_PARAM
            ws.cell(row_idx, 3).border = THIN_BORDER

            ws.cell(row_idx, 4, pr["param"]).font = FONT_PARAM
            ws.cell(row_idx, 4).border = THIN_BORDER

            ws.cell(row_idx, 5, pr["promised"]).font = FONT_DATA
            ws.cell(row_idx, 5).alignment = ALIGN_WRAP
            ws.cell(row_idx, 5).border = THIN_BORDER

            ws.cell(row_idx, 6, pr["delivered"]).font = FONT_DATA
            ws.cell(row_idx, 6).alignment = ALIGN_WRAP
            ws.cell(row_idx, 6).border = THIN_BORDER

            cell_m = ws.cell(row_idx, 7, f"{pr['match']} {pr['match_text']}")
            cell_m.font = FONT_PARAM
            cell_m.alignment = ALIGN_CENTER
            cell_m.border = THIN_BORDER

            fill = FILL_OK if pr["match"] == "✅" else (FILL_WARN if pr["match"] == "⚠️" else FILL_FAIL)
            cell_m.fill = fill

            ws.row_dimensions[row_idx].height = 22
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{row_idx - 1}"

    xlsx_path = os.path.join(output_dir, "match-prod-vs-preview.xlsx")
    wb.save(xlsx_path)
    print(f"[OK] Match .xlsx: {xlsx_path}")

    # ═══════════════════  .md  ═══════════════════
    md_lines = [
        f"# Match Prod vs Preview — {market}",
        "",
        f"**Fecha**: {today}",
        f"**Modo**: {mode}",
        f"**URLs**: {len(all_results)}",
        f"**Parámetros**: {total_params}",
        "",
        "## Resumen",
        "",
        "| Estado | Cantidad | Porcentaje |",
        "|--------|----------|------------|",
        f"| ✅ Match | {match_ok} | {ok_pct}% |",
        f"| ⚠️ Diferente | {match_warn} | {round(match_warn/total_params*100, 1) if total_params else 0}% |",
        f"| ❌ Sin dato | {match_fail} | {round(match_fail/total_params*100, 1) if total_params else 0}% |",
        "",
    ]

    # Mostrar solo diferencias
    diff_entries = [(e, [p for p in e["params"] if p["match"] != "✅"])
                    for e in all_results]
    diff_entries = [(e, pp) for e, pp in diff_entries if pp]

    if diff_entries:
        md_lines.append("## Diferencias encontradas")
        md_lines.append("")
        for ent, params in diff_entries:
            md_lines.append(f"### {ent['nombre'] or ent['production_url']}")
            md_lines.append(f"- **Preview**: {ent['preview_url']}")
            md_lines.append(f"- **Producción**: {ent['production_url']}")
            md_lines.append("")
            md_lines.append("| Parámetro | Prometido | Entregado | Estado |")
            md_lines.append("|-----------|-----------|-----------|--------|")
            for p in params:
                md_lines.append(f"| `{p['param']}` | {p['promised']} | {p['delivered']} | {p['match']} {p['match_text']} |")
            md_lines.append("")
    else:
        md_lines.append("✅ **Todo en match** — no hay diferencias entre lo prometido y lo entregado.")
        md_lines.append("")

    md_path = os.path.join(output_dir, "match-prod-vs-preview.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))
    print(f"[OK] Match .md: {md_path}")

    # ═══════════════════  .html  ═══════════════════
    def _esc(s: str) -> str:
        return (s.replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))

    health = "green" if ok_pct >= 80 else ("orange" if ok_pct >= 50 else "red")

    rows_html = []
    for ent, params in diff_entries:
        rows_html.append(
            f'<tr class="url-sep"><td colspan="4">'
            f'<strong>{_esc(ent["nombre"] or ent["production_url"])}</strong><br>'
            f'<small>P: {_esc(ent["preview_url"])}</small><br>'
            f'<small>D: {_esc(ent["production_url"])}</small>'
            f'</td></tr>')
        for p in params:
            color = {"✅": "green", "⚠️": "orange", "❌": "red"}.get(p["match"], "black")
            rows_html.append(
                f'<tr><td><code>{p["param"]}</code></td>'
                f'<td>{_esc(p["promised"])}</td>'
                f'<td>{_esc(p["delivered"])}</td>'
                f'<td style="color:{color}">{p["match"]} {_esc(p["match_text"])}</td></tr>')

    table_html = ""
    if rows_html:
        table_html = (
            "<table><thead><tr>"
            "<th>Parámetro</th><th>Prometido</th><th>Entregado</th><th>Estado</th>"
            "</tr></thead><tbody>" + "\n".join(rows_html) + "</tbody></table>")
    else:
        table_html = '<p style="color:green;font-size:1.1em">✅ Todo en match</p>'

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Match Prod vs Preview — {_esc(market)}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#f8f9fa; color:#333; padding:20px; }}
  h1 {{ font-size:1.5em; }}
  .meta {{ color:#666; font-size:0.9em; margin:8px 0 16px; }}
  .summary {{ display:flex; gap:12px; margin:16px 0; flex-wrap:wrap; }}
  .card {{ background:#fff; border-radius:8px; padding:16px 20px; flex:1; min-width:100px; box-shadow:0 1px 3px rgba(0,0,0,.08); text-align:center; }}
  .card .num {{ font-size:1.8em; font-weight:700; }}
  .card .label {{ font-size:0.8em; color:#666; }}
  .health {{ font-size:1.1em; margin:12px 0 20px; padding:10px 16px; border-radius:6px; color:#fff; font-weight:600; text-align:center; }}
  .health.green {{ background:#28a745; }}
  .health.orange {{ background:#fd7e14; }}
  .health.red {{ background:#dc3545; }}
  table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
  th {{ background:#4472C4; color:#fff; padding:10px 12px; text-align:left; font-size:0.85em; }}
  td {{ padding:8px 12px; border-bottom:1px solid #eee; font-size:0.85em; vertical-align:top; }}
  tr.url-sep td {{ background:#eef; font-size:0.95em; border-top:2px solid #4472C4; }}
  code {{ background:#f0f0f0; padding:1px 4px; border-radius:3px; font-size:0.9em; }}
</style>
</head>
<body>
<h1>Match Prod vs Preview — {_esc(market)}</h1>
<div class="meta">
  📅 {today} &nbsp;|&nbsp; Modo: {_esc(mode)} &nbsp;|&nbsp; {len(all_results)} URLs &nbsp;|&nbsp; {total_params} parámetros
</div>
<div class="health {health}">
  Match rate: {ok_pct}%
</div>
<div class="summary">
  <div class="card"><div class="num" style="color:#28a745">{match_ok}</div><div class="label">✅ Match</div></div>
  <div class="card"><div class="num" style="color:#fd7e14">{match_warn}</div><div class="label">⚠️ Diferente</div></div>
  <div class="card"><div class="num" style="color:#dc3545">{match_fail}</div><div class="label">❌ Sin dato</div></div>
</div>
{table_html}
</body>
</html>"""

    html_path = os.path.join(output_dir, "match-prod-vs-preview.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Match .html: {html_path}")

    # Resumen en consola
    print(f"\n  ✅ Match: {match_ok}  |  ⚠️ Diferente: {match_warn}  |  ❌ Sin dato: {match_fail}")
    print(f"  Match rate: {ok_pct}%")


# ════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════

def _generate_report_3way(mappings: list, all_results: list,
                           market: str, output_dir: str):
    """Genera reporte 3‑vías: US Expected vs Preview vs Producción."""
    today = date.today().isoformat()

    # ── Estadísticas ──
    total_params = sum(len(r["params"]) for r in all_results)
    ok_ep = sum(1 for r in all_results for p in r["params"] if p["match_ep"] == "✅")
    ok_ed = sum(1 for r in all_results for p in r["params"] if p["match_ed"] == "✅")
    ok_pd = sum(1 for r in all_results for p in r["params"] if p["match_pd"] == "✅")

    # ═════════ .xlsx ═════════
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"3Way {market}"

    for col_idx, header in enumerate(MATCH3_COLS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for col_idx, width in enumerate(MATCH3_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    row_idx = 2
    for ent in all_results:
        name = ent.get("nombre") or ent.get("production_url", "")
        url_display = ent.get("production_url", "")
        for pr in ent["params"]:
            ws.cell(row_idx, 1, f"{name}\n{url_display}").font = FONT_DATA
            ws.cell(row_idx, 1).alignment = ALIGN_WRAP
            ws.cell(row_idx, 1).border = THIN_BORDER

            ws.cell(row_idx, 2, pr["param"]).font = FONT_PARAM
            ws.cell(row_idx, 2).border = THIN_BORDER

            ws.cell(row_idx, 3, pr["expected"]).font = FONT_DATA
            ws.cell(row_idx, 3).alignment = ALIGN_WRAP
            ws.cell(row_idx, 3).border = THIN_BORDER

            ws.cell(row_idx, 4, pr["preview"]).font = FONT_DATA
            ws.cell(row_idx, 4).alignment = ALIGN_WRAP
            ws.cell(row_idx, 4).border = THIN_BORDER

            ws.cell(row_idx, 5, pr["production"]).font = FONT_DATA
            ws.cell(row_idx, 5).alignment = ALIGN_WRAP
            ws.cell(row_idx, 5).border = THIN_BORDER

            for col, key in [(6, "match_ep"), (7, "match_ed"), (8, "match_pd")]:
                val = pr[key]
                cell_m = ws.cell(row_idx, col, val)
                cell_m.font = FONT_PARAM
                cell_m.alignment = ALIGN_CENTER
                cell_m.border = THIN_BORDER
                fill = FILL_OK if val == "✅" else (FILL_WARN if val == "⚠️" else FILL_FAIL)
                cell_m.fill = fill

            ws.row_dimensions[row_idx].height = 28
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{row_idx - 1}"

    xlsx_path = os.path.join(output_dir, "match-3way.xlsx")
    wb.save(xlsx_path)
    print(f"[OK] 3Way .xlsx: {xlsx_path}")

    # ═════════ .md ═════════
    md_lines = [
        f"# Match 3‑Vías — {market}",
        f"**US Expected vs Preview vs Producción**",
        "",
        f"**Fecha**: {today}",
        f"**URLs**: {len(all_results)}",
        f"**Parámetros**: {total_params}",
        "",
        "## Resumen",
        "",
        "| Comparación | ✅ Match | ⚠️ Diferente | ❌ Sin dato |",
        "|------------|---------|-------------|------------|",
        f"| US vs Preview | {ok_ep} | {total_params - ok_ep - sum(1 for r in all_results for p in r['params'] if p['match_ep']=='❌')} | {sum(1 for r in all_results for p in r['params'] if p['match_ep']=='❌')} |",
        f"| US vs Producción | {ok_ed} | {total_params - ok_ed - sum(1 for r in all_results for p in r['params'] if p['match_ed']=='❌')} | {sum(1 for r in all_results for p in r['params'] if p['match_ed']=='❌')} |",
        f"| Preview vs Producción | {ok_pd} | {total_params - ok_pd - sum(1 for r in all_results for p in r['params'] if p['match_pd']=='❌')} | {sum(1 for r in all_results for p in r['params'] if p['match_pd']=='❌')} |",
        "",
    ]

    # mostrar diferencias
    diff_entries = [(e, [p for p in e["params"]
                         if p["match_ep"] != "✅" or p["match_ed"] != "✅" or p["match_pd"] != "✅"])
                    for e in all_results]
    diff_entries = [(e, pp) for e, pp in diff_entries if pp]

    if diff_entries:
        md_lines.append("## Diferencias")
        md_lines.append("")
        for ent, params in diff_entries:
            md_lines.append(f"### {ent.get('nombre') or ent.get('production_url', '')}")
            md_lines.append("")
            md_lines.append("| Parámetro | US Expected | Preview | Producción | E vs P | E vs D | P vs D |")
            md_lines.append("|-----------|------------|---------|------------|--------|--------|--------|")
            for p in params:
                md_lines.append(f"| `{p['param']}` | {p['expected']} | {p['preview']} | {p['production']} | {p['match_ep']} | {p['match_ed']} | {p['match_pd']} |")
            md_lines.append("")

    md_path = os.path.join(output_dir, "match-3way.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))
    print(f"[OK] 3Way .md: {md_path}")

    # ═════════ .html ═════════
    def _esc(s):
        return (s.replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))

    rows_html = []
    for ent, params in diff_entries:
        rows_html.append(
            f'<tr class="url-sep"><td colspan="8">'
            f'<strong>{_esc(ent.get("nombre") or "")}</strong><br>'
            f'<small>{_esc(ent.get("production_url", ""))}</small>'
            f'</td></tr>')
        for p in params:
            rows_html.append(
                f'<tr><td><code>{p["param"]}</code></td>'
                f'<td>{_esc(p["expected"])}</td>'
                f'<td>{_esc(p["preview"])}</td>'
                f'<td>{_esc(p["production"])}</td>'
                f'<td style="color:{"green" if p["match_ep"]=="✅" else "orange" if p["match_ep"]=="⚠️" else "red"}">{p["match_ep"]}</td>'
                f'<td style="color:{"green" if p["match_ed"]=="✅" else "orange" if p["match_ed"]=="⚠️" else "red"}">{p["match_ed"]}</td>'
                f'<td style="color:{"green" if p["match_pd"]=="✅" else "orange" if p["match_pd"]=="⚠️" else "red"}">{p["match_pd"]}</td>'
                f'</tr>')

    table_html = ""
    if rows_html:
        table_html = (
            "<table><thead><tr>"
            "<th>Param</th><th>US Exp.</th><th>Preview</th><th>Producción</th>"
            "<th>USvsPV</th><th>USvsPR</th><th>PVvsPR</th>"
            "</tr></thead><tbody>" + "\n".join(rows_html) + "</tbody></table>")
    else:
        table_html = '<p style="color:green;font-size:1.1em">✅ Todo alineado</p>'

    # cards de resumen
    ep_pct = round(ok_ep / total_params * 100, 1) if total_params else 0
    ed_pct = round(ok_ed / total_params * 100, 1) if total_params else 0
    pd_pct = round(ok_pd / total_params * 100, 1) if total_params else 0

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Match 3‑Vías — {_esc(market)}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#f8f9fa; color:#333; padding:20px; }}
  h1 {{ font-size:1.5em; }}
  .meta {{ color:#666; font-size:0.9em; margin:8px 0 16px; }}
  .summary {{ display:flex; gap:12px; margin:16px 0; flex-wrap:wrap; }}
  .card {{ background:#fff; border-radius:8px; padding:16px 20px; flex:1; min-width:140px; box-shadow:0 1px 3px rgba(0,0,0,.08); text-align:center; }}
  .card .num {{ font-size:1.6em; font-weight:700; }}
  .card .label {{ font-size:0.75em; color:#666; }}
  table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
  th {{ background:#4472C4; color:#fff; padding:8px 10px; text-align:left; font-size:0.8em; }}
  td {{ padding:6px 10px; border-bottom:1px solid #eee; font-size:0.8em; vertical-align:top; }}
  tr.url-sep td {{ background:#eef; font-size:0.9em; border-top:2px solid #4472C4; }}
  code {{ background:#f0f0f0; padding:1px 4px; border-radius:3px; font-size:0.9em; }}
</style>
</head>
<body>
<h1>Match 3‑Vías — {_esc(market)}</h1>
<div class="meta">
  📅 {today} &nbsp;|&nbsp; {len(all_results)} URLs &nbsp;|&nbsp; {total_params} parámetros &nbsp;|&nbsp; Modo 3‑vías
</div>
<div class="summary">
  <div class="card"><div class="num" style="color:#28a745">{ep_pct}%</div><div class="label">US vs Preview</div></div>
  <div class="card"><div class="num" style="color:#28a745">{ed_pct}%</div><div class="label">US vs Producción</div></div>
  <div class="card"><div class="num" style="color:#28a745">{pd_pct}%</div><div class="label">Preview vs Producción</div></div>
</div>
{table_html}
</body>
</html>"""

    html_path = os.path.join(output_dir, "match-3way.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] 3Way .html: {html_path}")

    print(f"\n  US vs Preview:     {ok_ep}/{total_params} ({ep_pct}%)")
    print(f"  US vs Producción:  {ok_ed}/{total_params} ({ed_pct}%)")
    print(f"  Preview vs Prod:   {ok_pd}/{total_params} ({pd_pct}%)")


def main():
    parser = argparse.ArgumentParser(
        description="Compara digitalData entre lo prometido (Preview/Expected) y lo entregado (Producción)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)

    parser.add_argument("--production", default=None,
                        help="Historial de producción (entregado). En modo 3way: "
                             "ruta al historial de producción.")
    parser.add_argument("--preview", default=None,
                        help="Historial de preview (prometido). Si se omite, se busca "
                             "{market}/preview/historial.xlsx automáticamente. "
                             "Si no existe, se usa expected.json como fallback.")
    parser.add_argument("--mapping", default="data/url-mapping.json",
                        help="JSON mapeo preview→producción")
    parser.add_argument("--expected", default="data/expected.json",
                        help="JSON con valores esperados (US Global)")
    parser.add_argument("--market", default="PR",
                        help="Código de mercado")
    parser.add_argument("--output", default=None,
                        help="Directorio de salida (default: {market}/match/)")
    parser.add_argument("--mode", default="auto",
                        choices=("auto", "2way", "3way"),
                        help="Modo de comparación: 2way (prometido vs entregado), "
                             "3way (US Expected vs Preview vs Producción), "
                             "auto (3way si existen ambos historiales, else 2way)")

    args = parser.parse_args()

    if not os.path.exists(args.mapping):
        print(f"[ERR] Mapping no encontrado: {args.mapping}")
        sys.exit(1)

    mappings = load_json(args.mapping)
    market_dir = args.market.upper()
    output_dir = args.output or os.path.join(market_dir, "match")
    os.makedirs(output_dir, exist_ok=True)

    expected_cfg = load_json(args.expected) if os.path.exists(args.expected) else {}

    # ── Resolver rutas por entorno ──
    prod_path = args.production or os.path.join(market_dir, "produccion", "historial.xlsx")
    prev_path = args.preview or os.path.join(market_dir, "preview", "historial.xlsx")
    has_prod = os.path.exists(prod_path)
    has_prev = os.path.exists(prev_path)

    # ── Determinar modo ──
    mode = args.mode
    if mode == "auto":
        mode = "3way" if (has_prod and has_prev) else "2way"

    if mode == "3way" and has_prod and has_prev:
        print(f"[INFO] Modo 3‑vías: Expected vs Preview ({prev_path}) vs Producción ({prod_path})")
        all_results = []
        for mapping in mappings:
            preview_url = mapping.get("preview_url", "")
            production_url = mapping.get("production_url", preview_url)
            page_key = mapping.get("page_key", "")
            nombre = mapping.get("nombre", "")
            if not page_key:
                continue

            # US Expected
            market_cfg = expected_cfg.get("markets", {}).get(args.market.upper(), {})
            expected_page = build_expected_page(page_key, market_cfg)

            # Preview real
            prev_dd = find_digitaldata_in_historial(prev_path, preview_url)
            prev_page = prev_dd.get("page", {}) if prev_dd else {}

            # Producción real
            prod_dd = find_digitaldata_in_historial(prod_path, production_url)
            prod_page = prod_dd.get("page", {}) if prod_dd else {}

            params = compare_params_3way(expected_page, prev_page, prod_page, PARAMS_ORDER)
            all_results.append({
                "preview_url": preview_url,
                "production_url": production_url,
                "page_key": page_key,
                "nombre": nombre,
                "params": params,
            })
            print(f"  {nombre or production_url[:50]}")

        generate_report(mappings, all_results, args.market.upper(), output_dir, "3way")

    else:
        # ── Modo 2‑way (original) ──
        prod_final = prod_path if has_prod else args.production
        if not prod_final or not os.path.exists(prod_final):
            # Fallback a ruta legacy {market}/historial.xlsx
            prod_final = os.path.join(market_dir, "historial.xlsx")
        if not os.path.exists(prod_final):
            print(f"[ERR] Historial de producción no encontrado")
            print(f"  Buscado en: {prod_path}")
            print(f"  Buscado en: {os.path.join(market_dir, 'historial.xlsx')}")
            sys.exit(1)

        print(f"[INFO] Producción: {prod_final}")

        # Detectar preview
        preview_path = prev_path if has_prev else None
        has_preview = preview_path is not None
        mode_label = ""

        if preview_path and os.path.exists(preview_path):
            has_preview = True
            mode_label = "preview-real"
            print(f"[INFO] Preview: {preview_path}")
        else:
            auto_legacy = os.path.join(market_dir, "historial_preview.xlsx")
            if os.path.exists(auto_legacy):
                preview_path = auto_legacy
                has_preview = True
                mode_label = "preview-real"
                print(f"[INFO] Preview (legacy): {auto_legacy}")
            else:
                if not expected_cfg:
                    print(f"[ERR] Sin preview ({prev_path}) ni expected ({args.expected})")
                    sys.exit(1)
                mode_label = "expected-fallback"
                print(f"[INFO] Modo expected (sin preview). Usando: {args.expected}")

        all_results = []
        for mapping in mappings:
            preview_url = mapping.get("preview_url", "")
            production_url = mapping.get("production_url", preview_url)
            page_key = mapping.get("page_key", "")
            nombre = mapping.get("nombre", "")
            if not page_key:
                continue

            if has_preview:
                promised_dd = find_digitaldata_in_historial(preview_path, preview_url)
                promised_page = promised_dd.get("page", {}) if promised_dd else {}
            else:
                market_cfg = expected_cfg.get("markets", {}).get(args.market.upper(), {})
                promised_page = build_expected_page(page_key, market_cfg)

            delivered_dd = find_digitaldata_in_historial(prod_final, production_url)
            delivered_page = delivered_dd.get("page", {}) if delivered_dd else {}

            params = compare_params(promised_page, delivered_page, PARAMS_ORDER)
            all_results.append({
                "preview_url": preview_url,
                "production_url": production_url,
                "page_key": page_key,
                "nombre": nombre,
                "params": params,
            })
            status = "✅" if all(p["match"] == "✅" for p in params) else "⚠️"
            print(f"  {status} {nombre or production_url[:50]}")

        generate_report(mappings, all_results, args.market.upper(), output_dir, mode_label)


if __name__ == "__main__":
    main()

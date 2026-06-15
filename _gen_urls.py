"""Extrae URLs de Excel → urls.json con detección automática de mercado desde la URL.

Uso:
  python _gen_urls.py                                         # urls.json desde RevisionManual.xlsx
  python _gen_urls.py --input otro.xlsx                       # desde otro archivo
  python _gen_urls.py --output mis_urls.json                  # nombre custom de salida
  python _gen_urls.py --default-market MX                     # override manual si no se detecta
"""
import openpyxl, json, argparse, re
from urllib.parse import urlparse

DEFAULT_INPUT = "RevisionManual.xlsx"
DEFAULT_OUTPUT = "urls.json"
_MARKET_RE = re.compile(r"/([a-z]{2})_([a-z]{2})/")

def detect_market_from_url(url: str) -> str | None:
    """Extrae el código de mercado desde la URL.

    Busca el patrón /{idioma}_{mercado}/ en el path.
    Ej: /es_pr/ → 'PR', /en_pr/ → 'PR', /es_mx/ → 'MX'
    """
    m = _MARKET_RE.search(url)
    return m.group(2).upper() if m else None


def make_name_from_url(url: str) -> str:
    """Genera un nombre legible desde la URL.

    Toma el último segmento del path (sin .html) y convierte
    guiones en espacios con capitalización.
    Ej: '.../driving-precautions.html' → 'Driving Precautions'
        '.../donativos-ambientales.html' → 'Donativos Ambientales'
    """
    path = urlparse(url).path
    # Sacar el archivo (último segmento)
    filename = path.rstrip("/").split("/")[-1]
    # Sacar extensión .html
    name = re.sub(r"\.html?$", "", filename)
    # Reemplazar guiones por espacios y capitalizar
    words = name.replace("-", " ").split()
    return " ".join(w.capitalize() for w in words) if words else ""


def extract(input_path: str, default_market: str = None) -> list:
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    entries = []
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row, 2).value
        if not url:
            continue
        url = str(url).strip()
        nombre = ws.cell(row, 1).value

        entry = {"url": url}
        if nombre and str(nombre).strip():
            entry["nombre"] = str(nombre).strip()
        else:
            # Generar nombre desde la URL si el Excel no tiene
            auto_name = make_name_from_url(url)
            if auto_name:
                entry["nombre"] = auto_name

        # Detectar mercado automáticamente desde la URL
        market = detect_market_from_url(url)
        if market:
            entry["market"] = market
        elif default_market:
            entry["market"] = default_market.upper()

        entries.append(entry)
    return entries

def main():
    parser = argparse.ArgumentParser(description="Extrae URLs de Excel a JSON con detección automática de mercado")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Excel de entrada")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="JSON de salida")
    parser.add_argument("--default-market", help="Override manual si la URL no tiene patrón de mercado")
    args = parser.parse_args()

    entries = extract(args.input, args.default_market)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

    # Estadísticas
    markets = set(e.get("market") for e in entries if e.get("market"))
    print(f"Generado {args.output} con {len(entries)} URLs desde {args.input}")
    if markets:
        print(f"  Mercados detectados: {', '.join(sorted(markets))}")
    else:
        print(f"  (sin mercado detectado en las URLs)")

if __name__ == "__main__":
    main()

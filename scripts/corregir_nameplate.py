"""
corregir_nameplate.py — Aplica los 20 fixes detectados por analizar_nameplate.py

Crea una copia corregida: FPR_MYCO_Nameplate_Bronco_CORREGIDO.xlsx

Fixes:
  Fix 1  (9x): Labels "Version Base - Color X" → "Version {correcta} - Color X"
  Fix 2  (2x): "Bronco Base" → "Bronco Badlands" en Badlands
  Fix 3  (1x): "#REF!" en contador → reemplazar con LEN correcto
  Fix 4  (8x): Color 8/9 duplicados → renumbered a Color 10/11
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("[ERR] openpyxl no instalado")

INPUT = "docs/FPR_MYCO_Nameplate_Bronco.xlsx"
OUTPUT = "docs/FPR_MYCO_Nameplate_Bronco_CORREGIDO.xlsx"

COL_A, COL_B, COL_C, COL_D, COL_E, COL_F, COL_G = range(7)
FIX_LOG = []


def n(val):
    return "" if val is None else str(val).strip()


def log(msg):
    FIX_LOG.append(msg)
    print(f"  {msg}")


def detect_ref_cells(path: str) -> list[tuple[str, int]]:
    """Pass 1: detecta celdas con #REF! (data_only=True)."""
    refs = []
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            val = str(row[COL_E].value or "")
            if val == "#REF!":
                refs.append((sheet_name, row[COL_E].row))
    wb.close()
    return refs


def fix_labels(ws, sheet_name, target):
    """Fix 1: 'Version Base - Color X' → 'Version {target} - Color X'."""
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[COL_B]
        val = n(cell.value)
        if "Version Base" in val and "Color" in val:
            new_val = val.replace("Version Base", f"Version {target}")
            if val != new_val:
                log(f"[{sheet_name}] Row {cell.row}: B '{val}' -> '{new_val}'")
                cell.value = new_val
                count += 1
    return count


def fix_badlands_titles(ws):
    """Fix 2: 'Bronco Base' → 'Bronco Badlands' en Title/Copy de Badlands."""
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[COL_D]
        tipo = n(row[COL_C].value)
        val = str(cell.value or "")
        if tipo in ("Title", "Sub Title", "Copy") and "Bronco" in val and "Base" in val:
            if "Badlands" not in val:
                old = val
                # Replace both "Bronco Base" and "Bronco® Base"
                new_val = val.replace("Bronco® Base", "Bronco® Badlands®")
                new_val = new_val.replace("Bronco Base", "Bronco Badlands")
                if new_val != old:
                    cell.value = new_val
                    count += 1
                    log(f"[Badlands] Row {cell.row}: D '{old[:60]}...' -> '{new_val[:60]}...'")
    return count


def fix_ref_cells(ws, sheet_name, ref_rows: set):
    """Fix 3: Reemplaza #REF! con LEN(Dn) calculado."""
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        r = row[COL_E].row
        if r in ref_rows:
            cell_d = row[COL_D]
            cell_e = row[COL_E]
            content_len = len(str(cell_d.value or ""))
            log(f"[{sheet_name}] Row {r}: E '#REF!' -> '{content_len}' (LEN de '{str(cell_d.value or "")[:50]}...')")
            cell_e.value = str(content_len)
            count += 1
    return count


def fix_color_duplicates(ws, sheet_name):
    """Fix 4: Color 8/9 duplicados → renumber a 10/11."""
    seen = {}
    count = 0
    next_free = 10
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[COL_B]
        val = n(cell.value)
        m = re.search(r"(Color\s+)(\d+)", val)
        if m:
            prefix = m.group(1)
            num = int(m.group(2))
            tag = f"Color {num}"
            if tag in seen:
                new_tag = f"Color {next_free}"
                new_val = val.replace(f"Color {num}", new_tag)
                log(f"[{sheet_name}] Row {cell.row}: B '{val}' -> '{new_val}'")
                cell.value = new_val
                next_free += 1
                count += 1
            else:
                seen[tag] = cell.row
    return count


def main():
    print("=" * 60)
    print("  CORREGIR NAMEPLATE — Aplicando fixes")
    print("=" * 60)

    input_path = Path(INPUT)
    output_path = Path(OUTPUT)

    if not input_path.exists():
        print(f"[ERR] No se encuentra: {INPUT}")
        sys.exit(1)

    # ── Pass 1: Detectar #REF! ──
    print(f"\n[Pass 1/3] Detectando #REF!...")
    ref_cells = detect_ref_cells(str(input_path))
    # Agrupar por sheet para lookup rápido
    ref_by_sheet: dict[str, set] = {}
    for s, r in ref_cells:
        if s not in ref_by_sheet:
            ref_by_sheet[s] = set()
        ref_by_sheet[s].add(r)
    if ref_cells:
        for s, r in ref_cells:
            log(f"  #REF! detectado en [{s}] row {r}")
    else:
        print("  Ninguno detectado")

    # ── Pass 2: Cargar con fórmulas y aplicar fixes ──
    print(f"\n[Pass 2/3] Cargando workbook (con formulas)...")
    wb = openpyxl.load_workbook(str(input_path))  # data_only=False (default)

    total_labels = 0
    total_titles = 0
    total_ref = 0
    total_colors = 0

    # Fix 1: Labels
    print("\n--- Fix 1: Labels 'Version Base' incorrectos ---")
    for sheet_name, target in [
        ("Bronco Version Big Bend", "Big Bend"),
        ("Bronco Version Outer Banks", "Outer Banks"),
        ("Bronco Version Badlands", "Badlands"),
    ]:
        if sheet_name in wb.sheetnames:
            total_labels += fix_labels(wb[sheet_name], sheet_name, target)

    # Fix 2: Títulos Badlands
    print("\n--- Fix 2: Títulos 'Bronco Base' en Badlands ---")
    if "Bronco Version Badlands" in wb.sheetnames:
        total_titles = fix_badlands_titles(wb["Bronco Version Badlands"])

    # Fix 3: #REF!
    print("\n--- Fix 3: #REF! en Contadores ---")
    for sheet_name in wb.sheetnames:
        if sheet_name in ref_by_sheet:
            total_ref += fix_ref_cells(wb[sheet_name], sheet_name, ref_by_sheet[sheet_name])

    # Fix 4: Colores duplicados
    print("\n--- Fix 4: Colores 8/9 duplicados ---")
    for sheet_name in ["Bronco Version Base", "Bronco Version Big Bend",
                        "Bronco Version Outer Banks", "Bronco Version Badlands"]:
        if sheet_name in wb.sheetnames:
            total_colors += fix_color_duplicates(wb[sheet_name], sheet_name)

    # ── Guardar ──
    print(f"\n[Pass 3/3] Guardando: {output_path.resolve()}")
    wb.save(str(output_path))
    wb.close()

    total = total_labels + total_titles + total_ref + total_colors
    print(f"\n{'='*60}")
    print(f"  TOTAL: {total} correcciones aplicadas")
    print(f"  {total_labels}/9 labels · {total_titles}/2 titulos · {total_ref} #REF! · {total_colors}/8 colores")
    print(f"{'='*60}")
    print(f"\nOutput: {output_path.resolve()}")

    # ── Verificar con analizar_nameplate ──
    print(f"\nVerificando con analizar_nameplate...")
    import subprocess
    result = subprocess.run(
        [sys.executable, "scripts/analizar_nameplate.py",
         "--input", str(output_path),
         "--ticket", "CORREGIDO",
         "--output-dir", "."],
        capture_output=True, text=True,
    )
    for line in result.stdout.split("\n"):
        if any(kw in line for kw in ("RESUMEN", "Quality flags", "labels", "titulo",
                                      "contador", "color", "incidencia", "Flags")):
            print(f"  {line.strip()}")
    if result.returncode != 0:
        for line in result.stderr.split("\n"):
            if line.strip():
                print(f"  [ERR] {line.strip()}")


if __name__ == "__main__":
    main()

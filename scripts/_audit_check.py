"""
_audit_check.py — Verifica calidad de archivos .xlsx generados.

Uso: python scripts/_audit_check.py
     python scripts/_audit_check.py --root-hist "C:/path/to/historial.xlsx"
"""
import argparse
import json
import os

import openpyxl


def main() -> None:
    parser = argparse.ArgumentParser(description="Verifica calidad de archivos .xlsx generados")
    parser.add_argument("--root-hist", default="", help="Ruta al historial.xlsx raíz (opcional)")
    args = parser.parse_args()

    print("=" * 60)
    print("1. reporte_auditoria.xlsx")
    print("=" * 60)
    wb = openpyxl.load_workbook("reporte_auditoria.xlsx", data_only=True)
    print("  Sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  [{sn}] {ws.max_row} rows x {ws.max_column} cols")
        h = [str(ws.cell(1, c).value or "")[:40] for c in range(1, ws.max_column + 1)]
        print(f"    Headers: {h}")
        for r in range(2, min(ws.max_row + 1, 5)):
            v = [str(ws.cell(r, c).value or "")[:50] for c in range(1, ws.max_column + 1)]
            print(f"    R{r}: {v}")
        if ws.max_row > 5:
            print(f"    ... ({ws.max_row - 1} data rows total)")
    wb.close()

    print()
    print("=" * 60)
    print("2. PR/con_aa.xlsx — content analysis")
    print("=" * 60)
    wb = openpyxl.load_workbook("PR/con_aa.xlsx", data_only=True)
    ws = wb.active
    print(f"  Rows: {ws.max_row - 1}, Cols: {ws.max_column}")
    headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    print(f"  Headers: {headers}")

    dd_real = 0
    dd_error = 0
    dd_empty = 0
    aa_real = 0
    aa_error = 0
    aa_empty = 0

    for r in range(2, ws.max_row + 1):
        col_d = str(ws.cell(r, 4).value or "").strip()
        col_f = str(ws.cell(r, 6).value or "").strip()

        if not col_d or col_d == "None":
            dd_empty += 1
        elif "error" in col_d.lower():
            dd_error += 1
        elif len(col_d) > 20:
            dd_real += 1
        else:
            dd_error += 1

        if not col_f or col_f == "None":
            aa_empty += 1
        elif "error" in col_f.lower():
            aa_error += 1
        elif len(col_f) > 20:
            aa_real += 1
        else:
            aa_error += 1

    total = ws.max_row - 1
    print(f"  Col D (digitaldata automatica): real={dd_real}, error={dd_error}, empty={dd_empty}")
    print(f"  Col F (AA analytics automatico): real={aa_real}, error={aa_error}, empty={aa_empty}")

    for r in range(2, ws.max_row + 1):
        col_d = str(ws.cell(r, 4).value or "").strip()
        if len(col_d) > 50 and "error" not in col_d.lower():
            nm = str(ws.cell(r, 1).value or "")[:30]
            print(f"\n  Sample REAL DD (row {r}, page={nm}):")
            print(f"    {col_d[:400]}")
            break

    for r in range(2, ws.max_row + 1):
        col_f = str(ws.cell(r, 6).value or "").strip()
        if len(col_f) > 50 and "error" not in col_f.lower():
            nm = str(ws.cell(r, 1).value or "")[:30]
            print(f"\n  Sample REAL AA (row {r}, page={nm}):")
            print(f"    {col_f[:400]}")
            break

    for r in range(2, ws.max_row + 1):
        col_f = str(ws.cell(r, 6).value or "").strip()
        if len(col_f) > 10 and "error" in col_f.lower():
            nm = str(ws.cell(r, 1).value or "")[:30]
            print(f"\n  Sample ERROR AA (row {r}, page={nm}):")
            print(f"    {col_f[:200]}")
            break

    wb.close()

    print()
    print("=" * 60)
    print("3. PR/sin_aa.xlsx — content analysis")
    print("=" * 60)
    wb = openpyxl.load_workbook("PR/sin_aa.xlsx", data_only=True)
    ws = wb.active
    print(f"  Rows: {ws.max_row - 1}, Cols: {ws.max_column}")

    dd_has = 0
    dd_empty = 0
    for r in range(2, ws.max_row + 1):
        col_d = str(ws.cell(r, 4).value or "").strip()
        if col_d and col_d != "None" and len(col_d) > 10:
            dd_has += 1
        else:
            dd_empty += 1
    print(f"  Col D: has content={dd_has}, empty={dd_empty}")

    for r in range(2, min(ws.max_row + 1, 7)):
        vals = [str(ws.cell(r, c).value or "")[:60] for c in range(1, ws.max_column + 1)]
        print(f"  R{r}: {vals}")
    wb.close()

    print()
    print("=" * 60)
    print("4. PR/preview_prod_match.xlsx")
    print("=" * 60)
    wb = openpyxl.load_workbook("PR/preview_prod_match.xlsx", data_only=True)
    ws = wb.active
    print(f"  Rows: {ws.max_row - 1}, Cols: {ws.max_column}")
    headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    print(f"  Headers: {headers}")
    for r in range(2, ws.max_row + 1):
        nm = str(ws.cell(r, 1).value or "")[:25]
        p_url = str(ws.cell(r, 2).value or "")[:50]
        pr_url = str(ws.cell(r, 3).value or "")[:50]
        p_dd = str(ws.cell(r, 4).value or "")
        pr_dd = str(ws.cell(r, 5).value or "")
        p_real = "REAL" if (len(p_dd) > 30 and "error" not in p_dd.lower()) else "ERROR/empty"
        pr_real = "REAL" if (len(pr_dd) > 30 and "error" not in pr_dd.lower()) else "ERROR/empty"
        print(f"  {nm:26s} | DD prev={p_real:12s} ({len(p_dd):4d}ch) | DD prod={pr_real:12s} ({len(pr_dd):4d}ch)")
    wb.close()

    print()
    print("=" * 60)
    print("5. Duplicados y relaciones preview/prod")
    print("=" * 60)

    root_hist = args.root_hist
    if root_hist and os.path.exists(root_hist):
        wb = openpyxl.load_workbook(root_hist, data_only=True)
        for sn in wb.sheetnames:
            if sn.startswith("_") or sn == "Sheet":
                continue
            ws = wb[sn]
            urls = []
            for r in range(2, ws.max_row + 1):
                u = str(ws.cell(r, 2).value or "").strip()
                if u:
                    urls.append(u)
            print(f"  Sheet '{sn}': {len(urls)} URLs")

            dupes = {u: urls.count(u) for u in set(urls) if urls.count(u) > 1}
            if dupes:
                print(f"  DUPLICATES in root historial: {len(dupes)}")
                for u, c in sorted(dupes.items())[:5]:
                    print(f"    x{c}: {u[:80]}")
            else:
                print("  No duplicates in root historial")
        wb.close()

    urls_data = json.load(open("data/urls.json"))
    env_types = {}
    url_by_env = {"preview": [], "production": [], "other": []}
    for e in urls_data:
        env = e.get("entorno", "other")
        env_types[env] = env_types.get(env, 0) + 1
        if env in url_by_env:
            url_by_env[env].append(e["url"])
    print(f"  urls.json envs: {env_types}")

    preview_urls = set(url_by_env["preview"])
    prod_urls = set(url_by_env["production"])
    overlap = preview_urls & prod_urls
    print(f"  Exact URL overlap preview↔prod: {len(overlap)}")

    preview_names = {}
    for e in urls_data:
        if e.get("entorno") == "preview":
            nm = e.get("nombre", "")
            url = e.get("url", "")
            preview_names[nm] = preview_names.get(nm, []) + [url]

    prod_names = {}
    for e in urls_data:
        if e.get("entorno") == "production":
            nm = e.get("nombre", "")
            url = e.get("url", "")
            prod_names[nm] = prod_names.get(nm, []) + [url]

    same_name = set(preview_names.keys()) & set(prod_names.keys())
    print(f"  Pages with SAME name in preview AND production: {len(same_name)}")
    for nm in sorted(same_name):
        pv = preview_names[nm]
        pr = prod_names[nm]
        print(f"    '{nm}': preview({len(pv)}) x production({len(pr)})")
        for u in pv:
            print(f"      PREVIEW: {u[:80]}")
        for u in pr:
            print(f"      PROD:    {u[:80]}")

    print("\nDone.")


if __name__ == "__main__":
    main()

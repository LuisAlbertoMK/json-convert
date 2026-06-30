"""
analizar_nameplate.py — Analiza Nameplate Excel y genera XLSX con formato original.

Workflow:
  1. Lee el Excel FPR_MYCO_Nameplate_Bronco.xlsx (todas las hojas visibles)
  2. Extrae URLs desde las filas donde el tipo = "URL"
  3. Preserva la estructura original del Excel (Visualización, Módulo, SEO/Formato, etc.)
  4. Genera XLSX con el MISMO formato visual: colores, bordes, fuentes, anchos
  5. Detecta anomalías de calidad automáticamente (hoja "Resumen")

Uso:
  python scripts/analizar_nameplate.py
  python scripts/analizar_nameplate.py --ticket GTBEMEAPUB-42479
  python scripts/analizar_nameplate.py --format csv     # solo CSV
  python scripts/analizar_nameplate.py --format both    # ambos

Columnas del XLSX (reflejan el Excel original):
  A: Visualización | B: Módulo / Sección | C: SEO/Formato
  D: Copies | E: Contadores | F: Otros/Notas/Comentarios | G: URL

Dependencias: openpyxl
"""

import argparse
import csv
import json
import logging
import re
import sys
from copy import copy
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, NamedStyle, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[ERR] openpyxl no instalado. Ejecutá: pip install openpyxl")

logger = logging.getLogger("analizar_nameplate")

# ── Constantes ──
DEFAULT_TICKET = "GTBEMEAPUB-42479"
DEFAULT_MERCADO = "PR"
DEFAULT_PAGINA = "Bronco"
EXCEL_PATH = "docs/FPR_MYCO_Nameplate_Bronco.xlsx"
SHEETS_A_INCLUIR = [
    "SEO Recommendations | ESP",
    "FPR_MYCO_Formato_Nameplate_v5(R",
    "Bronco Version Base",
    "Bronco Version Big Bend",
    "Bronco Version Outer Banks",
    "Bronco Version Badlands",
]
SHEET_VERSION_MAP = {
    "Bronco Version Base": "models/base",
    "Bronco Version Big Bend": "models/big-bend",
    "Bronco Version Outer Banks": "models/outer-banks",
    "Bronco Version Badlands": "models/badlands",
}
MAPA_HOJAS = {
    "SEO Recommendations | ESP": "SEO Metadatos",
    "FPR_MYCO_Formato_Nameplate_v5(R": "Maestra (Landing)",
    "Bronco Version Base": "Version Base",
    "Bronco Version Big Bend": "Version Big Bend",
    "Bronco Version Outer Banks": "Version Outer Banks",
    "Bronco Version Badlands": "Version Badlands",
}

# Columnas originales del Excel (0-indexed)
COL_VIS = 0     # A
COL_MOD = 1     # B
COL_TIPO = 2    # C
COL_CONT = 3    # D
COL_CTD = 4     # E
COL_NOTAS = 5   # F
COL_G = 6       # G

# Columnas del CSV
CSV_FIELDS = [
    "Visualización", "Módulo / Sección", "SEO/Formato",
    "Copies", "Contadores", "Otros/Notas/Comentarios",
    "URL", "Origen",
]

# ── Paleta de colores del original ──
CLR_DARK_BG = "44546A"       # Theme(8,tint=-0.25) — section headers (A)
CLR_LIGHT_GRAY = "BFBFBF"    # Column C — data labels
CLR_YELLOW = "FFFF00"        # SEO fields, URL, constraint notes
CLR_CYAN = "C5E0B3"         # Theme(6,tint=0.4) — category labels
CLR_MODULE_BG = "D9E2F3"    # Theme(9,tint=0) — module names (B)
CLR_RED = "FF0000"           # Counters, URL font
CLR_ORANGE = "ED7D31"        # Reminder notes
CLR_WHITE = "FFFFFF"
CLR_BLACK = "000000"
CLR_HEADER_BG = "D6DCE4"    # Theme(9,tint=0) header row

# ── Estilos reutilizables ──
_thin_side = Side(style="thin", color="808080")
_thin_all = Border(
    left=_thin_side, right=_thin_side,
    top=_thin_side, bottom=_thin_side,
)
_no_border = Border()

FONT_DEFAULT = Font(name="Calibri", size=11)
FONT_16 = Font(name="Calibri", size=16)
FONT_16_BOLD = Font(name="Calibri", size=16, bold=True)
FONT_16_BOLD_WHITE = Font(name="Calibri", size=16, bold=True, color=CLR_WHITE)
FONT_11_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_24_BOLD_RED = Font(name="Calibri", size=24, bold=True, color=CLR_RED)
FONT_16_RED = Font(name="Calibri", size=16, color=CLR_RED)
FONT_COUNTER = Font(name="Calibri", size=12, bold=True, color=CLR_WHITE)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

FILL_DARK = PatternFill(start_color=CLR_DARK_BG, end_color=CLR_DARK_BG, fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color=CLR_LIGHT_GRAY, end_color=CLR_LIGHT_GRAY, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=CLR_YELLOW, end_color=CLR_YELLOW, fill_type="solid")
FILL_CYAN = PatternFill(start_color=CLR_CYAN, end_color=CLR_CYAN, fill_type="solid")
FILL_MODULE = PatternFill(start_color=CLR_MODULE_BG, end_color=CLR_MODULE_BG, fill_type="solid")
FILL_RED = PatternFill(start_color=CLR_RED, end_color=CLR_RED, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=CLR_ORANGE, end_color=CLR_ORANGE, fill_type="solid")
FILL_HEADER = PatternFill(start_color=CLR_HEADER_BG, end_color=CLR_HEADER_BG, fill_type="solid")


# ── Utilidades ──

def _normalize(val) -> str:
    return "" if val is None else str(val).strip()


def _is_url_row(tipo: str, modulo: str, contenido: str) -> bool:
    return contenido.startswith("http") and (tipo.upper() == "URL" or modulo.upper() == "URL")


def _detect_content_type(tipo: str) -> str:
    """Clasifica el tipo de contenido para aplicar formato."""
    upper = tipo.upper()
    if upper in ("URL",):
        return "url"
    if upper in ("SEO IMAGE NAME", "SEO ALT. TXT.", "OG TITLE", "OG DESCRIPTION", "OG IMAGE NAME"):
        return "seo"
    if upper in ("CTA", "COPY", "COPY LABEL", "TITLE", "SUB TITLE", "EYEBROW",
                 "HEADER", "BODY COPY", "CARD BODY COPY", "CARD TITLE",
                 "LIST HEADER", "LIST INFORMATION", "NAMEPLATE 1", "NAMEPLATE 2",
                 "NAMEPLATE 3", "TOGGLE", "VEHICLE"):
        return "category"
    return "data"


def _es_cabecera(tipo: str, modulo: str) -> bool:
    """True si la fila es un header estructural (SECTION en col C o B)."""
    return tipo.upper() in ("SECTION", "SECCIÓN") or modulo.upper() in ("SECTION", "SECCIÓN")


# ── Extracción ──

def _extract_sheet_data(wb, sheet_name: str) -> tuple[list[dict], str]:
    """Extrae filas manteniendo la estructura original. Retorna (rows, page_url)."""
    ws = wb[sheet_name]
    rows: list[dict] = []
    page_url = ""
    seccion_actual = ""
    origen = MAPA_HOJAS.get(sheet_name, sheet_name)
    page_url_captured = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        vis = _normalize(row[COL_VIS])
        mod = _normalize(row[COL_MOD])
        tipo = _normalize(row[COL_TIPO])
        cont = _normalize(row[COL_CONT])
        ctd = _normalize(row[COL_CTD])
        notas = _normalize(row[COL_NOTAS])
        notas_g = _normalize(row[COL_G])

        if vis:
            seccion_actual = vis

        # Saltar solo filas completamente vacías (nada en A-D)
        if not vis and not mod and not tipo and not cont:
            continue

        # Saltar headers de columna
        if mod in ("Módulo / Sección",) or tipo == "SEO/Formato":
            if cont in ("Copies", "Contadores", "Otros/Notas/Comentarios") or \
               "SEO RECOMMENDATIONS" in cont:
                continue

        # Capturar URL de página (primera URL)
        es_url = _is_url_row(tipo, mod, cont)
        if es_url and not page_url_captured:
            page_url = cont
            page_url_captured = True

        # SECTION rows tienen tipo pero no contenido — incluirlas como agrupadores
        # Pero si no hay tipo NI contenido ni sección, skip
        if not tipo and not cont and not vis:
            continue

        notas_final = notas or notas_g
        url_fila = "" if _es_cabecera(tipo, mod) else page_url

        rows.append({
            "visualizacion": seccion_actual if (vis or seccion_actual) else "",
            "modulo": mod,
            "tipo": tipo,
            "contenido": cont,
            "contador": ctd,
            "notas": notas_final,
            "url": url_fila,
            "origen": origen,
            "_sheet": sheet_name,
        })

    return rows, page_url


def _extract_all(wb) -> tuple[list[dict], list[str], list[dict]]:
    """Extrae todos los datos. Retorna (rows, urls_por_sheet, issues)."""
    all_rows: list[dict] = []
    sheets_ok: list[str] = []

    for sheet_name in SHEETS_A_INCLUIR:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.sheet_state == "hidden":
            continue
        logger.info("Procesando: %s", sheet_name)
        rows, _ = _extract_sheet_data(wb, sheet_name)
        all_rows.extend(rows)
        sheets_ok.append(sheet_name)
        logger.debug("  -> %d filas", len(rows))

    # Validaciones globales
    issues = []
    all_text = " ".join(
        r.get("contenido", "") + " " + r.get("modulo", "")
        for r in all_rows
    ).lower()

    if "raptor" in all_text:
        has_sheet = any("Raptor" in s for s in SHEETS_A_INCLUIR)
        if not has_sheet:
            issues.append({
                "tipo": "cobertura", "severidad": "media",
                "descripcion": "Raptor existe en hoja maestra pero no tiene hoja dedicada",
            })

    for r in all_rows:
        c = r.get("notas", "") + r.get("contenido", "")
        if "falta actualizar" in c.lower():
            issues.append({
                "tipo": "obsolescencia", "severidad": "media",
                "descripcion": "Comentario: 'falta actualizar estas SEO Recommendations'",
            })
            break

    return all_rows, sheets_ok, issues


# ── Calidad ──

def _calc_quality(all_rows: list[dict]) -> dict:
    """Calcula estadísticas de calidad."""
    q = {"labels_incorrectos": 0, "titulo_incorrecto": 0,
         "contador_mixto": 0, "color_duplicado": 0}

    for r in all_rows:
        sheet = r.get("_sheet", "")
        mod = r.get("modulo", "")
        tipo = r.get("tipo", "")
        cont = r.get("contenido", "")
        ctd = r.get("contador", "")

        # Label check
        if sheet in SHEET_VERSION_MAP and sheet != "Bronco Version Base":
            if "Version Base" in mod:
                q["labels_incorrectos"] += 1

        # Title check
        if sheet == "Bronco Version Badlands" and tipo in ("Title", "Sub Title", "Copy"):
            if "Bronco" in cont and "Base" in cont and "Badlands" not in cont:
                q["titulo_incorrecto"] += 1

        # Contador check
        if ctd and not ctd.isdigit() and ctd not in ("", "Caracteres", "Palabras", "0"):
            q["contador_mixto"] += 1

        # Color duplicate (agrupa por sheet)
    for sheet_name in SHEET_VERSION_MAP:
        seen = set()
        for r in all_rows:
            if r.get("_sheet") != sheet_name:
                continue
            m = re.search(r"Color\s+(\d+)", r.get("modulo", ""))
            if m:
                tag = f"Color {m.group(1)}"
                if tag in seen:
                    q["color_duplicado"] += 1
                seen.add(tag)

    return q


# ── ESCRITURA CSV ──

def _write_csv(output_path: Path, all_rows: list[dict]):
    """Escribe CSV simple."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in all_rows:
            w.writerow({
                "Visualización": r["visualizacion"],
                "Módulo / Sección": r["modulo"],
                "SEO/Formato": r["tipo"],
                "Copies": r["contenido"],
                "Contadores": r["contador"],
                "Otros/Notas/Comentarios": r["notas"],
                "URL": r["url"],
                "Origen": r["origen"],
            })
    print(f"  CSV: {output_path.resolve()}  ({len(all_rows)} filas)")


# ── ESCRITURA XLSX (con formato original) ──

def _detect_style(tipo: str, cont: str, ctd: str, notas: str) -> dict:
    """Determina el estilo visual para una fila según su tipo y contenido."""
    upper = tipo.upper()
    ctype = _detect_content_type(tipo)
    has_notas_orange = "Se agregará" in notas
    has_notas_yellow = any(kw in notas for kw in ("caracteres", "APLICA", "máximo", "max"))

    # Estilo base
    style = {
        "font_b": None,  # B col font
        "font_c": None,  # C col font
        "font_d": FONT_16,  # D col font
        "fill_c": None,  # C col fill
        "fill_d": None,  # D col fill
        "fill_e": None,  # E col fill
        "fill_f": None,  # F col fill
        "font_e": None,  # E col font
    }

    # Columna C — tipo de contenido
    if upper == "URL":
        style["fill_c"] = FILL_YELLOW
        style["font_c"] = Font(name="Calibri", size=24, bold=True, color=CLR_RED)
        # El contenido (D) es la URL real: 16pt normal, sin fill especial
        style["font_d"] = Font(name="Calibri", size=16)
    elif ctype == "seo":
        style["fill_c"] = FILL_YELLOW
        style["font_c"] = FONT_11_BOLD
        # SEO image name values in D: no special fill needed
    elif ctype == "category":
        style["fill_c"] = FILL_CYAN
        style["font_c"] = FONT_11_BOLD
        # Title/subtitle/headers in D: bold
        if upper in ("TITLE", "SUB TITLE", "HEADER", "EYEBROW", "LIST HEADER", "NAMEPLATE"):
            style["font_d"] = FONT_16_BOLD
    else:
        style["fill_c"] = FILL_LIGHT_GRAY
        style["font_c"] = FONT_11_BOLD

    # Columna E — contadores
    ctd_digits = ctd.replace(",", "").replace(".", "")
    if ctd and not ctd_digits.isdigit():
        # Mixed semantics — not a pure number
        pass
    elif ctd and ctd != "0":
        style["fill_e"] = FILL_RED
        style["font_e"] = FONT_COUNTER

    # Columna F — notas
    if has_notas_orange:
        style["fill_f"] = FILL_ORANGE
    elif has_notas_yellow:
        style["fill_f"] = FILL_YELLOW

    return style


def _apply_row(ws, row_idx: int, r: dict, prev_mod: str):
    """Aplica datos y formato a una fila. Retorna si el módulo cambió."""
    vis = r["visualizacion"]
    mod = r["modulo"]
    tipo = r["tipo"]
    cont = r["contenido"]
    ctd = r["contador"]
    notas = r["notas"]
    url = r["url"]

    style = _detect_style(tipo, cont, ctd, notas)
    mod_changed = mod != prev_mod

    # Col A — Visualización (solo si cambia la sección)
    if vis:
        cell_a = ws.cell(row=row_idx, column=1, value=vis)
        cell_a.font = Font(name="Calibri", size=16, bold=True, color=CLR_WHITE)
        cell_a.fill = FILL_DARK
        cell_a.alignment = ALIGN_CENTER
        cell_a.border = _no_border

    # Col B — Módulo / Sección
    cell_b = ws.cell(row=row_idx, column=2, value=mod if mod_changed else "")
    if mod:
        cell_b.font = Font(name="Calibri", size=11, bold=True)
        cell_b.fill = FILL_MODULE
    cell_b.border = _thin_all
    cell_b.alignment = ALIGN_CENTER

    # Col C — SEO/Formato
    cell_c = ws.cell(row=row_idx, column=3, value=tipo)
    if style["font_c"]:
        cell_c.font = style["font_c"]
    else:
        cell_c.font = FONT_11_BOLD
    cell_c.fill = style.get("fill_c") or PatternFill()
    cell_c.border = _thin_all
    cell_c.alignment = ALIGN_CENTER

    # Col D — Copies
    cell_d = ws.cell(row=row_idx, column=4, value=cont)
    cell_d.font = style["font_d"]
    if style.get("fill_d"):
        cell_d.fill = style["fill_d"]
    cell_d.border = _thin_all
    cell_d.alignment = ALIGN_CENTER

    # Col E — Contadores
    cell_e = ws.cell(row=row_idx, column=5, value=ctd if ctd else None)
    if style["fill_e"]:
        cell_e.fill = style["fill_e"]
    if style["font_e"]:
        cell_e.font = style["font_e"]
    cell_e.border = _thin_all
    cell_e.alignment = ALIGN_CENTER

    # Col F — Otros/Notas
    cell_f = ws.cell(row=row_idx, column=6, value=notas if notas else None)
    if style["fill_f"]:
        cell_f.fill = style["fill_f"]
    cell_f.font = Font(name="Calibri", size=11, bold=True)
    cell_f.border = _thin_all
    cell_f.alignment = ALIGN_CENTER

    # Col G — URL de página
    cell_g = ws.cell(row=row_idx, column=7, value=url if url else None)
    if url:
        cell_g.font = Font(name="Calibri", size=10, color="666666")
    cell_g.border = _thin_all
    cell_g.alignment = ALIGN_CENTER

    return mod if mod else prev_mod


def _write_xlsx(output_path: Path, all_rows: list[dict], issues: list[dict],
                quality: dict, sheets_ok: list[str]):
    """Escribe XLSX con formato visual del Excel original + hoja Resumen."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # hoja default

    # Columnas: A=Visualización, B=Módulo, C=Tipo, D=Copies, E=Cont, F=Notas, G=URL
    col_widths = [60, 25, 22, 90, 14, 35, 45]
    headers = [
        "Visualización", "Módulo / Sección", "SEO/Formato",
        "Copies", "Contadores", "Otros/Notas/Comentarios", "URL",
    ]
    header_font = Font(name="Calibri", size=14, bold=True, color=CLR_BLACK)
    header_fill = PatternFill(start_color=CLR_HEADER_BG, end_color=CLR_HEADER_BG, fill_type="solid")

    # Agrupar filas por hoja de origen
    from collections import OrderedDict
    sheets_data: OrderedDict = OrderedDict()
    sheet_keys = list(MAPA_HOJAS.values())
    for r in all_rows:
        s = r.get("origen", "Otros")
        if s not in sheets_data:
            sheets_data[s] = []
        sheets_data[s].append(r)

    # Crear una hoja por origen
    for label, srows in sheets_data.items():
        safe_name = label[:31]  # límite de 31 chars para sheet names
        ws = wb.create_sheet(title=safe_name)

        # ── Column widths ──
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Header row ──
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = ALIGN_CENTER
            cell.border = _thin_all
        ws.row_dimensions[1].height = 30

        # ── Data rows ──
        prev_mod = ""
        prev_seccion = ""
        for i, r in enumerate(srows):
            row_idx = i + 2

            # Separador visual entre secciones
            vis = r.get("visualizacion", "")
            if vis and vis != prev_seccion:
                ws.row_dimensions[row_idx].height = 28
                prev_seccion = vis

            prev_mod = _apply_row(ws, row_idx, r, prev_mod)

        # ── Freeze header ──
        ws.freeze_panes = "A2"

        # ── Auto-filter ──
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(srows) + 1}"

    # ── Hoja: Resumen ──
    ws_res = wb.create_sheet(title="Resumen", index=0)
    ws_res.column_dimensions["A"].width = 25
    ws_res.column_dimensions["B"].width = 60
    ws_res.column_dimensions["C"].width = 20

    title_font = Font(name="Calibri", size=18, bold=True, color=CLR_DARK_BG)
    section_font = Font(name="Calibri", size=14, bold=True, color=CLR_WHITE)
    section_fill = PatternFill(start_color=CLR_DARK_BG, end_color=CLR_DARK_BG, fill_type="solid")
    normal_font = Font(name="Calibri", size=12)
    bold_font = Font(name="Calibri", size=12, bold=True)
    green_font = Font(name="Calibri", size=12, color="228B22", bold=True)
    red_font = Font(name="Calibri", size=12, color=CLR_RED, bold=True)

    row = 1
    ws_res.cell(row, 1, "RESUMEN DE ANALISIS").font = title_font
    ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # Stats
    ws_res.cell(row, 1, "Filas totales").font = bold_font
    ws_res.cell(row, 2, str(len(all_rows))).font = normal_font
    row += 1
    ws_res.cell(row, 1, "Sheets procesadas").font = bold_font
    ws_res.cell(row, 2, ", ".join(MAPA_HOJAS[s] for s in sheets_ok if s in MAPA_HOJAS)).font = normal_font
    row += 1
    ws_res.cell(row, 1, "URLs de pagina").font = bold_font
    urls = sorted(set(r["url"] for r in all_rows if r["url"]))
    ws_res.cell(row, 2, str(len(urls))).font = normal_font
    row += 2

    # Quality flags
    ws_res.cell(row, 1, "QUALITY FLAGS").font = section_font
    ws_res.cell(row, 1).fill = section_fill
    ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    total_flags = sum(quality.values())
    if total_flags == 0:
        ws_res.cell(row, 1, "Sin anomalias detectadas").font = green_font
        row += 1
    else:
        for k, v in quality.items():
            if v:
                ws_res.cell(row, 1, k).font = bold_font
                ws_res.cell(row, 2, str(v)).font = red_font if v > 0 else green_font
                row += 1
    row += 1

    # Incidencias
    ws_res.cell(row, 1, "INCIDENCIAS").font = section_font
    ws_res.cell(row, 1).fill = section_fill
    ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    if not issues:
        ws_res.cell(row, 1, "Sin incidencias").font = green_font
        row += 1
    else:
        for issue in issues:
            sev = issue["severidad"].upper()
            ws_res.cell(row, 1, f"[{sev}] {issue['tipo']}").font = bold_font
            ws_res.cell(row, 2, issue["descripcion"]).font = normal_font
            row += 1

    # URLs
    row += 1
    ws_res.cell(row, 1, "URLS POR PAGINA").font = section_font
    ws_res.cell(row, 1).fill = section_fill
    ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    for u in urls:
        ws_res.cell(row, 1, u).font = Font(name="Calibri", size=10, color="1F4E79")
        row += 1

    wb.save(str(output_path))
    print(f"  XLSX: {output_path.resolve()}  ({len(all_rows)} filas, {len(sheets_data)} hojas)")
    return True


# ── Summary JSON ──

def _write_summary(output_dir: Path, all_rows: list[dict], issues: list[dict],
                   quality: dict, ticket: str, mercado: str, pagina: str,
                   sheets_ok: list[str], ext: str):
    """Escribe resumen JSON."""
    by_sheet: dict = {}
    for r in all_rows:
        s = r.get("origen", "(sin origen)")
        by_sheet[s] = by_sheet.get(s, 0) + 1

    urls = sorted(set(r["url"] for r in all_rows if r["url"]))

    summary = {
        "archivo": f"{ticket}-{mercado}-{pagina}.{ext}",
        "ticket": ticket, "mercado": mercado, "pagina": pagina,
        "total_filas": len(all_rows),
        "sheets_procesadas": len(sheets_ok),
        "sheets": [MAPA_HOJAS.get(s, s) for s in sheets_ok],
        "urls_encontradas": urls,
        "filas_por_sheet": by_sheet,
        "calidad": {"total_flags": sum(quality.values()), "detalle": quality},
        "incidencias": [
            {"tipo": i["tipo"], "severidad": i["severidad"], "descripcion": i["descripcion"]}
            for i in issues
        ],
    }
    json_path = output_dir / f"{ticket}-{mercado}-{pagina}-summary.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    print(f"  JSON: {json_path.resolve()}")

    # Console
    print(f"\n{'='*50}")
    print(f"  RESUMEN — {ticket}-{mercado}-{pagina}")
    print(f"{'='*50}")
    print(f"  Sheets:             {len(sheets_ok)}")
    print(f"  Filas totales:      {len(all_rows)}")
    print(f"  URLs encontradas:   {len(urls)}")
    print(f"  Quality flags:      {sum(quality.values())}")
    for k, v in quality.items():
        if v:
            print(f"    {k}: {v}")
    for i in issues:
        print(f"  [{i['severidad'].upper()}] {i['descripcion']}")
    print(f"{'='*50}\n")


# ── Main ──

def main():
    parser = argparse.ArgumentParser(
        description="Analiza Nameplate Excel y genera XLSX con formato original.",
    )
    parser.add_argument("--input", default=EXCEL_PATH)
    parser.add_argument("--ticket", default=DEFAULT_TICKET)
    parser.add_argument("--mercado", default=DEFAULT_MERCADO)
    parser.add_argument("--pagina", default=DEFAULT_PAGINA)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--format", choices=["xlsx", "csv", "both"], default="xlsx",
                        help="Formato de salida (default: xlsx)")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S",
    )

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[ERR] No se encuentra: {input_path.resolve()}")
        sys.exit(1)

    output_dir = Path(args.output_dir) if args.output_dir else Path(args.ticket)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Cargar Excel ──
    print(f"\n[1/4] Leyendo Excel: {input_path.resolve()}")
    wb = openpyxl.load_workbook(str(input_path), data_only=True, read_only=True)

    # ── Extraer ──
    print("[2/4] Extrayendo datos...")
    all_rows, sheets_ok, issues = _extract_all(wb)
    wb.close()

    if not all_rows:
        print("[ERR] No se extrajeron datos")
        sys.exit(1)

    # ─── Calidad ──
    print("[3/4] Analizando calidad...")
    quality = _calc_quality(all_rows)

    # ─── Generar ──
    print("[4/4] Generando reportes...")

    ext = "xlsx"
    if args.format in ("xlsx", "both"):
        out = output_dir / f"{args.ticket}-{args.mercado}-{args.pagina}.xlsx"
        _write_xlsx(out, all_rows, issues, quality, sheets_ok)

    if args.format in ("csv", "both"):
        ext = "csv"
        out = output_dir / f"{args.ticket}-{args.mercado}-{args.pagina}.csv"
        _write_csv(out, all_rows)

    _write_summary(output_dir, all_rows, issues, quality,
                   args.ticket, args.mercado, args.pagina, sheets_ok, ext)

    print(f"[OK] Analisis completado -> {output_dir}/")


if __name__ == "__main__":
    main()

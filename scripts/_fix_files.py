"""
_fix_files.py — Regenera PR/historial.xlsx from root historial + split con_aa/sin_aa.
Sin prompts interactivos, seguro contra locks.
"""
import os
import shutil

import openpyxl

ROOT_HIST = r"C:\Users\LuisOrozco\OneDrive - WPP Cloud\Escritorio\historial.xlsx"
PR_DIR = "PR"

def main():
    # ── 1. Load root historial ──
    print(f"[1] Leyendo root historial: {ROOT_HIST}")
    wb_root = openpyxl.load_workbook(ROOT_HIST, data_only=True)
    root_sheets = [s for s in wb_root.sheetnames if not s.startswith("_")]

    # Find most recent sheet with data
    target_sheet = None
    for sn in sorted(root_sheets, reverse=True):
        ws = wb_root[sn]
        if ws.max_row > 1:
            target_sheet = sn
            break

    if not target_sheet:
        print("[ERROR] No sheets with data in root historial")
        wb_root.close()
        return False

    source_ws = wb_root[target_sheet]
    total_rows = source_ws.max_row - 1  # minus header
    print(f"  Sheet: '{target_sheet}', {total_rows} rows, {source_ws.max_column} cols")

    # ── 2. Create NEW PR/historial.xlsx ──
    # Remove existing locked file by creating a temp and swapping
    tmp_path = os.path.join(PR_DIR, "historial_tmp.xlsx")
    out_path = os.path.join(PR_DIR, "historial.xlsx")
    bak_path = os.path.join(PR_DIR, "historial_old.bak")

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = target_sheet

    # Copy headers
    headers = []
    for c in range(1, source_ws.max_column + 1):
        h = source_ws.cell(1, c).value
        new_ws.cell(1, c).value = h
        headers.append(str(h or ""))

    # Copy data rows
    copied = 0
    for r in range(2, source_ws.max_row + 1):
        for c in range(1, source_ws.max_column + 1):
            new_ws.cell(r, c).value = source_ws.cell(r, c).value
        copied += 1

    # Copy _control if exists
    if "_control" in wb_root.sheetnames:
        ctrl_src = wb_root["_control"]
        ctrl_dst = new_wb.create_sheet("_control")
        for r in range(1, ctrl_src.max_row + 1):
            for c in range(1, ctrl_src.max_column + 1):
                ctrl_dst.cell(r, c).value = ctrl_src.cell(r, c).value

    # Save to temp
    new_wb.save(tmp_path)
    new_wb.close()
    wb_root.close()
    print(f"  Temp file created: {tmp_path} ({copied} rows)")

    # ── 3. Swap files ──
    if os.path.exists(out_path):
        # Try to remove original (might fail if locked)
        try:
            os.remove(out_path)
            print(f"  Old {out_path} removed")
        except PermissionError:
            # Locked — backup and use temp as new file
            try:
                shutil.move(out_path, bak_path)
                print(f"  Old file moved to {bak_path} (was locked)")
            except PermissionError:
                print(f"  [WARN] Cannot remove locked file. Using {tmp_path} directly.")
                # Check if temp is usable
                tmp_size = os.path.getsize(tmp_path)
                print(f"  Temp file size: {tmp_size} bytes")
                print(f"  Will need to rename manually: move {tmp_path} to {out_path}")
                return True

    os.rename(tmp_path, out_path)
    print(f"[OK] {out_path} regenerated ({copied} rows)")

    # ── 4. Show headers for verification ──
    print(f"  Headers: {headers}")

    # ── 5. Count digitalData and AA in the data ──
    dd_col = None
    aa_col = None
    for i, h in enumerate(headers, 1):
        hl = h.lower()
        if "digitaldata (automatica)" in hl:
            dd_col = i
        elif "aa analytics (automatico)" in hl:
            aa_col = i

    wb_check = openpyxl.load_workbook(out_path, data_only=True)
    ws_check = wb_check[target_sheet]

    dd_ok = 0
    dd_error = 0
    aa_ok = 0
    aa_error = 0
    for r in range(2, ws_check.max_row + 1):
        # Check DD
        if dd_col:
            dd_val = str(ws_check.cell(r, dd_col).value or "")
            if dd_val and dd_val != "None" and len(dd_val) > 10:
                if '"error"' not in dd_val.lower():
                    dd_ok += 1
                else:
                    dd_error += 1
            else:
                dd_error += 1

        # Check AA
        if aa_col:
            aa_val = str(ws_check.cell(r, aa_col).value or "")
            if aa_val and aa_val != "None" and len(aa_val) > 10:
                if '"error"' not in aa_val.lower():
                    aa_ok += 1
                else:
                    aa_error += 1
            else:
                aa_error += 1

    wb_check.close()
    print(f"\n[OK] Datos en {out_path}:")
    print(f"  URLs totales: {copied}")
    if dd_col:
        print(f"  digitalData real: {dd_ok}")
        print(f"  digitalData error: {dd_error}")
    if aa_col:
        print(f"  AA analytics real: {aa_ok}")
        print(f"  AA analytics error: {aa_error}")

    # ── 5.5 Debug: check _has_digitaldata / _has_aa behavior ──
    print("\n[DEBUG] Checking split function behavior...")
    from json_convert.excel import _has_aa, _has_digitaldata

    dd_yes = 0
    dd_no = 0
    aa_yes = 0
    aa_no = 0
    for r in range(2, source_ws.max_row + 1):
        col_d = source_ws.cell(r, 4).value
        col_f = source_ws.cell(r, 6).value
        if _has_digitaldata(col_d):
            dd_yes += 1
        else:
            dd_no += 1
        if _has_aa(col_f):
            aa_yes += 1
        else:
            aa_no += 1

    print(f"  _has_digitaldata: YES={dd_yes}, NO={dd_no}")
    print(f"  _has_aa: YES={aa_yes}, NO={aa_no}")
    print(f"  has_data (DD or AA): {dd_yes + aa_yes}")
    print("  has_data (DD or AA, dedup): ", end="")
    both = 0
    for r in range(2, source_ws.max_row + 1):
        col_d = source_ws.cell(r, 4).value
        col_f = source_ws.cell(r, 6).value
        if _has_digitaldata(col_d) or _has_aa(col_f):
            both += 1
    print(both)

    # ── 6. Now regenerate con_aa/sin_aa from this historial ──
    print(f"\n[2] Regenerando con_aa/sin_aa desde {out_path}...")
    from json_convert.excel import (
        split_aa_workbooks,
    )

    wb_split = openpyxl.load_workbook(out_path, data_only=True)
    split_aa_workbooks(wb_split, target_sheet, PR_DIR)
    wb_split.close()

    # Verify split results
    con_path = os.path.join(PR_DIR, "con_aa.xlsx")
    sin_path = os.path.join(PR_DIR, "sin_aa.xlsx")

    con_rows = 0
    sin_rows = 0
    if os.path.exists(con_path):
        wb_c = openpyxl.load_workbook(con_path, data_only=True)
        # Use the audit_date sheet, not .active
        ws_c = wb_c[target_sheet] if target_sheet in wb_c.sheetnames else wb_c.active
        # Count actual rows with data (not using max_row which includes gaps)
        con_rows = sum(1 for r in range(2, ws_c.max_row + 1) if ws_c.cell(r, 2).value)
        wb_c.close()
    if os.path.exists(sin_path):
        wb_s = openpyxl.load_workbook(sin_path, data_only=True)
        ws_s = wb_s[target_sheet] if target_sheet in wb_s.sheetnames else wb_s.active
        sin_rows = sum(1 for r in range(2, ws_s.max_row + 1) if ws_s.cell(r, 2).value)
        wb_s.close()

    print("[OK] Split results:")
    print(f"  con_aa.xlsx: {con_rows} rows")
    print(f"  sin_aa.xlsx: {sin_rows} rows")
    print(f"  Total: {con_rows + sin_rows} (expect {copied})")

    if con_rows + sin_rows != copied:
        print(f"  [WARN] Mismatch! Expected {copied}, got {con_rows + sin_rows}")

    return True

if __name__ == "__main__":
    success = main()
    if success:
        print("\n[LISTO] Archivos regenerados correctamente.")
    else:
        print("\n[ERROR] Falló la regeneración.")

"""
prune_excel_columns.py — Elimina columnas inútiles de Excel generados.

Actualmente no elimina ninguna columna (todas son útiles).
Se mantiene el script como infraestructura para futuras limpiezas.

Uso:
  python prune_excel_columns.py                    # limpia todos los market/ y raíz
  python prune_excel_columns.py --dir PR           # solo un mercado
  python prune_excel_columns.py --dry-run          # muestra qué haría sin modificar
"""

import argparse
import os
import sys
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl no está instalado. Ejecutá: pip install openpyxl")
    sys.exit(1)

# Columnas a eliminar: (posicion, nombre)
# Ahora todas las columnas tienen uso real:
#   C = digitaldata (manual), D = digitaldata (automatica), E = AA, F = AA struct, G = metadata
DEAD_COLS = [
    # (3, "digitaldata (manual)"),  # ya no es muerta — columna activa
]


def find_excel_files(base_dir: str, specific_dir: str | None = None) -> list[str]:
    """Busca historial.xlsx, con_aa.xlsx, sin_aa.xlsx en market dirs y raíz."""
    targets = ["historial.xlsx", "con_aa.xlsx", "sin_aa.xlsx"]
    found = []

    if specific_dir:
        dirs = [Path(base_dir) / specific_dir]
    else:
        dirs = [d for d in Path(base_dir).iterdir() if d.is_dir() and not d.name.startswith(".")]
        dirs.append(Path(base_dir))  # también raíz

    for d in dirs:
        for fname in targets:
            fp = d / fname
            if fp.exists():
                found.append(str(fp))

    return found


def prune_file(path: str, dry_run: bool = False) -> dict:
    """Elimina columnas muertas de un Excel. Retorna stats."""
    stats = {"path": path, "removed_cols": [], "sheets": 0, "rows": 0}

    wb = openpyxl.load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_column < 3:
            continue

        # Procesar columnas en orden inverso para no afectar índices
        for col_idx, col_name in sorted(DEAD_COLS, reverse=True):
            if ws.max_column < col_idx:
                continue
            # Solo si el header coincide con la columna muerta
            header = str(ws.cell(1, col_idx).value or "").strip().lower()
            if col_name not in header:
                continue
            # Verificar que realmente está vacía
            all_empty = True
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(r, col_idx).value
                if cell_val is not None and str(cell_val).strip():
                    all_empty = False
                    break

            if all_empty or dry_run:
                if not dry_run:
                    ws.delete_cols(col_idx, 1)
                stats["removed_cols"].append(col_name)
                stats["sheets"] += 1
                stats["rows"] += ws.max_row

    if not dry_run:
        # Retry si OneDrive/Excel tiene el archivo bloqueado
        for attempt in range(4):
            try:
                wb.save(path)
                break
            except PermissionError:
                if attempt < 3:
                    time.sleep(2)
                else:
                    name, ext = os.path.splitext(path)
                    fallback = f"{name}_pruned{ext}"
                    wb.save(fallback)
                    print(f"  [!] Lock persistente, guardado como: {os.path.basename(fallback)}")
    wb.close()
    return stats


def main():
    parser = argparse.ArgumentParser(description="Elimina columnas inútiles de Excel generados")
    parser.add_argument("--dir", help="Mercado específico (ej: PR)")
    parser.add_argument("--dry-run", action="store_true", help="Solo mostrar qué haría")
    parser.add_argument("--verbose", action="store_true", help="Logging detallado")
    args = parser.parse_args()

    base = os.path.dirname(os.path.abspath(__file__))
    files = find_excel_files(base, args.dir)

    if not files:
        print("[!] No se encontraron archivos Excel para limpiar.")
        return

    total_removed = 0
    total_files = 0

    for fpath in files:
        stats = prune_file(fpath, args.dry_run)
        if stats["removed_cols"]:
            print(f"  {'[DRY]' if args.dry_run else '[OK]'} {os.path.relpath(fpath, base)}")
            for col in stats["removed_cols"]:
                print(f"         eliminada columna: {col}")
            total_removed += len(stats["removed_cols"])
            total_files += 1
        elif args.verbose:
            print(f"  [-] {os.path.relpath(fpath, base)} sin columnas muertas")

    print(f"\n{'='*55}")
    if args.dry_run:
        print(f"  SIMULACION: {total_files} archivos afectados, {total_removed} columnas a eliminar")
    else:
        print(f"  LIMPIEZA: {total_files} archivos modificados, {total_removed} columnas eliminadas")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()

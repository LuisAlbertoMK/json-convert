"""
test_gen_urls.py — Tests unitarios para gen_urls.py.

Requiere: openpyxl (no requiere playwright ni navegador).
"""

import json
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from gen_urls import extract


class TestGenUrlsExtract(unittest.TestCase):
    """Tests para extract() de _gen_urls.py."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.xlsx_path = os.path.join(self.tmpdir, "test.xlsx")

    def _make_excel(self, rows: list[tuple[str | None, str | None]]):
        """Crea un Excel de prueba con filas en col A (nombre) y col B (URL)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "nombre pagina auditada"
        ws.cell(1, 2).value = "pagina auditada (URL)"
        for i, (nombre, url) in enumerate(rows, start=2):
            if nombre is not None:
                ws.cell(i, 1).value = nombre
            if url is not None:
                ws.cell(i, 2).value = url
        wb.save(self.xlsx_path)
        wb.close()

    def tearDown(self):
        for f in os.listdir(self.tmpdir):
            try:
                os.remove(os.path.join(self.tmpdir, f))
            except PermissionError:
                pass
        os.rmdir(self.tmpdir)

    def test_extract_basic(self):
        """Extrae URLs con nombre."""
        self._make_excel([
            ("Home", "https://www.ford.com"),
            ("About", "https://www.ford.com/about"),
        ])
        entries = extract(self.xlsx_path)
        self.assertEqual(len(entries), 2)
        self.assertEqual(entries[0], {"url": "https://www.ford.com", "nombre": "Home"})
        self.assertEqual(entries[1], {"url": "https://www.ford.com/about", "nombre": "About"})

    def test_extract_without_nombre(self):
        """Celda A vacía → genera nombre desde la URL si es posible."""
        self._make_excel([
            (None, "https://www.ford.com"),
            ("", "https://www.ford.com/about"),
        ])
        entries = extract(self.xlsx_path)
        self.assertEqual(len(entries), 2)
        # Sin path → no se puede generar nombre
        self.assertEqual(entries[0], {"url": "https://www.ford.com"})
        # Con path /about → genera "About"
        self.assertEqual(entries[1], {"url": "https://www.ford.com/about", "nombre": "About"})

    def test_extract_skip_empty_urls(self):
        """Filas sin URL en col B se omiten."""
        self._make_excel([
            ("Home", "https://www.ford.com"),
            ("Empty", None),
            ("About", "https://www.ford.com/about"),
            ("Also Empty", ""),
        ])
        entries = extract(self.xlsx_path)
        self.assertEqual(len(entries), 2)

    def test_extract_default_market(self):
        """Con default_market, agrega campo market a todas."""
        self._make_excel([
            ("Home", "https://www.ford.com"),
        ])
        entries = extract(self.xlsx_path, default_market="PR")
        self.assertEqual(len(entries), 1)
        self.assertEqual(entries[0]["market"], "PR")

    def test_extract_market_case_normalized(self):
        """default_market se normaliza a uppercase."""
        self._make_excel([
            ("Home", "https://www.ford.com"),
        ])
        entries = extract(self.xlsx_path, default_market="mx")
        self.assertEqual(entries[0]["market"], "MX")

    def test_extract_no_market(self):
        """Sin default_market → no agrega market."""
        self._make_excel([
            ("Home", "https://www.ford.com"),
        ])
        entries = extract(self.xlsx_path)
        self.assertNotIn("market", entries[0])

    def test_extract_empty_excel(self):
        """Excel sin filas de datos → lista vacía."""
        self._make_excel([])
        entries = extract(self.xlsx_path)
        self.assertEqual(entries, [])

    def test_extract_whitespace_urls_trimmed(self):
        """URLs con espacios se trimean."""
        self._make_excel([
            ("Home", "  https://www.ford.com  "),
        ])
        entries = extract(self.xlsx_path)
        self.assertEqual(entries[0]["url"], "https://www.ford.com")

    def test_extract_output_json_serializable(self):
        """El resultado debe ser serializable a JSON."""
        self._make_excel([
            ("Home", "https://www.ford.com"),
            ("About", "https://www.ford.com/about"),
        ])
        entries = extract(self.xlsx_path, default_market="PR")
        dumped = json.dumps(entries, ensure_ascii=False, indent=2)
        self.assertIn("Home", dumped)
        self.assertIn("PR", dumped)
        self.assertEqual(len(json.loads(dumped)), 2)


if __name__ == "__main__":
    unittest.main(verbosity=2)

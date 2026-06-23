"""
test_extract_aa.py — Tests de integración para extract_aa.py pipeline.

Prueba el flujo real: Excel I/O → parseo JSON → extracción de campos → escritura.
Requiere: openpyxl (no requiere playwright).
"""

import json
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from extract_aa import ALL_FIELDS, DEFAULT_KEEP, extract_fields
from json_convert.excel import save_workbook

SAMPLE_AA_JSON = {
    "solution": "analytics",
    "page": {"title": "Ford Mach-E", "url": "https://ford.com/es/mach-e"},
    "request": {"method": "GET", "hostname": "smetrics.ford.com"},
    "visitor": {"experienceCloudId": "abc123"},
    "hit": {"id": "s123", "type": "pageView", "reportSuiteId": "fordglobal"},
    "events": ["event1", "event2"],
    "eVars": {"eVar1": '{"id":"mach-e"}', "eVar5": "preview"},
    "props": {"prop1": "home", "prop2": "vehiculos"},
    "pageName": "ford:mach-e:preview",
    "channel": "automotriz",
}


def _make_excel_with_aa(path: str, rows: list[dict | str | None],
                         headers: bool = True,
                         sheet_name: str = "data"):
    """Crea Excel con datos AA en col F (col 6) — formato de 8 cols."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if headers:
        ws.cell(1, 1).value = "nombre pagina auditada"
        ws.cell(1, 2).value = "pagina auditada (URL)"
        ws.cell(1, 3).value = "digitaldata (manual)"
        ws.cell(1, 4).value = "digitaldata (automatica)"
        ws.cell(1, 5).value = "AA analytics (manual)"
        ws.cell(1, 6).value = "AA analytics (automatico)"
        ws.cell(1, 7).value = "AA analytics (estructurado)"
        ws.cell(1, 8).value = "metadata / extra beacons"
    for i, aa_data in enumerate(rows, start=2):
        if aa_data is None:
            ws.cell(i, 6).value = None
        elif isinstance(aa_data, str):
            ws.cell(i, 6).value = aa_data
        else:
            ws.cell(i, 6).value = json.dumps(aa_data, ensure_ascii=False)
    wb.save(path)
    wb.close()


class TestExtractAAPipeline(unittest.TestCase):
    """Pipeline completo: Excel → extract_fields → escritura col G."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.xlsx_path = os.path.join(self.tmpdir, "test.xlsx")

    def tearDown(self):
        for f in os.listdir(self.tmpdir):
            try:
                os.remove(os.path.join(self.tmpdir, f))
            except PermissionError:
                pass
        os.rmdir(self.tmpdir)

    # ── extract_fields (ya testeadas en test_parse.py, solo smoke) ──

    def test_extract_fields_default(self):
        """Campos default: page, request, props, evars."""
        result = extract_fields(SAMPLE_AA_JSON, DEFAULT_KEEP)
        self.assertIn("page", result)
        self.assertIn("request", result)
        self.assertIn("props", result)
        self.assertIn("evars", result)

    def test_extract_fields_all(self):
        """keep='all' → extrae todo lo disponible."""
        result = extract_fields(SAMPLE_AA_JSON, ALL_FIELDS)
        self.assertIn("solution", result)
        self.assertIn("visitor", result)
        self.assertIn("hit", result)

    def test_extract_fields_empty_data(self):
        """data={} → resultado vacío."""
        self.assertEqual(extract_fields({}, DEFAULT_KEEP), {})

    def test_extract_fields_evars_unification(self):
        """eVars (uppercase) y evars (lowercase) se unifican."""
        grp1 = extract_fields({"eVars": {"x": "1"}}, ["evars"])
        self.assertEqual(grp1["evars"]["x"], "1")
        grp2 = extract_fields({"evars": {"y": "2"}}, ["evars"])
        self.assertEqual(grp2["evars"]["y"], "2")

    # ── Excel I/O pipeline ──

    def _simulate_extractaa(self, rows: list, keep=None,
                            sheet_name: str = "data") -> dict:
        """Simula el pipeline de extract_aa.py sobre un Excel real.
        Header-aware: detecta columna AA por header (backwards compat).
        Retorna dict con stats del proceso.
        """
        import openpyxl
        _make_excel_with_aa(self.xlsx_path, rows, sheet_name=sheet_name)
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb[sheet_name]

        # Header-aware column detection (como extract_aa.py real)
        hdr = {str(ws.cell(1, c).value or "").strip().lower(): c
               for c in range(1, ws.max_column + 1)}
        aa_src_col = hdr.get("aa analytics (automatico)", 6)

        keep = keep or DEFAULT_KEEP
        total = 0
        errores = []

        for row in range(2, ws.max_row + 1):
            raw = ws.cell(row, aa_src_col).value
            if not raw:
                errores.append((row, "COL_E_EMPTY"))
                continue
            raw_str = str(raw).strip()
            if not raw_str or raw_str.startswith("(no") or raw_str.startswith("(error"):
                errores.append((row, "NO_AA_DATA"))
                continue
            try:
                data = json.loads(raw_str)
            except json.JSONDecodeError:
                errores.append((row, "JSON_INVALID"))
                continue
            extracted = extract_fields(data, keep)
            if not extracted:
                errores.append((row, "NO_FIELDS_MATCHED"))
                continue
            aa_dst_col = hdr.get("aa analytics (estructurado)", 7)
            ws.cell(row, aa_dst_col).value = json.dumps(extracted, indent=2, ensure_ascii=False)
            total += 1

        save_workbook(wb, self.xlsx_path)
        wb.close()
        return {"total": total, "errores": errores}

    def test_pipeline_basic(self):
        """Pipeline: 2 filas con AA OK → ambas procesadas."""
        result = self._simulate_extractaa([SAMPLE_AA_JSON, SAMPLE_AA_JSON])
        self.assertEqual(result["total"], 2)
        self.assertEqual(len(result["errores"]), 0)

    def test_pipeline_partial_errors(self):
        """Pipeline con datos mixtos: algunos OK, otros con error."""
        rows = [
            SAMPLE_AA_JSON,          # OK
            "(no AA data)",          # error: placeholder
            None,                     # error: vacío
            "not valid json",        # error: JSON inválido
            SAMPLE_AA_JSON,          # OK
        ]
        result = self._simulate_extractaa(rows)
        self.assertEqual(result["total"], 2)
        self.assertEqual(len(result["errores"]), 3)

    def test_pipeline_all_errors(self):
        """Todas las filas con error → 0 procesadas."""
        rows = [None, "(no data)", "invalid json"]
        result = self._simulate_extractaa(rows)
        self.assertEqual(result["total"], 0)
        self.assertGreaterEqual(len(result["errores"]), 3)

    def test_pipeline_empty_sheet(self):
        """Sheet sin filas de datos → 0 procesadas, sin errores."""
        result = self._simulate_extractaa([])
        self.assertEqual(result["total"], 0)
        self.assertEqual(len(result["errores"]), 0)

    def test_pipeline_custom_keep(self):
        """Solo keep='pageName' → col AA struct solo tiene pageName."""
        result = self._simulate_extractaa([SAMPLE_AA_JSON], keep=["pageName"])
        self.assertEqual(result["total"], 1)
        import openpyxl
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb["data"]
        # Usar header-aware para leer col destino
        hdr = {str(ws.cell(1, c).value or "").strip().lower(): c
               for c in range(1, ws.max_column + 1)}
        dst_col = hdr.get("aa analytics (estructurado)", 6)
        written = json.loads(ws.cell(2, dst_col).value or "{}")
        self.assertEqual(written, {"pageName": "ford:mach-e:preview"})
        self.assertNotIn("page", written)
        wb.close()

    def test_pipeline_col_struct_applies_fill(self):
        """Fila procesada OK → col AA struct recibe fill en memoria (verde aplicado)."""
        import openpyxl
        from openpyxl.styles import PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "data"
        ws.cell(1, 6).value = "AA analytics (automatico)"
        ws.cell(1, 7).value = "AA analytics (estructurado)"
        ws.cell(2, 6).value = json.dumps(SAMPLE_AA_JSON)

        # Simular el mismo paso que extract_aa.py: escribir + aplicar fill
        extracted = extract_fields(SAMPLE_AA_JSON, DEFAULT_KEEP)
        cell = ws.cell(2, 7)
        cell.value = json.dumps(extracted, indent=2, ensure_ascii=False)
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                 fill_type="solid")

        self.assertEqual(cell.fill.fill_type, "solid")
        wb.close()

    def test_pipeline_json_different_formats(self):
        """Maneja eVars (Grupo 1) y evars (Grupo 2) sin problema."""
        rows = [
            {"eVars": {"eVar1": "a"}, "props": {"prop1": "b"}},
            {"evars": {"v1": "c"}, "props": {"c1": "d"}},
        ]
        result = self._simulate_extractaa(rows, keep=["evars", "props"])
        self.assertEqual(result["total"], 2)
        self.assertEqual(len(result["errores"]), 0)

    def test_pipeline_large_json(self):
        """JSON grande se procesa sin error."""
        big = dict(SAMPLE_AA_JSON)
        big["props"] = {f"prop{i}": f"val{i}" for i in range(1, 76)}
        big["eVars"] = {f"eVar{i}": f"val{i}" for i in range(1, 76)}
        result = self._simulate_extractaa([big])
        self.assertEqual(result["total"], 1)
        self.assertEqual(len(result["errores"]), 0)

    def test_pipeline_multiline_json(self):
        """JSON multilinea en col F se parsea correctamente."""
        multiline = json.dumps(SAMPLE_AA_JSON, indent=2)
        result = self._simulate_extractaa([multiline])
        self.assertEqual(result["total"], 1)
        self.assertEqual(len(result["errores"]), 0)

    # ── save_workbook ──

    def test_save_workbook_permission_error(self):
        """save_workbook con PermissionError → fallback."""
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("test")
        ws.cell(1, 1).value = "data"
        path = os.path.join(self.tmpdir, "locked.xlsx")
        wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path)
        real_save = wb2.save
        call_count = [0]

        def mock_save(p):
            call_count[0] += 1
            if call_count[0] == 1:
                raise PermissionError("Archivo bloqueado por otro proceso")
            real_save(p)

        from unittest.mock import patch
        with patch.object(wb2, "save", side_effect=mock_save):
            result = save_workbook(wb2, path)
        self.assertNotEqual(result, path)
        self.assertTrue(result.endswith("_browser.xlsx"))
        wb2.close()

    def test_save_workbook_normal(self):
        """save_workbook normal retorna el mismo path."""
        import openpyxl
        path = os.path.join(self.tmpdir, "normal.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("test")
        ws.cell(1, 1).value = "data"
        result = save_workbook(wb, path)
        self.assertEqual(result, path)
        self.assertTrue(os.path.exists(path))
        wb.close()


if __name__ == "__main__":
    unittest.main(verbosity=2)

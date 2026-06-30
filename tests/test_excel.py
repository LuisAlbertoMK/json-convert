"""
test_excel.py — Tests unitarios para funciones de excel.py no cubiertas.

Cubre: apply_data_fills, _has_digitaldata, _has_aa, split_aa_workbooks,
       update_vars_sheet, _is_json_error, _has_json_data, _pretty_json.

Uso:
  python -m pytest tests/test_excel.py -v
  python test_excel.py

Requiere: openpyxl.
"""

import json
import os
import sys
import tempfile
import unittest

import openpyxl

from json_convert.excel import (
    _has_aa,
    _has_digitaldata,
    _has_json_data,
    _is_json_error,
    _pretty_json,
    _safe_serialize,
    apply_data_fills,
    save_workbook,
    split_aa_workbooks,
    update_vars_sheet,
)


def _make_ws(rows: list[list]) -> openpyxl.Workbook:
    """Crea un workbook con una sheet 'test' y filas de datos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2025-01-15"
    ws.append(["url", "titulo", "dd_manual", "dd_auto", "aa_manual", "aa_auto", "aa_est", "beacons"])
    for r in rows:
        ws.append(r)
    return wb


# ═══════════════════════════════════════════════════════════════════════════
# _is_json_error / _has_json_data
# ═══════════════════════════════════════════════════════════════════════════

class TestJsonPredicates(unittest.TestCase):
    """Tests para _is_json_error() y _has_json_data()."""

    def test_is_json_error_true(self):
        """String con {"error":...} → True."""
        self.assertTrue(_is_json_error('{"error": "no data", "code": 404}'))

    def test_is_json_error_false(self):
        """String sin error → False."""
        self.assertFalse(_is_json_error('{"page": "ok"}'))

    def test_is_json_error_not_json(self):
        """String no-JSON → False."""
        self.assertFalse(_is_json_error("not json"))

    def test_has_json_data_true(self):
        """Dict sin 'error' → True."""
        self.assertTrue(_has_json_data('{"page": "ok", "id": 1}'))

    def test_has_json_data_has_error_key(self):
        """Dict con key 'error' → False."""
        self.assertFalse(_has_json_data('{"error": "no data"}'))

    def test_has_json_data_not_json(self):
        """String no-JSON → False."""
        self.assertFalse(_has_json_data("not json"))

    def test_has_json_data_empty(self):
        """String vacío → False."""
        self.assertFalse(_has_json_data(""))


# ═══════════════════════════════════════════════════════════════════════════
# _pretty_json / _safe_serialize
# ═══════════════════════════════════════════════════════════════════════════

class TestSafeSerialize(unittest.TestCase):
    """Tests para _safe_serialize() y _pretty_json()."""

    def test_serialize_dict(self):
        """_safe_serialize retorna un dict JSON-safe."""
        result = _safe_serialize({"a": 1, "b": [2, 3]})
        self.assertEqual(result["a"], 1)
        self.assertEqual(result["b"], [2, 3])

    def test_serialize_cycle(self):
        """Objeto con ciclo no debe explotar — se trunca por profundidad."""
        d = {"name": "parent"}
        d["child"] = d
        result = _safe_serialize(d)
        self.assertIn("name", str(result))
        # depth >20 (estricto) se convierte a str → navegar 21 niveles
        child = result
        for _ in range(21):
            child = child["child"]
        self.assertIsInstance(child, str)

    def test_serialize_long_list(self):
        """Lista >50 items se trunca a 50."""
        long_list = list(range(100))
        result = _safe_serialize(long_list)
        self.assertEqual(len(result), 50)

    def test_serialize_non_serializable(self):
        """Objeto no serializable → fallback a str()."""
        result = _safe_serialize(object())
        self.assertIn("object", result.lower())

    def test_pretty_json_returns_str(self):
        result = _pretty_json({"a": 1})
        self.assertIn('"a"', result)

    def test_pretty_json_fallback_on_error(self):
        """Si _safe_serialize tira, retorna str()."""
        result = _pretty_json(set())  # sets no se pueden serializar facil
        self.assertIsInstance(result, str)


# ═══════════════════════════════════════════════════════════════════════════
# _has_digitaldata
# ═══════════════════════════════════════════════════════════════════════════

class TestHasDigitalData(unittest.TestCase):
    """Tests para _has_digitaldata()."""

    def test_none(self):
        self.assertFalse(_has_digitaldata(None))

    def test_empty(self):
        self.assertFalse(_has_digitaldata(""))

    def test_placeholder_dash(self):
        self.assertFalse(_has_digitaldata("-"))

    def test_na(self):
        self.assertFalse(_has_digitaldata("N/A"))

    def test_no_digitaldata_placeholder(self):
        self.assertFalse(_has_digitaldata("(no digitaldata)"))

    def test_valid_json_dict(self):
        self.assertTrue(_has_digitaldata('{"page": "home", "id": 1}'))

    def test_valid_but_error_dict(self):
        """Dict con 'error':'no digitaldata' → False."""
        self.assertFalse(_has_digitaldata('{"error": "no digitaldata"}'))

    def test_invalid_json(self):
        """String no-JSON → False."""
        self.assertFalse(_has_digitaldata("not json"))


# ═══════════════════════════════════════════════════════════════════════════
# _has_aa
# ═══════════════════════════════════════════════════════════════════════════

class TestHasAA(unittest.TestCase):
    """Tests para _has_aa()."""

    def test_none(self):
        self.assertFalse(_has_aa(None))

    def test_empty(self):
        self.assertFalse(_has_aa(""))

    def test_placeholder_dash(self):
        self.assertFalse(_has_aa("-"))

    def test_valid_aa_has_solution(self):
        self.assertTrue(_has_aa('{"solution": "analytics", "pageName": "test"}'))

    def test_valid_no_solution(self):
        self.assertFalse(_has_aa('{"pageName": "test"}'))

    def test_has_error_key(self):
        self.assertFalse(_has_aa('{"solution": "analytics", "error": "fail"}'))

    def test_invalid_json(self):
        self.assertFalse(_has_aa("not json"))


# ═══════════════════════════════════════════════════════════════════════════
# apply_data_fills
# ═══════════════════════════════════════════════════════════════════════════

class TestApplyDataFills(unittest.TestCase):
    """Tests para apply_data_fills()."""

    def test_red_fill_on_error_in_col_c_and_d(self):
        """Col C y D con error JSON → fill rojo."""
        error_json = '{"error": "no data", "code": 500}'
        wb = _make_ws([
            ["url", "t", error_json, error_json, "", "{}", "{}"],
        ])
        ws = wb.active
        apply_data_fills(ws)
        # Col C (3) debe tener fill rojo
        self.assertEqual(ws.cell(2, 3).fill.start_color.rgb, "00FFC7CE")
        # Col D (4) debe tener fill rojo
        self.assertEqual(ws.cell(2, 4).fill.start_color.rgb, "00FFC7CE")

    def test_yellow_fill_on_valid_aa(self):
        """Col F (AA auto) con JSON válido → fill amarillo."""
        aa_json = '{"solution": "analytics", "pageName": "test"}'
        wb = _make_ws([
            ["url", "t", "", "{}", "", aa_json, ""],
        ])
        ws = wb.active
        apply_data_fills(ws)
        self.assertEqual(ws.cell(2, 6).fill.start_color.rgb, "00FFEB9C")

    def test_no_fill_on_empty(self):
        """Celdas vacías → sin fill."""
        wb = _make_ws([
            ["url", "t", "", "", "", "", "", ""],
        ])
        ws = wb.active
        apply_data_fills(ws)
        # Todas deben tener fill sin color (tema por defecto)
        for col in range(3, 9):
            fill = ws.cell(2, col).fill
            self.assertEqual(fill.fill_type, None)

    def test_skips_header_row(self):
        """Fila 1 (header) no debe recibir fill."""
        wb = _make_ws([
            ["url", "t", '{"error":"x"}', '{"error":"x"}', "", '{"solution":"a"}', '{"page":"ok"}'],
        ])
        ws = wb.active
        apply_data_fills(ws)
        for col in range(1, 9):
            fill = ws.cell(1, col).fill
            self.assertEqual(fill.fill_type, None)


# ═══════════════════════════════════════════════════════════════════════════
# split_aa_workbooks
# ═══════════════════════════════════════════════════════════════════════════

class TestSplitAAWorkbooks(unittest.TestCase):
    """Tests para split_aa_workbooks()."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_split_with_aa_and_without(self):
        """Filas con AA → con_aa.xlsx, sin_aa.xlsx con TODAS las filas (solo 4 cols)."""
        wb = _make_ws([
            ["url1", "t1", "", '{"page":"ok"}', "", '{"solution":"a"}', "", ""],  # tiene aa+dd
            ["url2", "t2", "", "", "", "", "", ""],  # no tiene nada
        ])
        ws = wb.active
        split_aa_workbooks(wb, ws.title, self.tmpdir)

        con_path = os.path.join(self.tmpdir, "con_aa.xlsx")
        sin_path = os.path.join(self.tmpdir, "sin_aa.xlsx")
        self.assertTrue(os.path.exists(con_path))
        self.assertTrue(os.path.exists(sin_path))

        # con_aa: header + row 2 (con datos)
        con_wb = openpyxl.load_workbook(con_path)
        con_ws = con_wb[ws.title]
        self.assertEqual(con_ws.cell(2, 1).value, "url1")
        self.assertEqual(con_ws.cell(2, 6).value, '{"solution":"a"}')
        # url2 NO esta en con_aa (no tiene datos)
        self.assertIsNone(con_ws.cell(3, 1).value)
        con_wb.close()

        # sin_aa: TODAS las filas, solo primeras 4 columnas
        sin_wb = openpyxl.load_workbook(sin_path)
        sin_ws = sin_wb[ws.title]
        self.assertEqual(sin_ws.max_row, 3)     # header + 2 datos
        self.assertEqual(sin_ws.max_column, 4)   # solo 4 cols
        self.assertEqual(sin_ws.cell(2, 1).value, "url1")   # url1 tambien
        self.assertEqual(sin_ws.cell(3, 1).value, "url2")   # url2 tambien
        # col 5+ debe estar vacia
        self.assertIsNone(sin_ws.cell(2, 5).value)
        sin_wb.close()

    def test_split_all_con_aa(self):
        """Todas las filas tienen AA → sin_aa tiene copia ligera de todas."""
        wb = _make_ws([
            ["url1", "t1", "", "", "", '{"solution":"a"}', "", ""],
            ["url2", "t2", "", "", "", '{"solution":"b"}', "", ""],
        ])
        ws = wb.active
        split_aa_workbooks(wb, ws.title, self.tmpdir)

        con_path = os.path.join(self.tmpdir, "con_aa.xlsx")
        sin_path = os.path.join(self.tmpdir, "sin_aa.xlsx")
        self.assertTrue(os.path.exists(con_path))
        self.assertTrue(os.path.exists(sin_path))

        con_wb = openpyxl.load_workbook(con_path)
        self.assertEqual(con_wb.active.max_row, 3)       # header + 2 filas
        con_wb.close()

        sin_wb = openpyxl.load_workbook(sin_path)
        self.assertEqual(sin_wb.active.max_row, 3)       # header + 2 filas (todas)
        self.assertEqual(sin_wb.active.max_column, 4)    # solo 4 cols
        sin_wb.close()

    def test_split_only_dd_no_aa(self):
        """Solo digitalData (col D) sin AA → con_aa igual, sin_aa con copia."""
        wb = _make_ws([
            ["url1", "t1", "", '{"page":"ok"}', "", "", "", ""],
        ])
        ws = wb.active
        split_aa_workbooks(wb, ws.title, self.tmpdir)
        con_path = os.path.join(self.tmpdir, "con_aa.xlsx")
        self.assertTrue(os.path.exists(con_path))
        con_wb = openpyxl.load_workbook(con_path)
        self.assertEqual(con_wb.active.cell(2, 4).value, '{"page":"ok"}')
        con_wb.close()
        # sin_aa tambien tiene la fila (solo 4 cols)
        sin_path = os.path.join(self.tmpdir, "sin_aa.xlsx")
        sin_wb = openpyxl.load_workbook(sin_path)
        self.assertEqual(sin_wb.active.max_row, 2)
        self.assertEqual(sin_wb.active.cell(2, 1).value, "url1")
        sin_wb.close()

    def test_split_empty_sheet(self):
        """Sheet sin filas de datos → ambos archivos con solo header."""
        wb = _make_ws([])
        ws = wb.active
        split_aa_workbooks(wb, ws.title, self.tmpdir)
        con_path = os.path.join(self.tmpdir, "con_aa.xlsx")
        sin_path = os.path.join(self.tmpdir, "sin_aa.xlsx")
        self.assertTrue(os.path.exists(con_path))
        self.assertTrue(os.path.exists(sin_path))


# ═══════════════════════════════════════════════════════════════════════════
# update_vars_sheet
# ═══════════════════════════════════════════════════════════════════════════

class TestUpdateVarsSheet(unittest.TestCase):
    """Tests para update_vars_sheet()."""

    def test_creates_sheet_with_vars(self):
        """Sheet _vars se crea con eVars/props deduplicados."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("2025-01-15")
        rows_aa = [
            (2, {"eVars": {"eVar1": "val1", "eVar2": "val2"}, "props": {"prop1": "v1"}}),
        ]
        update_vars_sheet(wb, "2025-01-15", rows_aa)
        vs = wb["_vars"]
        self.assertEqual(vs.cell(1, 1).value, "variable")
        # Debe tener 3 variables (eVar1, eVar2, prop1) + header = 4 filas
        self.assertEqual(vs.max_row, 4)

    def test_replaces_existing(self):
        """Si _vars ya existe, se reemplaza."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("2025-01-15")
        vs = wb.create_sheet("_vars")
        vs.append(["variable", "descripcion"])
        vs.append(["old_var", ""])

        rows_aa = [(2, {"eVars": {"eVar1": "val1"}})]
        update_vars_sheet(wb, "2025-01-15", rows_aa)
        vs = wb["_vars"]
        # Solo debe tener eVar1 (old_var debe haber sido borrado)
        self.assertEqual(vs.max_row, 2)  # header + eVar1

    def test_deduplicates_vars(self):
        """Misma variable en múltiples filas → solo una aparición."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("2025-01-15")
        rows_aa = [
            (2, {"eVars": {"eVar1": "a"}}),
            (3, {"eVars": {"eVar1": "b"}}),
        ]
        update_vars_sheet(wb, "2025-01-15", rows_aa)
        vs = wb["_vars"]
        self.assertEqual(vs.max_row, 2)  # header + eVar1 sola vez


if __name__ == "__main__":
    unittest.main(verbosity=2)

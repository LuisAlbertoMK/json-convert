#!/usr/bin/env python3
"""
generate_migration_catalog.py — Genera catálogo de migración AEM.

Compara los valores REALES de digitalData (capturados por extract_browser.py)
contra el estándar ESPERADO (definido en expected.json) y genera un catálogo
con JSON pretty-printed por URL y acciones detalladas.

Uso:
  # 1. Generar template de url-mapping desde RevisionManual.xlsx
  python generate_migration_catalog.py --gen-template

  # 2. Generar catálogo (después de llenar url-mapping.json)
  python generate_migration_catalog.py \
      --historial PR/historial.xlsx \
      --mapping url-mapping.json \
      --expected expected.json \
      --market PR

Flujo completo:
  RevisionManual.xlsx
    → _gen_urls.py → urls.json
    → extract_browser.py → {market}/historial.xlsx
    → generate_migration_catalog.py → {market}/catalogo-migracion.xlsx
"""

import json
import argparse
import os
import sys
import re
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Constantes ──
PARAMS_ORDER = ["pageName", "siteSection", "pageNameNoVehicle",
                "client", "site", "variantName", "pageType"]

COLS = [
    "URL de Producción",
    "Path de AEM Migrado",
    "Parámetro (digitalData.page)",
    "Valor en Producción (Actual)",
    "Valor Esperado en AEM (Nuevo Componente)",
    "Estado / Acción para Authoring",
]
COL_WIDTHS = [45, 45, 30, 50, 50, 60]

# ── Estilos ──
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_DATA = Font(name="Consolas", size=9)  # Monospace for JSON
FONT_PARAM = Font(name="Calibri", size=10)
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
THICK_TOP = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="medium", color="4472C4"),
    bottom=Side(style="thin", color="D9D9D9"),
)

FILL_OK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_WARN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_CREATE = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
FILL_REMOVE = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_JSON_BG = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# ── Emojis (solo para Excel, no para print) ──
EMOJI = {"ok": "✅", "warn": "⚠️", "create": "❌", "remove": "🗑️"}


# ════════════════════════════════════════════
#  CORE
# ════════════════════════════════════════════

def load_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def find_digitaldata_in_historial(historial_path: str, preview_url: str) -> dict:
    """Busca el digitalData de una URL en el historial."""
    wb = openpyxl.load_workbook(historial_path, data_only=True)
    data_sheet = None
    for sn in wb.sheetnames:
        if sn not in ("_control", "_vars"):
            data_sheet = wb[sn]
            break
    if not data_sheet:
        wb.close()
        return {}

    for row in range(2, data_sheet.max_row + 1):
        url_cell = data_sheet.cell(row, 2).value
        if url_cell and preview_url in str(url_cell):
            dd_raw = data_sheet.cell(row, 4).value
            wb.close()
            if dd_raw and isinstance(dd_raw, str):
                try:
                    return json.loads(dd_raw)
                except json.JSONDecodeError:
                    return {}
            return {}
    wb.close()
    return {}


def resolve_expected(param_name: str, page_key: str, cfg: dict) -> tuple:
    """
    Resuelve el valor esperado para un parámetro según el page_key.

    Retorna (expected_value, action, detail).
    action: ok | warn | create | remove
    """
    param_cfg = cfg.get("params", {}).get(param_name, {})
    rule = param_cfg.get("rule", "pattern")

    # No tenemos actual acá, solo resolvemos el expected
    if rule == "deprecated":
        return (None, "remove", param_cfg.get("note", "Eliminar"))

    if rule == "fixed":
        return (param_cfg.get("value", ""), "warn",
                param_cfg.get("note", "Valor fijo requerido"))

    if rule == "mirror":
        return ("(ver pageName)", "warn",
                param_cfg.get("note", "Alinear con pageName"))

    if rule == "required":
        val = param_cfg.get("mapping", {}).get(page_key, "")
        if not val:
            return ("", "create", "Sin regla")
        return (val, "create", param_cfg.get("note", "Nuevo parámetro"))

    if rule == "mapping":
        val = param_cfg.get("mapping", {}).get(page_key, "")
        if not val:
            return ("", "create", "Sin regla")
        return (val, "warn", "Mapear sección")

    # pattern (default)
    val = param_cfg.get("patterns", {}).get(page_key, "")
    if not val:
        return ("", "create", "Sin regla")
    return (val, "warn", param_cfg.get("default_note", "Alinear nomenclatura"))


def evaluate_param(param_name: str, actual: str | None,
                   page_key: str, cfg: dict) -> tuple:
    """
    Evalúa un parámetro individual contra el estándar.

    Retorna (expected_value, action_emoji, action_detail).
    """
    expected, base_action, detail = resolve_expected(param_name, page_key, cfg)

    if base_action == "remove":
        if actual is not None:
            return ("🗑️ Eliminar", "🗑️", detail)
        return ("—", "✅", "No aplica (ya eliminado)")

    if actual is None:
        return (f"❌ Crear: {expected}", "❌",
                f"Crear: {detail}")

    if str(actual) == str(expected):
        return (f"✅ {actual}", "✅", "Componente alineado")

    # Diferente: determinar detalle específico
    if param_name == "pageName":
        prefix = cfg.get("prefix", "")
        if not str(actual).startswith(prefix):
            return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                    f"Agregar prefijo '{prefix}'")
        return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                f"Alinear: {detail}")

    if param_name == "siteSection":
        return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                f"Sección legacy '{actual}' → '{expected}'")

    if param_name == "site":
        return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
                f"Unificar sitio '{actual}' → '{expected}'")

    return (f"⚠️ Cambiar: {actual} → {expected}", "⚠️",
            f"Modificar: {detail}")


def build_expected_page(page_key: str, cfg: dict) -> dict:
    """Construye el dict digitalData.page esperado según las reglas."""
    expected = {}
    for param in PARAMS_ORDER:
        val, action, _ = resolve_expected(param, page_key, cfg)
        if action == "remove":
            continue  # no incluir deprecated
        if val is not None and val != "":
            expected[param] = val
    return expected


def compare_page(actual_page: dict, expected_page: dict, cfg: dict,
                 page_key: str) -> list[dict]:
    """Compara parámetro por parámetro y genera acciones."""
    results = []
    for param in PARAMS_ORDER:
        actual_val = actual_page.get(param)
        actual_str = str(actual_val) if actual_val is not None else None

        expected_val, base_action, detail = resolve_expected(param, page_key, cfg)

        action_emoji = "✅"
        action_text = "Componente alineado"

        if base_action == "remove":
            if actual_val is not None:
                action_emoji = "🗑️"
                action_text = f"Eliminar: {detail}"
                expected_str = "—"
            else:
                action_emoji = "✅"
                action_text = f"No aplica ({detail})"
                expected_str = "—"
            results.append({
                "param": param,
                "actual": actual_str,
                "expected": expected_str,
                "action_emoji": action_emoji,
                "action_text": action_text,
            })
            continue

        expected_str = str(expected_val) if expected_val is not None else ""

        if actual_str is None:
            action_emoji = "❌"
            action_text = f"Crear: {detail}"
        elif actual_str == expected_str:
            action_emoji = "✅"
            action_text = "Componente alineado"
        else:
            action_emoji = "⚠️"
            if param == "pageName":
                prefix = cfg.get("prefix", "")
                if not actual_str.startswith(prefix):
                    action_text = f"Agregar prefijo '{prefix}'"
                else:
                    action_text = f"Alinear nomenclatura"
            elif param == "siteSection":
                action_text = f"Sección legacy → '{expected_str}'"
            elif param == "site":
                action_text = f"Unificar sitio → '{expected_str}'"
            else:
                action_text = f"Alinear: {detail}"

        results.append({
            "param": param,
            "actual": actual_str,
            "expected": expected_str,
            "action_emoji": action_emoji,
            "action_text": action_text,
        })
    return results


# ════════════════════════════════════════════
#  OUTPUT
# ════════════════════════════════════════════

def pp_json(obj: dict) -> str:
    """Pretty-print JSON con indentación para celdas Excel."""
    if not obj:
        return "{ }"
    return json.dumps(obj, indent=2, ensure_ascii=False)


def generate_catalog(historial_path: str, mapping_path: str,
                     expected_path: str, market: str, output_path: str):
    """Genera el catálogo de migración completo."""
    expected_cfg = load_json(expected_path)
    mappings = load_json(mapping_path)
    market_cfg = expected_cfg.get("markets", {}).get(market, {})

    if not market_cfg:
        print(f"[ERR] Mercado '{market}' no encontrado en {expected_path}")
        sys.exit(1)

    # Crear workbook
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = f"Catálogo Migración {market}"

    # Headers (fila 1)
    for col_idx, header in enumerate(COLS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ─── Datos ───
    row_idx = 2

    for mapping in mappings:
        preview_url = mapping.get("preview_url", "")
        production_url = mapping.get("production_url", preview_url)
        aem_path = mapping.get("aem_path", "")
        page_key = mapping.get("page_key", "")
        nombre = mapping.get("nombre", "")

        if not page_key:
            print(f"  [!] Saltando URL sin page_key: {preview_url}")
            continue

        # Obtener digitalData real
        dd = find_digitaldata_in_historial(historial_path, preview_url)
        actual_page = dd.get("page", {}) if dd else {}

        if dd:
            print(f"  [+] {preview_url[:55]}... ({nombre})")
        else:
            print(f"  [!] Sin digitalData: {preview_url[:55]}... ({nombre})")

        # Construir expected
        expected_page = build_expected_page(page_key, market_cfg)

        # Comparar
        param_results = compare_page(actual_page, expected_page,
                                     market_cfg, page_key)

        # --- FILA 1: URL + JSON actual ---
        actual_json = pp_json(actual_page) if actual_page else "{ }"
        expected_json = pp_json(expected_page)

        # Escribir celdas de la URL (cols 1-2, se combinan para esta URL)
        ws.cell(row_idx, 1, production_url).font = FONT_DATA
        ws.cell(row_idx, 1).alignment = ALIGN_WRAP
        ws.cell(row_idx, 1).border = THICK_TOP

        ws.cell(row_idx, 2, aem_path).font = FONT_DATA
        ws.cell(row_idx, 2).alignment = ALIGN_WRAP
        ws.cell(row_idx, 2).border = THICK_TOP

        ws.cell(row_idx, 3, "📦 digitalData.page (completo)").font = Font(
            name="Calibri", size=10, bold=True)
        ws.cell(row_idx, 3).alignment = ALIGN_WRAP
        ws.cell(row_idx, 3).border = THICK_TOP

        # Actual JSON
        cell_actual = ws.cell(row_idx, 4, actual_json)
        cell_actual.font = Font(name="Consolas", size=9)
        cell_actual.alignment = ALIGN_WRAP
        cell_actual.fill = FILL_JSON_BG
        cell_actual.border = THICK_TOP

        # Expected JSON
        cell_expected = ws.cell(row_idx, 5, expected_json)
        cell_expected.font = Font(name="Consolas", size=9)
        cell_expected.alignment = ALIGN_WRAP
        cell_expected.fill = FILL_JSON_BG
        cell_expected.border = THICK_TOP

        ws.cell(row_idx, 6, "").border = THICK_TOP

        # Alto de fila generoso para el JSON
        max_lines = max(len(actual_json.split("\n")),
                        len(expected_json.split("\n")))
        ws.row_dimensions[row_idx].height = max(60, min(max_lines * 15, 400))

        row_idx += 1

        # --- FILAS SIGUIENTES: un parámetro por fila ---
        for pr in param_results:
            ws.cell(row_idx, 1, "").border = THIN_BORDER
            ws.cell(row_idx, 2, "").border = THIN_BORDER

            cell_param = ws.cell(row_idx, 3, pr["param"])
            cell_param.font = FONT_PARAM
            cell_param.alignment = ALIGN_WRAP
            cell_param.border = THIN_BORDER

            cell_act = ws.cell(row_idx, 4, pr["actual"] if pr["actual"] else "No existe")
            cell_act.font = FONT_PARAM
            cell_act.alignment = ALIGN_WRAP
            cell_act.border = THIN_BORDER

            cell_exp = ws.cell(row_idx, 5, pr["expected"])
            cell_exp.font = FONT_PARAM
            cell_exp.alignment = ALIGN_WRAP
            cell_exp.border = THIN_BORDER

            action_text = f"{pr['action_emoji']} {pr['action_text']}"
            cell_action = ws.cell(row_idx, 6, action_text)
            cell_action.font = FONT_PARAM
            cell_action.alignment = ALIGN_WRAP
            cell_action.border = THIN_BORDER

            # Color según acción
            fill = FILL_OK
            if pr["action_emoji"] == "⚠️":
                fill = FILL_WARN
            elif pr["action_emoji"] == "❌":
                fill = FILL_CREATE
            elif pr["action_emoji"] == "🗑️":
                fill = FILL_REMOVE

            cell_action.fill = fill
            ws.row_dimensions[row_idx].height = 22

            row_idx += 1

        # Fila separadora entre URLs
        row_idx += 1

    # ─── Fila de firma ───
    row_idx += 1
    total_expected = sum(
        1 + len(PARAMS_ORDER) for m in mappings if m.get("page_key")
    )
    ws.cell(row_idx, 1,
            f"Generado: {date.today().isoformat()} | "
            f"Mercado: {market} | "
            f"URLs: {len(mappings)}").font = Font(italic=True, color="999999")

    # Freeze pane
    ws.freeze_panes = "A2"

    # Auto-filter en headers
    ws.auto_filter.ref = f"A1:F{row_idx - 1}"

    # Guardar
    out_wb.save(output_path)
    print(f"\n[OK] Catalogo generado: {output_path}")
    print(f"     {len(mappings)} URLs × {len(PARAMS_ORDER)} parámetros")


# ════════════════════════════════════════════
#  CLI
# ════════════════════════════════════════════

def generate_template(historial_path: str, output: str):
    """Genera url-mapping.json template desde RevisionManual.xlsx."""
    wb = openpyxl.load_workbook(historial_path, data_only=True)
    ws = wb.active

    mappings = []
    for row in range(2, ws.max_row + 1):
        nombre = ws.cell(row, 1).value
        url = ws.cell(row, 2).value
        if not url:
            continue
        url = str(url).strip()
        if not url:
            continue

        name_str = str(nombre).strip() if nombre else ""
        suggested_key = re.sub(r'[^a-z0-9]+', '-', name_str.lower()).strip('-')
        if not suggested_key:
            suggested_key = "page"

        mappings.append({
            "preview_url": url,
            "production_url": "",
            "aem_path": "",
            "page_key": suggested_key,
            "nombre": name_str
        })

    wb.close()

    with open(output, "w", encoding="utf-8") as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)

    print(f"[OK] Template generado: {output}")
    print(f"     {len(mappings)} URLs. Editar production_url, aem_path y page_key.")
    return mappings


def main():
    parser = argparse.ArgumentParser(
        description="Genera catálogo de migración AEM",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)

    parser.add_argument("--historial", default="PR/historial.xlsx",
                        help="Ruta al historial.xlsx del pipeline")
    parser.add_argument("--mapping", default="url-mapping.json",
                        help="JSON mapeo preview→producción")
    parser.add_argument("--expected", default="expected.json",
                        help="JSON con valores esperados por mercado")
    parser.add_argument("--market", default="PR",
                        help="Código de mercado (PR, MX...)")
    parser.add_argument("--output", default=None,
                        help="Ruta de salida (default: {market}/catalogo-migracion.xlsx)")
    parser.add_argument("--gen-template", action="store_true",
                        help="Generar url-mapping.json template")
    parser.add_argument("--input", default="RevisionManual.xlsx",
                        help="Excel para --gen-template")

    args = parser.parse_args()

    if args.gen_template:
        generate_template(args.input, args.mapping)
        return

    # Validar archivos
    for fpath, fname in [(args.historial, "historial"),
                         (args.mapping, "mapping"),
                         (args.expected, "expected")]:
        if not os.path.exists(fpath):
            print(f"[ERR] Archivo {fname} no encontrado: {fpath}")
            print(f"  Usar --gen-template para crear el mapping")
            sys.exit(1)

    # Output default: {market}/catalogo-migracion.xlsx
    if args.output is None:
        market_dir = args.market.upper()
        os.makedirs(market_dir, exist_ok=True)
        output_path = os.path.join(market_dir, "catalogo-migracion.xlsx")
    else:
        output_path = args.output
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    generate_catalog(args.historial, args.mapping,
                     args.expected, args.market, output_path)


if __name__ == "__main__":
    main()

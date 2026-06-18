"""
analisis_urls.py — Análisis HTTP de URLs de produccion (sin Playwright).

Workflow:
  1. Lee urls.json del proyecto
  2. Filtra solo URLs con entorno="produccion"
  3. HTTP GET a cada URL (stdlib, sin navegador)
  4. Analiza: status, digitalData, GTM, GA4, meta tags, redirects
  5. Genera reporte Excel + resumen JSON

Uso:
  python analisis_urls.py                          # Todas las URLs produccion
  python analisis_urls.py --urls urls.json          # explícito
  python analisis_urls.py --market PR               # filtrar por mercado
  python analisis_urls.py --output reporte.xlsx     # output personalizado
  python analisis_urls.py --workers 5               # concurrencia
  python analisis_urls.py --verbose                 # logging debug

Dependencias: openpyxl (ya incluida en requirements.txt)
"""

import argparse
import json
import logging
import re
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# ── Constantes ──
TIMEOUT = 30
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
MAX_REDIRECTS = 5
DEFAULT_WORKERS = 3
OUTPUT_JSON = "reporte-analisis.json"
OUTPUT_XLSX = "reporte-analisis.xlsx"

# ── Logging ──
logger = logging.getLogger("analisis_urls")

# ── Patrones de busqueda en HTML ──
PATTERNS = {
    "digitaldata": re.compile(
        r"window\.digitaldata|"
        r"digitalData\s*=|"
        r"digitalData\s*;|"
        r"\"digitalData\"",
        re.IGNORECASE,
    ),
    "gtm": re.compile(
        r"googletagmanager\.com/gtm\.js|"
        r"GTM-|"
        r"google_tag_manager",
        re.IGNORECASE,
    ),
    "ga4": re.compile(
        r"gtag/gtag\.js|"
        r"G-[A-Z0-9]+|"
        r"google-analytics\.com/g/collect",
        re.IGNORECASE,
    ),
}

META_TAGS = ["title", "description", "og:title", "og:description", "og:image", "twitter:card"]


def analyze_url(url: str) -> dict:
    """HTTP GET + analisis basico de una URL."""
    result = {
        "url": url,
        "status": None,
        "status_text": "",
        "final_url": url,
        "redirect_count": 0,
        "content_type": "",
        "content_length": 0,
        "error": "",
        "digitaldata": False,
        "gtm": False,
        "ga4": False,
        "meta_tags": {},
        "load_time_ms": 0,
        "has_doctype": False,
        "html_size_kb": 0,
    }

    start = time.perf_counter()
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
            elapsed = int((time.perf_counter() - start) * 1000)
            result["status"] = resp.status
            result["status_text"] = _status_text(resp.status)
            result["final_url"] = resp.url
            result["content_type"] = resp.headers.get("Content-Type", "")
            result["content_length"] = resp.headers.get("Content-Length", 0)
            result["load_time_ms"] = elapsed

            if result["redirect_count"] == 0 and url != resp.url:
                result["redirect_count"] = 1

            html = resp.read()
            result["html_size_kb"] = round(len(html) / 1024, 1)
            result["has_doctype"] = html[:100].lower().startswith(b"<!doctype")

            _analyze_html(html.decode("utf-8", errors="replace"), result)

    except urllib.error.HTTPError as e:
        result["status"] = e.code
        result["status_text"] = _status_text(e.code)
        result["error"] = str(e)
        try:
            html = e.read()
            _analyze_html(html.decode("utf-8", errors="replace"), result)
        except Exception:
            pass

    except urllib.error.URLError as e:
        result["error"] = str(e.reason) if hasattr(e, "reason") else str(e)

    except Exception as e:
        result["error"] = str(e)

    return result


def _analyze_html(html: str, result: dict):
    """Busca patrones y extrae meta tags del HTML."""
    for key, pat in PATTERNS.items():
        result[key] = bool(pat.search(html))

    # Meta tags por regex (sin parser externo)
    for tag in META_TAGS:
        # <title>...</title>
        if tag == "title":
            m = re.search(r"<title[^>]*>(.*?)</title>", html, re.IGNORECASE | re.DOTALL)
            if m:
                result["meta_tags"][tag] = m.group(1).strip()
        # <meta name="..." content="..."> o <meta property="..." content="...">
        else:
            for attr in ("name", "property"):
                pat_meta = re.compile(
                    rf'<meta\s+{attr}=["\']{re.escape(tag)}["\']\s+content=["\'](.*?)["\']',
                    re.IGNORECASE,
                )
                m = pat_meta.search(html)
                if m:
                    result["meta_tags"][tag] = m.group(1).strip()
                    break
            # Si no se encontro, no se agrega


def _status_text(code: int) -> str:
    """Texto corto para codigo HTTP."""
    return {
        200: "OK",
        201: "Creado",
        301: "Movido permanentemente",
        302: "Redireccion temporal",
        303: "See Other",
        304: "No modificado",
        307: "Redireccion temporal",
        308: "Redireccion permanente",
        400: "Solicitud incorrecta",
        401: "No autorizado",
        403: "Prohibido",
        404: "No encontrado",
        410: "Eliminado",
        429: "Demasiadas solicitudes",
        500: "Error interno del servidor",
        502: "Bad Gateway",
        503: "Servicio no disponible",
        504: "Gateway Timeout",
    }.get(code, f"Codigo {code}")


def load_urls(path: str = "urls.json") -> list[dict]:
    """Carga URLs desde JSON y filtra solo produccion."""
    p = Path(path)
    if not p.exists():
        print(f"[ERR] No se encuentra: {path}")
        sys.exit(1)
    with p.open(encoding="utf-8") as f:
        all_urls = json.load(f)
    return [u for u in all_urls if u.get("entorno") == "produccion"]


def save_json(results: list[dict], path: str = OUTPUT_JSON):
    """Guarda resultados como JSON."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"  JSON: {path}")


def save_excel(results: list[dict], path: str = OUTPUT_XLSX):
    """Genera reporte Excel con resultados del analisis."""
    if openpyxl is None:
        print("  [WARN] openpyxl no instalado, no se genera Excel")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis URLs"

    # ── Headers ──
    headers = [
        "URL",
        "Status",
        "Final URL",
        "Redirects",
        "Content-Type",
        "HTML (KB)",
        "Tiempo (ms)",
        "digitalData",
        "GTM",
        "GA4",
        "Title",
        "Description",
        "Error",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── Colores de status ──
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # ── Datos ──
    for row, r in enumerate(results, 2):
        ws.cell(row, 1, r["url"])
        ws.cell(row, 2, f"{r['status']} {r['status_text']}" if r["status"] else "ERROR")
        ws.cell(row, 3, r["final_url"])
        ws.cell(row, 4, r["redirect_count"])
        ws.cell(row, 5, r["content_type"][:80])
        ws.cell(row, 6, r["html_size_kb"])
        ws.cell(row, 7, r["load_time_ms"])
        ws.cell(row, 8, "✅" if r["digitaldata"] else "❌")
        ws.cell(row, 9, "✅" if r["gtm"] else "❌")
        ws.cell(row, 10, "✅" if r["ga4"] else "❌")

        meta = r.get("meta_tags", {})
        ws.cell(row, 11, meta.get("title", ""))
        ws.cell(row, 12, meta.get("description", ""))
        ws.cell(row, 13, r["error"])

        # Colorear fila segun status
        if r["error"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row, col).fill = red_fill
        elif r["status"] and 200 <= r["status"] < 300:
            for col in range(1, len(headers) + 1):
                ws.cell(row, col).fill = green_fill
        elif r["status"] and 300 <= r["status"] < 400:
            for col in range(1, len(headers) + 1):
                ws.cell(row, col).fill = yellow_fill
        elif r["status"] and r["status"] >= 400:
            for col in range(1, len(headers) + 1):
                ws.cell(row, col).fill = red_fill

    # ── Ancho columnas ──
    col_widths = [50, 20, 50, 10, 30, 10, 12, 12, 10, 10, 40, 60, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Fijar header ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"

    wb.save(path)
    print(f"  Excel: {path}")


def print_summary(results: list[dict]):
    """Resumen de resultados en consola."""
    total = len(results)
    ok = sum(1 for r in results if r["status"] and 200 <= r["status"] < 300)
    redirects = sum(1 for r in results if r["status"] and 300 <= r["status"] < 400)
    errors_http = sum(1 for r in results if r["status"] and r["status"] >= 400)
    errors_net = sum(1 for r in results if r["error"] and not r["status"])
    dd_count = sum(1 for r in results if r["digitaldata"])
    gtm_count = sum(1 for r in results if r["gtm"])
    ga4_count = sum(1 for r in results if r["ga4"])

    print(f"\n{'='*50}")
    print("  RESUMEN ANALISIS URLs")
    print(f"{'='*50}")
    print(f"  Total URLs:        {total}")
    print(f"  200 OK:            {ok}")
    print(f"  Redirecciones:     {redirects}")
    print(f"  Errores HTTP:      {errors_http}")
    print(f"  Errores de red:    {errors_net}")
    print(f"  {''.join(['-']*20)}")
    print(f"  digitalData:       {dd_count} / {total} ({dd_count/total*100:.0f}%)" if total else "  digitalData:       0")
    print(f"  GTM:               {gtm_count} / {total} ({gtm_count/total*100:.0f}%)" if total else "  GTM:               0")
    print(f"  GA4:               {ga4_count} / {total} ({ga4_count/total*100:.0f}%)" if total else "  GA4:               0")
    print(f"{'='*50}")


def main():
    parser = argparse.ArgumentParser(
        description="Analisis HTTP de URLs de produccion (sin Playwright)"
    )
    parser.add_argument("--urls", default="urls.json", help="Path a urls.json")
    parser.add_argument("--market", default=None, help="Filtrar por mercado")
    parser.add_argument("--output", default=OUTPUT_XLSX, help="Path del Excel de salida")
    parser.add_argument("--json", default=OUTPUT_JSON, help="Path del JSON de salida")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS, help="Concurrencia")
    parser.add_argument("--verbose", action="store_true", help="Logging debug")
    parser.add_argument("--no-excel", action="store_true", help="Solo JSON, sin Excel")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    # ── Cargar URLs ──
    urls_data = load_urls(args.urls)
    if args.market:
        urls_data = [u for u in urls_data if u.get("market", "").upper() == args.market.upper()]

    if not urls_data:
        print("[WARN] No hay URLs de produccion para analizar")
        return

    urls = [u["url"] for u in urls_data]
    total = len(urls)
    print(f"\nAnalizando {total} URLs de produccion ({args.workers} workers)...\n")

    # ── Ejecutar analisis ──
    results = []
    done = 0
    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(analyze_url, url): url for url in urls}
        for future in as_completed(futures):
            url = futures[future]
            done += 1
            try:
                result = future.result()
                status_display = result["status"] or "ERR"
                logger.info(f"[{done}/{total}] {status_display} {url}")
                results.append(result)
            except Exception as e:
                logger.error(f"[{done}/{total}] EXCEPTION {url}: {e}")
                results.append({"url": url, "error": str(e)})

    # Reordenar en el orden original
    url_order = {u: i for i, u in enumerate(urls)}
    results.sort(key=lambda r: url_order.get(r["url"], 999))

    # ── Guardar ──
    save_json(results, args.json)
    if not args.no_excel:
        save_excel(results, args.output)

    print_summary(results)
    print(f"\n[OK] Analisis completado. {len(results)} URLs procesadas.")


if __name__ == "__main__":
    main()

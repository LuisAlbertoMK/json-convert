"""
audit_report.py — Reporte de URLs fallidas desde archivos de auditoría.

Escanea archivos historial.xlsx en directorios de mercado (PR/, MX/, etc.),
identifica qué páginas NO están funcionando y genera un Excel consolidado
con el ÚLTIMO estado de cada URL (sin duplicados).

Uso:
  python audit_report.py                                          # busca en todos los directorios
  python audit_report.py --dir PR                                 # solo un mercado
  python audit_report.py --dir PR --dir MX                        # mercados específicos
  python audit_report.py --input PR/historial.xlsx                # archivo directo
  python audit_report.py --output reporte_fallos.xlsx             # custom output

Formato del reporte (Excel):
  - Una hoja "Resumen" con la fecha de generación y stats
  - Una hoja "Fallos" con las URLs que NO funcionan (último estado)
  - Una hoja "Todos" con el estado completo de todas las URLs
  - Sin duplicados: por cada URL solo aparece su estado más reciente
"""

import argparse
import io
import json
import os
import sys
from datetime import datetime
from pathlib import Path

# Force UTF-8 stdout for Windows cp1252 terminal
if sys.stdout.encoding and sys.stdout.encoding.upper() not in ("UTF-8", "CP65001"):
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("[ERROR] openpyxl no está instalado. Ejecutá: pip install openpyxl")
    sys.exit(1)

# ── Estilos ──
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Columnas del reporte de salida
OUTPUT_HEADERS = [
    "nombre pagina", "URL", "mercado", "estado", "detalle error",
    "fecha ultima revision", "score", "digitaldata", "hoja origen",
]

ERROR_STATES = {"NO_AA_DATA", "TIMEOUT", "HTTP_403", "HTTP_ERROR",
                "URL_INVALID", "NETWORK_ERROR", "NAV_ERROR", "UNKNOWN"}


def find_historial_files(base_dir: str, specific_dirs: list[str] = None) -> list[tuple[str, str]]:
    """Busca archivos historial.xlsx en directorios de mercado.
    
    Returns:
        Lista de (ruta_completa, nombre_mercado)
    """
    results = []
    base = Path(base_dir)

    if specific_dirs:
        dirs = [base / d for d in specific_dirs]
    else:
        # Auto-detectar directorios con historial.xlsx
        dirs = [d for d in base.iterdir() if d.is_dir() and not d.name.startswith(".")]

    for d in dirs:
        candidates = [
            d / "historial.xlsx",
            d / "con_aa.xlsx",
            d / "sin_aa.xlsx",
        ]
        for hpath in candidates:
            if hpath.exists():
                results.append((str(hpath), d.name.upper()))
                break  # Solo el primero por directorio

    # También buscar en la raíz
    root_hist = base / "historial.xlsx"
    if root_hist.exists():
        results.append((str(root_hist), "ROOT"))

    return results


def parse_meta_col(val: str) -> dict:
    """Parsea el JSON de metadata (col G). Si falla, devuelve dict vacío."""
    if not val or not isinstance(val, str):
        return {}
    try:
        return json.loads(val)
    except (json.JSONDecodeError, ValueError):
        return {}


def determine_status(col_e: str, col_d: str, meta: dict) -> tuple[str, str, int]:
    """Determina si una URL está funcionando.
    
    Returns:
        (estado: "OK"|"FALLO"|"SIN_DATOS", detalle_error: str, score: int)
    """
    score = meta.get("score", 0)
    error_code = meta.get("code", "")
    error_msg = meta.get("error", "")

    # Error explícito en metadata
    if error_code in ERROR_STATES:
        return ("FALLO", error_msg or error_code, score)

    # Col E empieza con ( → error textual
    if col_e and isinstance(col_e, str) and col_e.startswith("("):
        detail = col_e.strip("()")
        return ("FALLO", detail, score)

    # Sin col E ni error → sin datos
    if not col_e or (isinstance(col_e, str) and col_e.strip() in ("", "-", "N/A")):
        # Ver si col D tiene digitaldata
        if col_d and isinstance(col_d, str) and "(no digitaldata)" in col_d:
            return ("SIN_DATOS", "Sin digitalData ni AA", score)
        return ("SIN_DATOS", "Sin datos de AA", score)

    # Tiene datos en col E → OK
    return ("OK", "", score)


def extract_pages_from_historial(path: str, market: str) -> list[dict]:
    """Extrae todas las páginas de un archivo historial, tomando el ÚLTIMO sheet de cada URL.
    
    Returns:
        Lista de dicts con datos de cada página (una por URL, la más reciente)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = [s for s in wb.sheetnames if s not in ("_control", "_vars") and not s.startswith("_")]
    # Ordenar sheets por fecha descendente (YYYY-MM-DD)
    sheets.sort(reverse=True)

    if not sheets:
        wb.close()
        return []

    pages: dict[str, dict] = {}  # URL → datos más recientes

    for sheet_name in sheets:
        ws = wb[sheet_name]
        headers = {}
        for c in range(1, ws.max_column + 1):
            hval = ws.cell(1, c).value
            if hval:
                headers[c] = str(hval).strip().lower()

        for row in range(2, ws.max_row + 1):
            url = ws.cell(row, 2).value
            if not url:
                continue
            url = str(url).strip()

            # Si ya tenemos esta URL de un sheet más reciente, saltar
            if url in pages:
                continue

            nombre = ws.cell(row, 1).value or ""
            col_d = ws.cell(row, 3).value  # digitaldata automatica
            col_e = ws.cell(row, 4).value  # AA analytics
            col_f = ws.cell(row, 5).value  # AA estructurado
            col_g = ws.cell(row, 6).value  # metadata

            meta = parse_meta_col(str(col_g) if col_g else "")
            estado, detalle, score = determine_status(
                str(col_e) if col_e else "",
                str(col_d) if col_d else "",
                meta,
            )

            # Detectar si tiene digitaldata
            dd_status = "OK"
            if not col_d or "(no digitaldata)" in str(col_d):
                dd_status = "NO"

            pages[url] = {
                "nombre": str(nombre).strip() if nombre else "(sin nombre)",
                "url": url,
                "mercado": market,
                "estado": estado,
                "detalle": detalle,
                "fecha": sheet_name,
                "score": score,
                "digitaldata": dd_status,
                "hoja": sheet_name,
            }

    wb.close()

    # También leer sin_aa.xlsx / con_aa.xlsx del mismo directorio si existen
    base_dir = os.path.dirname(path)
    for suffix in ["sin_aa.xlsx", "con_aa.xlsx"]:
        companion = os.path.join(base_dir, suffix)
        if not os.path.exists(companion) or companion == path:
            continue
        try:
            wb2 = openpyxl.load_workbook(companion, data_only=True)
            ws2 = wb2.active
            for row in range(2, ws2.max_row + 1):
                url = ws2.cell(row, 2).value
                if not url or str(url).strip() in pages:
                    continue
                url = str(url).strip()
                nombre = ws2.cell(row, 1).value or ""
                col_e = ws2.cell(row, 4).value
                col_g = ws2.cell(row, 6).value
                meta = parse_meta_col(str(col_g) if col_g else "")
                estado, detalle, score = determine_status(
                    str(col_e) if col_e else "", "", meta
                )
                pages[url] = {
                    "nombre": str(nombre).strip() if nombre else "(sin nombre)",
                    "url": url,
                    "mercado": market,
                    "estado": estado,
                    "detalle": detalle,
                    "fecha": os.path.basename(companion).replace(".xlsx", ""),
                    "score": score,
                    "digitaldata": "?",
                    "hoja": f"{suffix} ({sheet_name})",
                }
            wb2.close()
        except Exception:
            pass  # Si el archivo está corrupto, lo ignoramos

    return list(pages.values())


def build_report(all_pages: list[dict]) -> tuple[list[dict], list[dict]]:
    """Separa en working y failed, ordena.
    
    Returns:
        (failed_pages, all_sorted)
    """
    failed = [p for p in all_pages if p["estado"] != "OK"]
    failed.sort(key=lambda x: (x["mercado"], x["score"], x["nombre"]))
    all_sorted = sorted(all_pages, key=lambda x: (x["mercado"], x["estado"], x["nombre"]))
    return failed, all_sorted


def write_report(failed: list[dict], all_sorted: list[dict], output_path: str):
    """Escribe el Excel de reporte."""
    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ──
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_summary.cell(1, 1).value = "Reporte de Estado - Auditoria AA"
    ws_summary.cell(1, 1).font = Font(bold=True, size=14)
    ws_summary.cell(2, 1).value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary.cell(3, 1).value = ""

    total = len(all_sorted)
    ok_count = sum(1 for p in all_sorted if p["estado"] == "OK")
    fail_count = len(failed)
    nodata_count = sum(1 for p in all_sorted if p["estado"] == "SIN_DATOS")

    stats = [
        ("Total URLs auditadas", total),
        ("Funcionando (OK)", ok_count),
        ("Con fallos", fail_count),
        ("Sin datos", nodata_count),
        ("", ""),
        ("Tasa de exito", f"{ok_count / total * 100:.1f}%" if total > 0 else "N/A"),
    ]
    for i, (label, val) in enumerate(stats, start=5):
        ws_summary.cell(i, 1).value = label
        ws_summary.cell(i, 1).font = Font(bold=True)
        ws_summary.cell(i, 2).value = val

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # ── Hoja 2: Fallos ──
    _write_data_sheet(wb, "Fallos", failed, FAIL_FILL, "FALLO")

    # ── Hoja 3: Todos ──
    _write_data_sheet(wb, "Todos", all_sorted, None, None)

    wb.save(output_path)
    return output_path


def _write_data_sheet(wb, title: str, data: list[dict], highlight_fill, highlight_state):
    """Escribe una hoja con datos tabulares."""
    ws = wb.create_sheet(title=title)

    # Headers
    for c, h in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data
    for r, page in enumerate(data, start=2):
        vals = [
            page["nombre"], page["url"], page["mercado"],
            page["estado"], page["detalle"], page["fecha"],
            page["score"], page["digitaldata"], page["hoja"],
        ]
        estado = page["estado"]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(r, c)
            cell.value = val
            cell.border = THIN_BORDER
            # Color por estado
            if highlight_fill and estado == highlight_state:
                cell.fill = highlight_fill
            elif estado == "OK":
                cell.fill = OK_FILL
            elif estado == "SIN_DATOS":
                cell.fill = WARN_FILL

    # Ajustar anchos
    col_widths = [35, 55, 10, 12, 40, 18, 8, 12, 25]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    # Anchos para columnas más allá de la Z
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 25

    # Filtro automático
    ws.auto_filter.ref = f"A1:{chr(64 + len(OUTPUT_HEADERS))}{len(data) + 1}"


def generate_per_market_reports(pages: list[dict], base_dir: str, verbose: bool = False):
    """Genera un reporte individual por mercado en su carpeta."""
    by_market: dict[str, list[dict]] = {}
    for p in pages:
        m = p.get("mercado", "ROOT")
        if m not in by_market:
            by_market[m] = []
        by_market[m].append(p)

    for market, market_pages in by_market.items():
        if market == "ROOT":
            continue  # el global ya cubre ROOT
        market_dir = os.path.join(base_dir, market)
        os.makedirs(market_dir, exist_ok=True)
        output_path = os.path.join(market_dir, "reporte-auditoria.xlsx")
        failed, sorted_pages = build_report(market_pages)
        write_report(failed, sorted_pages, output_path)
        if verbose:
            print(f"    >> {market}/reporte-auditoria.xlsx ({len(market_pages)} URLs, {len(failed)} fallos)")


def main():
    parser = argparse.ArgumentParser(
        description="Reporte de URLs fallidas desde archivos de auditoría AA",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python audit_report.py                                # auto-detecta mercados
  python audit_report.py --dir PR                       # solo PR
  python audit_report.py --dir PR --dir MX              # PR y MX
  python audit_report.py --input PR/historial.xlsx      # archivo directo
  python audit_report.py --output fallos.xlsx           # nombre custom
        """,
    )
    parser.add_argument("--dir", action="append", dest="dirs",
                        help="Directorio(s) de mercado (ej: PR, MX)")
    parser.add_argument("--input", help="Ruta directa a un historial.xlsx")
    parser.add_argument("--output", default="reporte_auditoria.xlsx",
                        help="Archivo Excel de salida (default: reporte_auditoria.xlsx)")
    parser.add_argument("--verbose", action="store_true", help="Logging detallado")

    args = parser.parse_args()
    base = os.path.dirname(os.path.abspath(__file__))

    # Recolectar fuentes de datos
    sources = []
    if args.input:
        if os.path.exists(args.input):
            parent = os.path.basename(os.path.dirname(os.path.abspath(args.input)))
            market = parent.upper() if parent and parent != "." and not parent.startswith("_") else "ROOT"
            sources.append((args.input, market))
        else:
            print(f"[ERROR] No se encuentra: {args.input}")
            sys.exit(1)

    if args.input is None:
        sources = find_historial_files(base, args.dirs)

    if not sources:
        print("[!] No se encontraron archivos historial.xlsx en ningún directorio.")
        print("    Ejecutá primero extract_browser.py para generar auditorías.")
        print("    O usá --input para apuntar a un archivo específico.")
        sys.exit(1)

    # Extraer páginas
    all_pages = []
    for path, market in sources:
        if args.verbose:
            print(f"  Leyendo {path} (mercado: {market})...")
        try:
            pages = extract_pages_from_historial(path, market)
            all_pages.extend(pages)
            if args.verbose:
                print(f"    → {len(pages)} URLs encontradas")
        except Exception as e:
            print(f"  [WARN] Error leyendo {path}: {e}")

    if not all_pages:
        print("[!] No se encontraron datos de auditoría en los archivos.")
        sys.exit(1)

    # Generar reporte global
    failed, all_sorted = build_report(all_pages)
    output_path = write_report(failed, all_sorted, args.output)

    # Generar reportes por mercado (default: siempre)
    generate_per_market_reports(all_pages, base, args.verbose)

    # Mostrar resumen
    total = len(all_pages)
    ok_count = sum(1 for p in all_pages if p["estado"] == "OK")
    fail_count = len(failed)
    nodata_count = sum(1 for p in all_pages if p["estado"] == "SIN_DATOS")

    print(f"{'='*55}")
    print(f"  REPORTE GLOBAL:   {output_path}")
    markets = sorted(set(p.get("mercado", "ROOT") for p in all_pages if p.get("mercado") != "ROOT"))
    for m in markets:
        print(f"  {m}/reporte-auditoria.xlsx")
    print(f"{'='*55}")
    print(f"  Total URLs:       {total}")
    print(f"  Funcionando (OK): {ok_count}")
    print(f"  Con fallos:       {fail_count}")
    print(f"  Sin datos:        {nodata_count}")
    if total > 0:
        print(f"  Tasa de éxito:    {ok_count / total * 100:.1f}%")
    print(f"{'='*55}")

    if fail_count > 0:
        print(f"\n  PÁGINAS CON FALLOS:")
        for p in failed:
            print(f"    [{p['mercado']}] {p['nombre']}")
            print(f"           {p['url']}")
            print(f"           Estado: {p['detalle'] or p['estado']} (score: {p['score']})")
            print()

    if nodata_count > 0:
        print(f"\n  PÁGINAS SIN DATOS (pueden ser páginas sin AA implementado):")
        for p in all_sorted:
            if p["estado"] == "SIN_DATOS":
                print(f"    [{p['mercado']}] {p['nombre']}")
        print()


if __name__ == "__main__":
    main()

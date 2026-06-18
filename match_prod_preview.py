#!/usr/bin/env python3
"""
match_prod_preview.py — Compara digitalData entre lo prometido (Preview/Expected)
y lo entregado (Producción).

Modos de operación (automáticos):
  1. Preview existente → compara preview_url (AEM component) vs production_url (live)
  2. Sin preview → compara expected.json (estándar) vs production_url (live)

Uso:
  # Comparar expected vs producción (modo estándar)
  python match_prod_preview.py --production PR/historial.xlsx --mapping url-mapping.json --expected expected.json --market PR

  # Comparar preview vs producción (si existe PR/historial_preview.xlsx)
  python match_prod_preview.py --production PR/historial.xlsx --preview PR/historial_preview.xlsx --mapping url-mapping.json --market PR

Output: {market}/match-prod-vs-preview.{xlsx,md,html}
"""

import argparse
import io
import json
import os
import sys
from datetime import date

# Force UTF-8 for emoji-safe output
if sys.stdout.encoding and sys.stdout.encoding.upper() != "UTF-8":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Constantes ──
PARAMS_ORDER = ["pageName", "siteSection", "pageNameNoVehicle",
                "client", "site", "variantName", "pageType"]

MATCH_COLS = [
    "URL Preview (Prometido)",
    "URL Producción (Entregado)",
    "Nombre",
    "Parámetro",
    "Valor Prometido",
    "Valor Entregado",
    "Match",
]
MATCH_COL_WIDTHS = [45, 45, 25, 20, 50, 50, 12]

# ── Estilos ──
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_DATA = Font(name="Consolas", size=9)
FONT_PARAM = Font(name="Calibri", size=10)
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
FILL_OK = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_WARN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_MISS = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


def load_json(path: str) -> dict | list:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def find_digitaldata_in_historial(historial_path: str, url_fragment: str) -> dict:
    """Busca digitalData por URL en un historial."""
    if not os.path.exists(historial_path):
        return {}
    wb = openpyxl.load_workbook(historial_path, data_only=True)
    data_sheet = None
    for sn in wb.sheetnames:
        if sn not in ("_control", "_vars", "Sheet"):
            data_sheet = wb[sn]
            break
    if not data_sheet and "Sheet" in wb.sheetnames:
        data_sheet = wb["Sheet"]
    if not data_sheet:
        wb.close()
        return {}

    dd_col = 3
    for c in range(1, data_sheet.max_column + 1):
        hv = data_sheet.cell(1, c).value
        if hv and "digitaldata (automatica)" in str(hv).strip().lower():
            dd_col = c
            break

    for row in range(2, data_sheet.max_row + 1):
        url_cell = data_sheet.cell(row, 2).value
        if url_cell and url_fragment in str(url_cell):
            dd_raw = data_sheet.cell(row, dd_col).value
            wb.close()
            if dd_raw and isinstance(dd_raw, str):
                try:
                    return json.loads(dd_raw)
                except json.JSONDecodeError:
                    return {}
            return {}
    wb.close()
    return {}


def build_expected_page(page_key: str, cfg: dict) -> dict:
    """Construye dict digitalData.page según expected.json."""
    expected = {}
    for param in PARAMS_ORDER:
        param_cfg = cfg.get("params", {}).get(param, {})
        rule = param_cfg.get("rule", "pattern")
        if rule == "deprecated":
            continue
        if rule == "fixed":
            val = param_cfg.get("value", "")
        elif rule == "mirror":
            continue  # se resuelve con pageName
        elif rule in ("mapping", "required"):
            val = param_cfg.get("mapping", {}).get(page_key, "")
        else:  # pattern
            val = param_cfg.get("patterns", {}).get(page_key, "")
        if val:
            expected[param] = val
    return expected


def compare_params(promised: dict, delivered: dict,
                   param_list: list[str]) -> list[dict]:
    """Compara dos dicts digitalData.page parámetro por parámetro."""
    results = []
    for param in param_list:
        p_val = promised.get(param)
        d_val = delivered.get(param)
        p_str = str(p_val) if p_val is not None else ""
        d_str = str(d_val) if d_val is not None else ""

        if not p_str and not d_str:
            results.append({"param": param, "promised": "—", "delivered": "—",
                            "match": "⚠️", "match_text": "Sin datos en ambos"})
        elif not p_str:
            results.append({"param": param, "promised": "—", "delivered": d_str,
                            "match": "❌", "match_text": "Sin dato prometido"})
        elif not d_str:
            results.append({"param": param, "promised": p_str, "delivered": "—",
                            "match": "❌", "match_text": "Sin dato entregado"})
        elif p_str == d_str:
            results.append({"param": param, "promised": p_str, "delivered": d_str,
                            "match": "✅", "match_text": "Match"})
        else:
            results.append({"param": param, "promised": p_str, "delivered": d_str,
                            "match": "⚠️", "match_text": "Diferente"})
    return results


# ════════════════════════════════════════════
#  OUTPUT
# ════════════════════════════════════════════

def generate_report(mappings: list, all_results: list,
                    market: str, output_dir: str, mode: str):
    """Genera .xlsx, .md y .html con los resultados de la comparación."""
    today = date.today().isoformat()

    # ── Contar estadísticas ──
    total_params = sum(len(r["params"]) for r in all_results)
    match_ok = sum(1 for r in all_results for p in r["params"] if p["match"] == "✅")
    match_warn = sum(1 for r in all_results for p in r["params"] if p["match"] == "⚠️")
    match_fail = sum(1 for r in all_results for p in r["params"] if p["match"] == "❌")
    ok_pct = round(match_ok / total_params * 100, 1) if total_params else 0

    # ═══════════════════  .xlsx  ═══════════════════
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Match {market}"

    for col_idx, header in enumerate(MATCH_COLS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for col_idx, width in enumerate(MATCH_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    row_idx = 2
    for ent in all_results:
        for pr in ent["params"]:
            ws.cell(row_idx, 1, ent["preview_url"]).font = FONT_DATA
            ws.cell(row_idx, 1).alignment = ALIGN_WRAP
            ws.cell(row_idx, 1).border = THIN_BORDER

            ws.cell(row_idx, 2, ent["production_url"]).font = FONT_DATA
            ws.cell(row_idx, 2).alignment = ALIGN_WRAP
            ws.cell(row_idx, 2).border = THIN_BORDER

            ws.cell(row_idx, 3, ent["nombre"]).font = FONT_PARAM
            ws.cell(row_idx, 3).border = THIN_BORDER

            ws.cell(row_idx, 4, pr["param"]).font = FONT_PARAM
            ws.cell(row_idx, 4).border = THIN_BORDER

            ws.cell(row_idx, 5, pr["promised"]).font = FONT_DATA
            ws.cell(row_idx, 5).alignment = ALIGN_WRAP
            ws.cell(row_idx, 5).border = THIN_BORDER

            ws.cell(row_idx, 6, pr["delivered"]).font = FONT_DATA
            ws.cell(row_idx, 6).alignment = ALIGN_WRAP
            ws.cell(row_idx, 6).border = THIN_BORDER

            cell_m = ws.cell(row_idx, 7, f"{pr['match']} {pr['match_text']}")
            cell_m.font = FONT_PARAM
            cell_m.alignment = ALIGN_CENTER
            cell_m.border = THIN_BORDER

            fill = FILL_OK if pr["match"] == "✅" else (FILL_WARN if pr["match"] == "⚠️" else FILL_MISS)
            cell_m.fill = fill

            ws.row_dimensions[row_idx].height = 22
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{row_idx - 1}"

    xlsx_path = os.path.join(output_dir, "match-prod-vs-preview.xlsx")
    wb.save(xlsx_path)
    print(f"[OK] Match .xlsx: {xlsx_path}")

    # ═══════════════════  .md  ═══════════════════
    md_lines = [
        f"# Match Prod vs Preview — {market}",
        "",
        f"**Fecha**: {today}",
        f"**Modo**: {mode}",
        f"**URLs**: {len(all_results)}",
        f"**Parámetros**: {total_params}",
        "",
        "## Resumen",
        "",
        "| Estado | Cantidad | Porcentaje |",
        "|--------|----------|------------|",
        f"| ✅ Match | {match_ok} | {ok_pct}% |",
        f"| ⚠️ Diferente | {match_warn} | {round(match_warn/total_params*100, 1) if total_params else 0}% |",
        f"| ❌ Sin dato | {match_fail} | {round(match_fail/total_params*100, 1) if total_params else 0}% |",
        "",
    ]

    # Mostrar solo diferencias
    diff_entries = [(e, [p for p in e["params"] if p["match"] != "✅"])
                    for e in all_results]
    diff_entries = [(e, pp) for e, pp in diff_entries if pp]

    if diff_entries:
        md_lines.append("## Diferencias encontradas")
        md_lines.append("")
        for ent, params in diff_entries:
            md_lines.append(f"### {ent['nombre'] or ent['production_url']}")
            md_lines.append(f"- **Preview**: {ent['preview_url']}")
            md_lines.append(f"- **Producción**: {ent['production_url']}")
            md_lines.append("")
            md_lines.append("| Parámetro | Prometido | Entregado | Estado |")
            md_lines.append("|-----------|-----------|-----------|--------|")
            for p in params:
                md_lines.append(f"| `{p['param']}` | {p['promised']} | {p['delivered']} | {p['match']} {p['match_text']} |")
            md_lines.append("")
    else:
        md_lines.append("✅ **Todo en match** — no hay diferencias entre lo prometido y lo entregado.")
        md_lines.append("")

    md_path = os.path.join(output_dir, "match-prod-vs-preview.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))
    print(f"[OK] Match .md: {md_path}")

    # ═══════════════════  .html  ═══════════════════
    def _esc(s: str) -> str:
        return (s.replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))

    health = "green" if ok_pct >= 80 else ("orange" if ok_pct >= 50 else "red")

    rows_html = []
    for ent, params in diff_entries:
        rows_html.append(
            f'<tr class="url-sep"><td colspan="4">'
            f'<strong>{_esc(ent["nombre"] or ent["production_url"])}</strong><br>'
            f'<small>P: {_esc(ent["preview_url"])}</small><br>'
            f'<small>D: {_esc(ent["production_url"])}</small>'
            f'</td></tr>')
        for p in params:
            color = {"✅": "green", "⚠️": "orange", "❌": "red"}.get(p["match"], "black")
            rows_html.append(
                f'<tr><td><code>{p["param"]}</code></td>'
                f'<td>{_esc(p["promised"])}</td>'
                f'<td>{_esc(p["delivered"])}</td>'
                f'<td style="color:{color}">{p["match"]} {_esc(p["match_text"])}</td></tr>')

    table_html = ""
    if rows_html:
        table_html = (
            "<table><thead><tr>"
            "<th>Parámetro</th><th>Prometido</th><th>Entregado</th><th>Estado</th>"
            "</tr></thead><tbody>" + "\n".join(rows_html) + "</tbody></table>")
    else:
        table_html = '<p style="color:green;font-size:1.1em">✅ Todo en match</p>'

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Match Prod vs Preview — {_esc(market)}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:#f8f9fa; color:#333; padding:20px; }}
  h1 {{ font-size:1.5em; }}
  .meta {{ color:#666; font-size:0.9em; margin:8px 0 16px; }}
  .summary {{ display:flex; gap:12px; margin:16px 0; flex-wrap:wrap; }}
  .card {{ background:#fff; border-radius:8px; padding:16px 20px; flex:1; min-width:100px; box-shadow:0 1px 3px rgba(0,0,0,.08); text-align:center; }}
  .card .num {{ font-size:1.8em; font-weight:700; }}
  .card .label {{ font-size:0.8em; color:#666; }}
  .health {{ font-size:1.1em; margin:12px 0 20px; padding:10px 16px; border-radius:6px; color:#fff; font-weight:600; text-align:center; }}
  .health.green {{ background:#28a745; }}
  .health.orange {{ background:#fd7e14; }}
  .health.red {{ background:#dc3545; }}
  table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:8px; overflow:hidden; box-shadow:0 1px 3px rgba(0,0,0,.08); }}
  th {{ background:#4472C4; color:#fff; padding:10px 12px; text-align:left; font-size:0.85em; }}
  td {{ padding:8px 12px; border-bottom:1px solid #eee; font-size:0.85em; vertical-align:top; }}
  tr.url-sep td {{ background:#eef; font-size:0.95em; border-top:2px solid #4472C4; }}
  code {{ background:#f0f0f0; padding:1px 4px; border-radius:3px; font-size:0.9em; }}
</style>
</head>
<body>
<h1>Match Prod vs Preview — {_esc(market)}</h1>
<div class="meta">
  📅 {today} &nbsp;|&nbsp; Modo: {_esc(mode)} &nbsp;|&nbsp; {len(all_results)} URLs &nbsp;|&nbsp; {total_params} parámetros
</div>
<div class="health {health}">
  Match rate: {ok_pct}%
</div>
<div class="summary">
  <div class="card"><div class="num" style="color:#28a745">{match_ok}</div><div class="label">✅ Match</div></div>
  <div class="card"><div class="num" style="color:#fd7e14">{match_warn}</div><div class="label">⚠️ Diferente</div></div>
  <div class="card"><div class="num" style="color:#dc3545">{match_fail}</div><div class="label">❌ Sin dato</div></div>
</div>
{table_html}
</body>
</html>"""

    html_path = os.path.join(output_dir, "match-prod-vs-preview.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Match .html: {html_path}")

    # Resumen en consola
    print(f"\n  ✅ Match: {match_ok}  |  ⚠️ Diferente: {match_warn}  |  ❌ Sin dato: {match_fail}")
    print(f"  Match rate: {ok_pct}%")


# ════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Compara digitalData entre lo prometido (Preview/Expected) y lo entregado (Producción)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)

    parser.add_argument("--production", default="PR/historial.xlsx",
                        help="Historial de producción (entregado)")
    parser.add_argument("--preview", default=None,
                        help="Historial de preview (prometido). Si se omite, se busca "
                             "{market}/historial_preview.xlsx automáticamente. "
                             "Si no existe, se usa expected.json como fallback.")
    parser.add_argument("--mapping", default="url-mapping.json",
                        help="JSON mapeo preview→producción")
    parser.add_argument("--expected", default="expected.json",
                        help="JSON con valores esperados (fallback si no hay preview)")
    parser.add_argument("--market", default="PR",
                        help="Código de mercado")
    parser.add_argument("--output", default=None,
                        help="Directorio de salida (default: {market}/)")

    args = parser.parse_args()

    # Validar archivos
    if not os.path.exists(args.production):
        print(f"[ERR] Historial de producción no encontrado: {args.production}")
        sys.exit(1)
    if not os.path.exists(args.mapping):
        print(f"[ERR] Mapping no encontrado: {args.mapping}")
        sys.exit(1)

    mappings = load_json(args.mapping)
    market_dir = args.market.upper()
    output_dir = args.output or market_dir
    os.makedirs(output_dir, exist_ok=True)

    # Detectar preview historial
    preview_path = args.preview
    has_preview = False
    mode = ""

    if preview_path and os.path.exists(preview_path):
        has_preview = True
        mode = "preview-real"
        print(f"[INFO] Usando preview real: {preview_path}")
    else:
        auto_preview = os.path.join(market_dir, "historial_preview.xlsx")
        if os.path.exists(auto_preview):
            preview_path = auto_preview
            has_preview = True
            mode = "preview-real"
            print(f"[INFO] Preview detectado automáticamente: {auto_preview}")
        else:
            if not os.path.exists(args.expected):
                print(f"[ERR] No hay preview ({auto_preview}) ni expected ({args.expected})")
                sys.exit(1)
            mode = "expected-fallback"
            print(f"[INFO] Modo expected (sin preview). Usando: {args.expected}")

    # Comparar
    all_results = []
    expected_cfg = load_json(args.expected) if os.path.exists(args.expected) else {}

    for mapping in mappings:
        preview_url = mapping.get("preview_url", "")
        production_url = mapping.get("production_url", preview_url)
        page_key = mapping.get("page_key", "")
        nombre = mapping.get("nombre", "")

        if not page_key:
            continue

        if has_preview:
            # Modo preview real: comparar historiales
            promised_dd = find_digitaldata_in_historial(preview_path, preview_url)
            promised_page = promised_dd.get("page", {}) if promised_dd else {}
        else:
            # Modo expected: construir desde expected.json
            market_cfg = expected_cfg.get("markets", {}).get(args.market.upper(), {})
            promised_page = build_expected_page(page_key, market_cfg)

        delivered_dd = find_digitaldata_in_historial(args.production, production_url)
        delivered_page = delivered_dd.get("page", {}) if delivered_dd else {}

        params = compare_params(promised_page, delivered_page, PARAMS_ORDER)

        all_results.append({
            "preview_url": preview_url,
            "production_url": production_url,
            "page_key": page_key,
            "nombre": nombre,
            "params": params,
        })

        status = "✅" if all(p["match"] == "✅" for p in params) else "⚠️"
        print(f"  {status} {nombre or production_url[:50]}")

    # Generar reporte
    generate_report(mappings, all_results, args.market.upper(), output_dir, mode)


if __name__ == "__main__":
    main()
